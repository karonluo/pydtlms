from io import BytesIO

from openpyxl import load_workbook

from app.services.system_user_excel_service import build_system_user_import_template, parse_system_user_import_template


def test_build_system_user_import_template_omits_system_generated_columns() -> None:
    content = build_system_user_import_template()

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]

    assert "用户ID" not in headers
    assert "最近登录" not in headers
    assert headers[0] == "账号"


def test_parse_system_user_import_template_accepts_legacy_headers() -> None:
    legacy_headers = [
        "用户ID",
        "账号",
        "姓名",
        "岗位角色",
        "角色编码",
        "部门",
        "邮箱",
        "电话",
        "账号状态",
        "最近登录",
        "用户介绍",
        "登录密码",
    ]
    workbook = load_workbook(BytesIO(build_system_user_import_template()))
    worksheet = workbook.active
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(legacy_headers)
    worksheet.append([
        123,
        "zhangsan",
        "张三",
        "导师",
        "advisor",
        "人工智能中心",
        "zhangsan@example.com",
        "13800001111",
        "启用",
        "2026-05-25 10:00:00",
        "导师介绍",
        "Secret123!",
    ])
    stream = BytesIO()
    workbook.save(stream)

    rows = parse_system_user_import_template(stream.getvalue())

    assert len(rows) == 1
    assert rows[0]["username"] == "zhangsan"
    assert rows[0]["full_name"] == "张三"
    assert "id" not in rows[0]
    assert "last_login_at" not in rows[0]