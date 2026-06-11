from datetime import datetime
from io import BytesIO
from types import SimpleNamespace
from threading import RLock

from openpyxl import load_workbook
import pytest

from app.schemas.auth import UserProfileUpdate
from app.schemas.portal import PortalStudentRecord
from app.schemas.recruitment import (
    AdvisorScreeningBatchSubmitRequest,
    AdvisorScreeningSubmitItem,
    InitialScreeningConfirmationRequest,
    RecruitApplicationRecord,
    RecruitPlanUpsert,
)
from app.schemas.system import RoleUpsert, SystemUserUpsert
from app.schemas.student import CenterUpsert
from app.schemas.student import RegisteredPortalStudentRollbackStageRequest
from app.core.exceptions import DatabaseUnavailableError
from app.services.management_service import LazyRuntimeManagementStore, RuntimeManagementStore
from app.services.management_service_shared import CACHE_NULL_SENTINEL_KEY, PASSWORD_CONTEXT
from app.services.management_service_students import RuntimeManagementStoreStudentsMixin
from app.services.email_service import NotificationEmailService
from app.services.recruitment_excel_service import build_registered_portal_students_template
from app.services.runtime_seed_data import build_runtime_seed_state


class FakeRuntimeManagementStore:
    init_count = 0

    def __init__(self) -> None:
        type(self).init_count += 1
        self.label = "runtime-store"

    def ping(self) -> str:
        return self.label


class FakePostgresStateStore:
    def __init__(self) -> None:
        self.created_centers: list[dict] = []
        self.updated_centers: list[dict] = []
        self.created_students: list[dict] = []
        self.updated_students: list[dict] = []
        self.updated_portal_students: list[dict] = []
        self.deleted_center_ids: list[int] = []
        self.deleted_recruitment_plan_ids: list[int] = []
        self.portal_application_submissions: list[dict] = []
        self.updated_recruitment_applications: list[dict] = []
        self.background_assessments: list[dict] = []
        self.advisor_screening_batches: list[dict] = []
        self.initial_screening_confirmations: list[dict] = []
        self.initial_screening_notifications: list[dict] = []
        self.rollback_recruitment_applications: list[dict] = []
        self.saved_states: list[dict] = []
        self.synced_workflow_tasks: list[dict] = []
        self.synced_roles: list[dict] = []
        self.synced_operation_logs: list[dict] = []
        self.workflow_task_snapshot: dict | None = None
        self.recruitment_application_rows: list[dict] | None = None
        self.dict_options_map: dict[str, list[dict]] = {}
        self.system_user_rows: dict[int, dict] = {}
        self.role_rows: list[dict] | None = None
        self.system_user_state_rows: list[dict] | None = None
        self.profile_rows: dict[str, dict] | None = None
        self.audit_policy_rows: list[dict] | None = None
        self.integration_rows: list[dict] | None = None
        self.system_stats_snapshot: dict[str, int] | None = None
        self.missing_permission_codes_result: list[str] = []
        self.operation_log_max_id = 0

    def load_state(self) -> dict | None:
        return build_runtime_seed_state()

    def get_operation_log_max_id(self) -> int:
        return int(self.operation_log_max_id)

    def load_team_state(self):
        raise AttributeError("load_team_state")

    def list_dict_options(self, dict_type: str) -> list[dict]:
        return list(self.dict_options_map.get(dict_type, []))

    def load_role_state(self):
        return None if self.role_rows is None else [dict(item) for item in self.role_rows]

    def load_system_user_state(self):
        return None if self.system_user_state_rows is None else [dict(item) for item in self.system_user_state_rows]

    def load_user_profile_state(self):
        return None if self.profile_rows is None else {key: dict(value) for key, value in self.profile_rows.items()}

    def load_audit_policy_state(self):
        return None if self.audit_policy_rows is None else [dict(item) for item in self.audit_policy_rows]

    def load_integration_state(self):
        return None if self.integration_rows is None else [dict(item) for item in self.integration_rows]

    def get_system_stats_snapshot(self):
        if self.system_stats_snapshot is None:
            raise AttributeError("get_system_stats_snapshot")
        return dict(self.system_stats_snapshot)

    def get_role_by_id(self, role_id: int):
        if self.role_rows is None:
            return None
        for row in self.role_rows:
            if int(row.get("id") or 0) == int(role_id):
                return dict(row)
        return None

    def role_code_exists(self, role_code: str, *, exclude_role_id: int | None = None) -> bool:
        if self.role_rows is None:
            return False
        excluded = int(exclude_role_id) if exclude_role_id is not None else None
        return any(
            str(row.get("role_code") or "") == str(role_code) and (excluded is None or int(row.get("id") or 0) != excluded)
            for row in self.role_rows
        )

    def missing_permission_codes(self, permission_codes: list[str]) -> list[str]:
        configured_codes = {str(code) for code in permission_codes}
        return [code for code in self.missing_permission_codes_result if code in configured_codes]

    def get_audit_policy_by_id(self, policy_id: int):
        if self.audit_policy_rows is None:
            return None
        for row in self.audit_policy_rows:
            if int(row.get("id") or 0) == int(policy_id):
                return dict(row)
        return None

    def get_integration_by_id(self, integration_id: int):
        if self.integration_rows is None:
            return None
        for row in self.integration_rows:
            if int(row.get("id") or 0) == int(integration_id):
                return dict(row)
        return None

    def get_system_user_by_username(self, username: str):
        normalized_username = str(username)
        runtime_state = self.load_state() or {}
        source_rows = self.system_user_state_rows
        if source_rows is None:
            source_rows = [dict(item) for item in runtime_state.get("system_users", [])]
        for row in [*source_rows, *self.system_user_rows.values()]:
            if str(row.get("username") or "") != normalized_username:
                continue
            role_rows = self.role_rows
            if role_rows is None:
                role_rows = [dict(item) for item in runtime_state.get("roles", [])]
            role = next((item for item in role_rows if str(item.get("role_code") or "") == str(row.get("role_code") or "")), None)
            profile_rows = self.profile_rows
            if profile_rows is None:
                profile_rows = {key: dict(value) for key, value in (runtime_state.get("profiles", {}) or {}).items()}
            profile = dict(profile_rows.get(normalized_username, {}))
            return {
                **dict(row),
                "role_name": (role or {}).get("role_name") or row.get("role_code") or "",
                "department_name": profile.get("department_name") or row.get("department_name") or "",
                "introduction": profile.get("introduction") or row.get("introduction"),
                "email": profile.get("email") or row.get("email"),
                "phone_number": profile.get("phone_number") or row.get("phone_number"),
                "permissions": list((role or {}).get("permissions") or []),
            }
        return None

    def get_user_profile(self, username: str):
        normalized_username = str(username)
        if self.profile_rows is not None:
            profile = self.profile_rows.get(normalized_username)
            return None if profile is None else dict(profile)
        runtime_state = self.load_state() or {}
        profile = (runtime_state.get("profiles", {}) or {}).get(normalized_username)
        return None if profile is None else dict(profile)

    def get_portal_student_detail(self, student_id):
        del student_id
        return None

    def list_active_advisors(self):
        return [
            {
                "id": 11,
                "full_name": "刘亚",
                "advisor_no": "A011",
                "organization_name": "智能感知中心",
                "introduction": "长期从事具身智能与多模态学习研究。",
            },
            {
                "id": 12,
                "full_name": "王青",
                "advisor_no": "A012",
                "organization_name": "通用智能中心",
                "introduction": "关注基础模型、智能体与系统优化。",
            },
        ]

    def get_recruitment_application_detail(self, application_id):
        del application_id
        return None

    def list_recruitment_applications_page(self, *, keyword=None, plan_id=None, status=None, advisor_name=None, advisor_names=None, page=1, page_size=10):
        rows = [] if self.recruitment_application_rows is None else [dict(item) for item in self.recruitment_application_rows]
        if plan_id is not None:
            rows = [item for item in rows if int(item.get("plan_id") or 0) == int(plan_id)]
        if status:
            statuses = {segment.strip() for segment in str(status).split(",") if segment.strip()}
            rows = [item for item in rows if str(item.get("application_status") or "") in statuses]
        if advisor_name:
            normalized_advisor_name = str(advisor_name)
            rows = [
                item
                for item in rows
                if (
                    str(item.get("advisor_screening_round") or "") == "second_choice"
                    and str(item.get("second_choice") or "") == normalized_advisor_name
                )
                or (
                    str(item.get("advisor_screening_round") or "") != "second_choice"
                    and str(item.get("first_choice") or item.get("intended_advisor_name") or "") == normalized_advisor_name
                )
            ]
        normalized_advisor_names = {str(item).strip() for item in (advisor_names or []) if str(item).strip()}
        if normalized_advisor_names:
            rows = [
                item
                for item in rows
                if (
                    str(item.get("advisor_screening_round") or "") == "second_choice"
                    and str(item.get("second_choice") or "") in normalized_advisor_names
                )
                or (
                    str(item.get("advisor_screening_round") or "") != "second_choice"
                    and str(item.get("first_choice") or item.get("intended_advisor_name") or "") in normalized_advisor_names
                )
            ]
        if keyword:
            normalized_keyword = str(keyword)
            rows = [
                item
                for item in rows
                if normalized_keyword in str(item.get("business_key") or "")
                or normalized_keyword in str(item.get("student_name") or "")
            ]
        total = len(rows)
        start = max(int(page) - 1, 0) * int(page_size)
        end = start + int(page_size)
        return rows[start:end], total

    def save_state(self, state) -> None:
        self.saved_states.append(state)

    def sync_created_center(self, team_payload, operation_log=None, counters=None) -> None:
        self.created_centers.append({"team_payload": team_payload, "operation_log": operation_log, "counters": counters})

    def sync_updated_center(self, team_payload, affected_students, operation_log=None, counters=None) -> None:
        self.updated_centers.append(
            {
                "team_payload": team_payload,
                "affected_students": affected_students,
                "operation_log": operation_log,
                "counters": counters,
            }
        )

    def sync_deleted_center(self, center_id, operation_log=None, counters=None) -> None:
        self.deleted_center_ids.append(int(center_id))

    def delete_recruitment_plan(self, plan_id) -> None:
        self.deleted_recruitment_plan_ids.append(int(plan_id))

    def sync_created_student(self, student_payload, operation_log=None, counters=None) -> None:
        self.created_students.append({"student_payload": student_payload, "operation_log": operation_log, "counters": counters})

    def sync_updated_student(self, student_payload, operation_log=None, counters=None) -> None:
        self.updated_students.append({"student_payload": student_payload, "operation_log": operation_log, "counters": counters})

    def sync_portal_student(self, student_payload, operation_log=None, counters=None) -> None:
        self.updated_portal_students.append({"student_payload": student_payload, "operation_log": operation_log, "counters": counters})
        self.saved_states.append({"portal_student_payload": student_payload, "operation_log": operation_log, "counters": counters})

    def sync_portal_application_submission(
        self,
        portal_student_payload,
        application_payload,
        operation_log=None,
        *,
        workflow_task=None,
        counters=None,
    ) -> None:
        self.portal_application_submissions.append(
            {
                "portal_student_payload": portal_student_payload,
                "application_payload": application_payload,
                "operation_log": operation_log,
                "workflow_task": workflow_task,
                "counters": counters,
            }
        )

    def sync_recruitment_application_status(self, application_id, payload) -> None:
        self.updated_recruitment_applications.append({"application_id": int(application_id), "payload": dict(payload)})

    def list_background_assessments(self, application_id: int) -> list[dict]:
        return [dict(item) for item in self.background_assessments if int(item.get("application_id") or 0) == int(application_id)]

    def sync_recruitment_background_assessment(
        self,
        application_id,
        assessment_payload,
        application_payload,
        workflow_task_payload,
        operation_log=None,
        *,
        counters=None,
    ) -> None:
        normalized_application_id = int(application_id)
        incoming = dict(assessment_payload)
        self.background_assessments = [
            item
            for item in self.background_assessments
            if not (
                int(item.get("application_id") or 0) == normalized_application_id
                and str(item.get("evaluator_username") or "") == str(incoming.get("evaluator_username") or "")
            )
        ]
        self.background_assessments.append({**incoming, "application_id": normalized_application_id})
        self.updated_recruitment_applications.append({"application_id": normalized_application_id, "payload": dict(application_payload)})
        self.synced_workflow_tasks.append({"task_payload": dict(workflow_task_payload), "operation_log": operation_log, "counters": counters})

    def sync_recruitment_qualification_review(
        self,
        application_id,
        review_payload,
        application_payload,
        workflow_task_payload,
        operation_log=None,
        *,
        counters=None,
    ) -> None:
        normalized_application_id = int(application_id)
        self.updated_recruitment_applications.append({"application_id": normalized_application_id, "payload": dict(application_payload)})
        self.synced_workflow_tasks.append({"task_payload": dict(workflow_task_payload), "operation_log": operation_log, "counters": counters})

    def sync_advisor_screening_batch(
        self,
        batch_payload,
        application_payloads,
        workflow_task_payloads,
        operation_logs,
        *,
        counters=None,
    ):
        batch_id = len(self.advisor_screening_batches) + 1
        self.advisor_screening_batches.append(
            {
                "id": batch_id,
                "batch_payload": dict(batch_payload),
                "application_payloads": [dict(item) for item in application_payloads],
                "workflow_task_payloads": [dict(item) for item in workflow_task_payloads],
                "operation_logs": [None if item is None else dict(item) for item in operation_logs],
                "counters": counters,
            }
        )
        for application_payload in application_payloads:
            self.updated_recruitment_applications.append(
                {"application_id": int(application_payload["id"]), "payload": dict(application_payload)}
            )
        for workflow_task_payload, operation_log in zip(workflow_task_payloads, operation_logs):
            self.synced_workflow_tasks.append(
                {"task_payload": dict(workflow_task_payload), "operation_log": operation_log, "counters": counters}
            )
        return batch_id

    def sync_initial_screening_confirmation(
        self,
        confirmation_payload,
        application_payload,
        workflow_task_payload,
        notification_payloads,
        operation_log=None,
        *,
        counters=None,
    ) -> None:
        self.initial_screening_confirmations.append(
            {
                "confirmation_payload": dict(confirmation_payload),
                "application_payload": dict(application_payload),
                "workflow_task_payload": dict(workflow_task_payload),
                "operation_log": None if operation_log is None else dict(operation_log),
                "counters": counters,
            }
        )
        self.initial_screening_notifications.extend(dict(item) for item in notification_payloads)
        self.updated_recruitment_applications.append(
            {"application_id": int(application_payload["id"]), "payload": dict(application_payload)}
        )
        self.synced_workflow_tasks.append(
            {"task_payload": dict(workflow_task_payload), "operation_log": operation_log, "counters": counters}
        )

    def sync_workflow_task(self, task_payload, operation_log=None, *, counters=None) -> None:
        self.synced_workflow_tasks.append({"task_payload": dict(task_payload), "operation_log": operation_log, "counters": counters})

    def rollback_recruitment_application_stage(
        self,
        application_payload,
        workflow_task_payload,
        *,
        clear_background_assessments=False,
        clear_initial_screening_confirmation=False,
        operation_log=None,
        counters=None,
    ) -> None:
        application_id = int(application_payload["id"])
        self.rollback_recruitment_applications.append(
            {
                "application_payload": dict(application_payload),
                "workflow_task_payload": dict(workflow_task_payload),
                "clear_background_assessments": bool(clear_background_assessments),
                "clear_initial_screening_confirmation": bool(clear_initial_screening_confirmation),
                "operation_log": None if operation_log is None else dict(operation_log),
                "counters": counters,
            }
        )
        if clear_background_assessments:
            self.background_assessments = [item for item in self.background_assessments if int(item.get("application_id") or 0) != application_id]
        if clear_initial_screening_confirmation:
            self.initial_screening_confirmations = [
                item
                for item in self.initial_screening_confirmations
                if int(item.get("application_payload", {}).get("id") or item.get("confirmation_payload", {}).get("application_id") or 0) != application_id
            ]
            self.initial_screening_notifications = [
                item for item in self.initial_screening_notifications if int(item.get("application_id") or 0) != application_id
            ]
        self.updated_recruitment_applications.append({"application_id": application_id, "payload": dict(application_payload)})
        self.synced_workflow_tasks.append({"task_payload": dict(workflow_task_payload), "operation_log": operation_log, "counters": counters})

    def sync_role(self, role_payload, operation_log=None, *, counters=None) -> None:
        current_roles = [] if self.role_rows is None else [dict(item) for item in self.role_rows]
        incoming_role = dict(role_payload)
        updated = False
        for index, item in enumerate(current_roles):
            if int(item.get("id") or 0) == int(incoming_role.get("id") or 0):
                current_roles[index] = {**item, **incoming_role}
                updated = True
                break
        if not updated:
            current_roles.insert(0, incoming_role)
        self.role_rows = current_roles
        self.synced_roles.append({"role_payload": role_payload, "operation_log": operation_log, "counters": counters})

    def sync_audit_policy(self, policy_payload, operation_log=None, *, counters=None) -> None:
        self.saved_states.append({"policy_payload": policy_payload, "operation_log": operation_log, "counters": counters})

    def delete_audit_policy(self, policy_id) -> None:
        self.deleted_center_ids.append(int(policy_id))

    def sync_integration(self, integration_payload, operation_log=None, *, counters=None) -> None:
        self.saved_states.append({"integration_payload": integration_payload, "operation_log": operation_log, "counters": counters})

    def delete_integration(self, integration_id, integration_name=None) -> None:
        del integration_name
        self.deleted_center_ids.append(int(integration_id))

    def delete_role(self, role_id) -> None:
        self.deleted_center_ids.append(int(role_id))

    def sync_operation_log(self, operation_log, *, counters=None) -> None:
        self.synced_operation_logs.append({"operation_log": operation_log, "counters": counters})

    def _workflow_task_rows(self) -> list[dict]:
        rows_by_id: dict[int, dict] = {}
        if self.workflow_task_snapshot is not None:
            rows_by_id[int(self.workflow_task_snapshot["id"])] = dict(self.workflow_task_snapshot)
        for item in self.portal_application_submissions:
            workflow_task = item.get("workflow_task")
            if isinstance(workflow_task, dict) and workflow_task.get("id") is not None:
                rows_by_id[int(workflow_task["id"])] = dict(workflow_task)
        for item in self.synced_workflow_tasks:
            task_payload = item.get("task_payload")
            if isinstance(task_payload, dict) and task_payload.get("id") is not None:
                rows_by_id[int(task_payload["id"])] = dict(task_payload)
        return [rows_by_id[key] for key in sorted(rows_by_id.keys())]

    def list_workflow_tasks_page(self, *, status=None, module=None, keyword=None, page=1, page_size=10):
        filtered = self._workflow_task_rows()
        if status:
            filtered = [item for item in filtered if str(item.get("status") or "") == str(status)]
        if module:
            filtered = [item for item in filtered if str(item.get("business_module") or "") == str(module)]
        if keyword:
            normalized_keyword = str(keyword)
            filtered = [
                item
                for item in filtered
                if normalized_keyword in str(item.get("business_key") or "")
                or normalized_keyword in str(item.get("title") or "")
                or normalized_keyword in str(item.get("applicant_name") or "")
            ]
        total = len(filtered)
        start = max(int(page) - 1, 0) * int(page_size)
        end = start + int(page_size)
        return [dict(item) for item in filtered[start:end]], total

    def get_workflow_task_snapshot(self, task_id):
        normalized_task_id = int(task_id)
        for item in self._workflow_task_rows():
            if int(item.get("id") or 0) == normalized_task_id:
                return dict(item)
        return None

    def sync_user_profile(self, profile) -> None:
        self.saved_states.append({"profile": profile})

    def sync_system_user(self, user_payload, profile_payload=None, operation_log=None, *, counters=None) -> None:
        self.system_user_rows[int(user_payload["id"])] = {
            **self.system_user_rows.get(int(user_payload["id"]), {}),
            **dict(user_payload),
        }
        self.saved_states.append(
            {
                "user_payload": user_payload,
                "profile_payload": profile_payload,
                "operation_log": operation_log,
                "counters": counters,
            }
        )

    def get_system_user_by_id(self, user_id: int):
        row = self.system_user_rows.get(int(user_id))
        return dict(row) if row is not None else None

    def system_username_exists(self, username: str, *, exclude_user_id: int | None = None) -> bool:
        normalized_username = str(username)
        excluded = int(exclude_user_id) if exclude_user_id is not None else None
        return any(
            str(row.get("username") or "") == normalized_username and (excluded is None or int(row.get("id") or 0) != excluded)
            for row in self.system_user_rows.values()
        )

    def list_system_users_page(
        self,
        keyword: str | None = None,
        role_code: str | None = None,
        account_status: str | None = None,
        department_name: str | None = None,
        page: int = 1,
        page_size: int = 10,
    ) -> tuple[list[dict], int]:
        if not self.system_user_rows:
            raise AttributeError("list_system_users_page")

        runtime_state = self.load_state() or {}
        role_rows = self.role_rows
        if role_rows is None:
            role_rows = [dict(item) for item in runtime_state.get("roles", [])]
        profile_rows = self.profile_rows
        if profile_rows is None:
            profile_rows = {key: dict(value) for key, value in (runtime_state.get("profiles", {}) or {}).items()}

        items: list[dict] = []
        for row in sorted(self.system_user_rows.values(), key=lambda item: int(item.get("id") or 0), reverse=True):
            profile = dict(profile_rows.get(str(row.get("username") or ""), {}))
            role = next((item for item in role_rows if str(item.get("role_code") or "") == str(row.get("role_code") or "")), None)
            items.append(
                {
                    **dict(row),
                    "role_name": (role or {}).get("role_name") or row.get("role_code") or "",
                    "department_name": profile.get("department_name") or row.get("department_name") or "",
                    "introduction": profile.get("introduction") or row.get("introduction"),
                    "email": profile.get("email") or row.get("email"),
                    "phone_number": profile.get("phone_number") or row.get("phone_number"),
                    "account_status": row.get("account_status") or ("启用" if row.get("is_active", True) else "停用"),
                }
            )

        def _contains(source: str | None, target: str | None) -> bool:
            if not target:
                return True
            return str(target).lower() in str(source or "").lower()

        filtered = [
            item for item in items
            if _contains(item.get("username"), keyword)
            or _contains(item.get("full_name"), keyword)
            or _contains(item.get("department_name"), keyword)
        ] if keyword else list(items)
        if role_code:
            filtered = [item for item in filtered if str(item.get("role_code") or "") == str(role_code)]
        if account_status:
            filtered = [item for item in filtered if str(item.get("account_status") or "") == str(account_status)]
        if department_name:
            filtered = [item for item in filtered if _contains(item.get("department_name"), department_name)]

        total = len(filtered)
        offset = max(page - 1, 0) * page_size
        return filtered[offset: offset + page_size], total


class FakeNotificationEmailService:
    def __init__(self, log_delivery=None) -> None:
        self.log_delivery = log_delivery
        self.portal_registration_calls: list[dict] = []
        self.portal_password_reset_calls: list[dict] = []
        self.portal_admin_password_reset_calls: list[dict] = []
        self.recruitment_status_calls: list[dict] = []
        self.recruitment_stage_rollback_calls: list[dict] = []
        self.recruitment_material_review_rejection_calls: list[dict] = []
        self.custom_message_calls: list[dict] = []
        self.is_enabled = True
        self.workflow_notifications_is_enabled = True

    def enabled(self) -> bool:
        return self.is_enabled

    def workflow_notifications_enabled(self) -> bool:
        return self.workflow_notifications_is_enabled

    def send_portal_registration_success(self, full_name: str, email: str) -> None:
        self.portal_registration_calls.append({"full_name": full_name, "email": email})

    def send_portal_registration_success_async(self, full_name: str, email: str) -> None:
        self.send_portal_registration_success(full_name, email)

    def send_portal_password_reset_success(self, full_name: str, email: str, account: str) -> None:
        self.portal_password_reset_calls.append({"full_name": full_name, "email": email, "account": account})

    def send_portal_admin_password_reset(self, full_name: str, email: str, temporary_password: str) -> None:
        self.portal_admin_password_reset_calls.append({"full_name": full_name, "email": email, "temporary_password": temporary_password})

    def send_recruitment_status_update(self, **payload) -> None:
        self.recruitment_status_calls.append(payload)

    def send_recruitment_stage_rollback(self, **payload) -> None:
        self.recruitment_stage_rollback_calls.append(payload)

    def send_recruitment_material_review_rejection(self, **payload) -> None:
        self.recruitment_material_review_rejection_calls.append(payload)

    def send_message(self, *, to_email: str, subject: str, text_body: str) -> None:
        self.custom_message_calls.append({"to_email": to_email, "subject": subject, "text_body": text_body})

    def send_portal_registration_verification_code(self, email: str, verification_code: str) -> None:
        self.custom_message_calls.append(
            {
                "to_email": email,
                "subject": "学生门户邮箱验证码",
                "text_body": f"验证码：{verification_code}",
            }
        )

    def send_portal_login_verification_code(self, email: str, verification_code: str) -> None:
        self.custom_message_calls.append(
            {
                "to_email": email,
                "subject": "学生门户登录验证码",
                "text_body": f"验证码：{verification_code}",
            }
        )


def test_recruitment_reject_email_text_contains_review_comment() -> None:
    captured_messages: list[dict] = []

    class CaptureEmailService(NotificationEmailService):
        def send_message(self, **payload) -> None:  # type: ignore[override]
            captured_messages.append(payload)

    service = CaptureEmailService()
    service.send_recruitment_status_update(
        student_name="测试学生",
        email="student@example.com",
        business_key="SH202606020001",
        application_status="报名终止",
        plan_name="2026 秋季博士招生",
        review_comment="科研背景与项目方向匹配不足",
    )

    assert captured_messages
    text_body = captured_messages[0]["text_body"]
    assert "当前状态：报名终止" in text_body
    assert "原因：科研背景与项目方向匹配不足" in text_body


def test_recruitment_material_review_rejection_email_text_matches_template() -> None:
    captured_messages: list[dict] = []

    class CaptureEmailService(NotificationEmailService):
        def send_message(self, **payload) -> None:  # type: ignore[override]
            captured_messages.append(payload)

    service = CaptureEmailService()
    service.send_recruitment_material_review_rejection(
        student_name="测试学生",
        email="student@example.com",
        business_key="SH202705110001",
        plan_name="2027级联培博士生夏令营",
        review_comment="材料缺失",
    )

    assert captured_messages
    text_body = captured_messages[0]["text_body"]
    assert "测试学生同学，您好:" in text_body
    assert "业务编号: SH202705110001" in text_body
    assert "当前状态：驳回" in text_body
    assert "原因：材料缺失" in text_body
    assert "您的申请已被驳回，请于 24H 内登录系统补充或修改信息后重新提交。" in text_body
    assert "招生计划: 2027级联培博士生夏令营" in text_body


class FakeCacheClient:
    def __init__(self) -> None:
        self.values: dict[str, str] = {}
        self.ttl_values: dict[str, int] = {}

    def get(self, key: str):
        return self.values.get(key)

    def set(self, key: str, value, ex: int | None = None) -> bool:
        self.values[key] = str(value)
        if ex is not None:
            self.ttl_values[key] = int(ex)
        return True

    def exists(self, key: str) -> bool:
        return key in self.values

    def ttl(self, key: str) -> int:
        return int(self.ttl_values.get(key, -1))

    def delete(self, *keys: str) -> int:
        removed = 0
        for key in keys:
            if key in self.values:
                removed += 1
                self.values.pop(key, None)
            self.ttl_values.pop(key, None)
        return removed

    def incr(self, key: str) -> int:
        current = int(self.values.get(key, "0")) + 1
        self.values[key] = str(current)
        return current

    def expire(self, key: str, seconds: int) -> bool:
        self.ttl_values[key] = int(seconds)
        return True


def _build_center_payload(name: str) -> CenterUpsert:
    return CenterUpsert(
        center_name=name,
        director_name="刘亚",
        advisor_names=["刘亚", "袁野"],
        is_enabled=True,
        created_date="2026-04-21",
    )


def _build_plan_payload(name: str, description: str = "聚焦研究潜力与简章展示") -> RecruitPlanUpsert:
    return RecruitPlanUpsert(
        plan_name=name,
        academic_year="2026",
        semester="秋",
        brochure_image_url="/portal-brochures/test-plan.svg",
        plan_description=description,
    )


def _build_recruitment_application_payload():
    from app.schemas.recruitment import RecruitApplicationUpsert

    return RecruitApplicationUpsert(
        plan_id=1,
        student_name="邮件联调考生",
        graduation_school="江南大学",
        highest_degree="硕士",
        intended_field="智能制造",
        email="candidate-email@example.com",
        phone_number="13800009999",
        id_number="32000019990101123X",
        material_status="待审核",
        application_status="报名已提交",
    )


def _build_portal_application_payload(plan_id: int, team_name: str, advisor_name: str):
    from app.schemas.portal import PortalApplicationUpsert

    return PortalApplicationUpsert(
        plan_id=plan_id,
        profile={
            "full_name_pinyin": "zhangsan",
            "profile_photo_url": "/portal-attachments/uploads/student-7/profile_photo/photo-a.jpg",
            "id_card_collage_url": "/portal-attachments/uploads/student-7/id_card_collage/id-card-a.jpg",
            "gender": "男",
            "ethnic_group": "汉族",
            "political_status": "共青团员",
            "mailing_address": "上海市徐汇区某路 100 号",
            "emergency_contact_name": "张家长",
            "emergency_contact_phone": "13800002222",
        },
        graduation_school="江南大学",
        highest_degree="硕士",
        preferences=[
            {
                "preference_order": 1,
                "research_center_name": team_name,
                "team_id": 1,
                "advisor_name": advisor_name,
                "advisor_user_id": 11,
                "is_optional": False,
            }
        ],
        selected_team_id=1,
        selected_team_name=team_name,
        selected_advisor_user_id=11,
        selected_advisor_name=advisor_name,
        intended_field=team_name,
        source_channel="上海人工智能实验室官网",
        english_proficiencies=[
            {
                "exam_name": "CET-6",
                "score_text": "520",
                "certificate_attachment_url": "/portal-attachments/uploads/student-7/english_certificate/cet6-a.pdf",
            }
        ],
        family_members=[
            {"member_name": "张父", "relation_type": "父亲", "contact_phone": "13800001111"},
        ],
        personal_statement={
            "personal_statement_text": "个人陈述" + "丁" * 280,
            "growth_experience_text": "成长" + "甲" * 280,
            "program_application_reason_text": "申报" + "乙" * 280,
            "career_plan_text": "规划" + "丙" * 280,
            "resume_attachment_url": "/portal-attachments/uploads/student-7/resume/resume-a.pdf",
        },
        declaration={
            "has_read_declaration": True,
        },
        signed_agreement=True,
    )


def _advance_recruitment_application_to_advisor_screening(store, task_id: int) -> None:
    store.execute_workflow_action(
        task_id,
        "approve",
        "资料审核通过",
        principal={
            "username": "academy.admin",
            "full_name": "书院管理员",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )
    store.execute_workflow_action(
        task_id,
        "approve",
        "第一位书院管理员通过",
        principal={
            "username": "reviewer.a",
            "full_name": "评估人甲",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )
    store.execute_workflow_action(
        task_id,
        "approve",
        "第二位书院管理员通过",
        principal={
            "username": "reviewer.b",
            "full_name": "评估人乙",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )


def _sync_recruitment_application_rows_from_state(store, fake_postgres: FakePostgresStateStore) -> None:
    fake_postgres.recruitment_application_rows = [
        {
            "id": int(item.get("id") or 0),
            "plan_id": item.get("plan_id"),
            "business_key": item.get("business_key"),
            "student_name": item.get("student_name"),
            "graduation_school": item.get("graduation_school"),
            "highest_degree": item.get("highest_degree"),
            "intended_field": item.get("intended_field"),
            "material_status": item.get("material_status"),
            "application_status": item.get("application_status"),
            "first_choice": item.get("first_choice"),
            "second_choice": item.get("second_choice"),
            "intended_advisor_name": item.get("intended_advisor_name"),
            "advisor_screening_round": item.get("advisor_screening_round"),
            "advisor_screening_status": item.get("advisor_screening_status"),
        }
        for item in store.state["recruitment_applications"]
    ]


def test_lazy_store_does_not_create_instance_until_attribute_access(monkeypatch) -> None:
    FakeRuntimeManagementStore.init_count = 0
    monkeypatch.setattr("app.services.management_service.RuntimeManagementStore", FakeRuntimeManagementStore)

    store = LazyRuntimeManagementStore()

    assert FakeRuntimeManagementStore.init_count == 0
    assert store._instance is None


def test_lazy_store_reuses_single_runtime_instance(monkeypatch) -> None:
    FakeRuntimeManagementStore.init_count = 0
    monkeypatch.setattr("app.services.management_service.RuntimeManagementStore", FakeRuntimeManagementStore)

    store = LazyRuntimeManagementStore()

    assert store.ping() == "runtime-store"
    assert store.label == "runtime-store"
    assert FakeRuntimeManagementStore.init_count == 1


def test_create_center_uses_incremental_sync(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    record = store.create_center(_build_center_payload("增量新增研究中心"))

    assert record.center_name == "增量新增研究中心"
    assert len(fake_postgres.created_centers) == 1
    assert fake_postgres.created_centers[0]["team_payload"]["team_name"] == "增量新增研究中心"


def test_delete_center_uses_incremental_sync(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_center(_build_center_payload("增量删除研究中心"))
    store.delete_center(created.id)

    assert created.id in fake_postgres.deleted_center_ids


def test_update_center_falls_back_to_full_save_when_incremental_sync_fails(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()

    def raising_sync_updated_center(*args, **kwargs) -> None:
        raise RuntimeError("sync failed")

    fake_postgres.sync_updated_center = raising_sync_updated_center
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_center(_build_center_payload("增量回退研究中心"))
    save_calls: list[bool] = []
    monkeypatch.setattr(store, "_save", lambda: save_calls.append(True))

    updated = store.update_center(
        created.id,
        _build_center_payload("增量回退研究中心-已更新"),
    )

    assert updated.center_name == "增量回退研究中心-已更新"
    assert save_calls == [True]


def test_get_centers_filters_by_director_id(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()

    def list_centers_page(*, keyword=None, is_enabled=None, director_id=None, page=1, page_size=10):
        del keyword, is_enabled, page, page_size
        items = [
            {
                "id": 1,
                "center_name": "智能制造联合团队",
                "director_name": "林达华",
                "director_id": 1002,
                "advisor_names": ["林达华", "刘亚"],
                "advisor_ids": [1002, 1001],
                "advisor_relation_ids": [11, 12],
                "is_enabled": True,
                "created_date": "2026-04-21",
                "member_student_count": 2,
                "active_student_count": 2,
            },
            {
                "id": 2,
                "center_name": "前沿探索中心",
                "director_name": "刘亚",
                "director_id": 1001,
                "advisor_names": ["刘亚"],
                "advisor_ids": [1001],
                "advisor_relation_ids": [13],
                "is_enabled": True,
                "created_date": "2026-04-21",
                "member_student_count": 1,
                "active_student_count": 1,
            },
        ]
        filtered = [item for item in items if director_id is None or item["director_id"] == director_id]
        return filtered, len(filtered)

    fake_postgres.list_centers_page = list_centers_page
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_centers(director_id=1002)

    assert response.total == 1
    assert response.items[0].center_name == "智能制造联合团队"
    assert response.items[0].director_name == "林达华"
    assert response.items[0].director_id == 1002


def test_migrate_state_backfills_recruitment_plan_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.load_state = lambda: {
        "roles": [],
        "profiles": {},
        "teams": [],
        "students": [],
        "portal_students": [],
        "system_users": [],
        "audit_policies": [],
        "recruitment_plans": [{"id": 9, "plan_name": "历史计划", "academic_year": "2025", "semester": "春"}],
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    migrated_plan = store.state["recruitment_plans"][0]
    assert migrated_plan["plan_description"] is None
    assert migrated_plan["current_stage"] == "报名配置"
    assert migrated_plan["target_quota"] == 0
    assert migrated_plan["interview_group_count"] == 0
    assert migrated_plan["is_open"] is True
    assert migrated_plan["brochure_image_url"] is None


def test_get_recruitment_plans_matches_keyword_with_plan_description(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_recruitment_plans(keyword="产业问题导向", page=1, page_size=10)

    assert response.total == 1
    record = response.items[0]
    assert record.plan_name == "2026 工程博士专项"
    assert record.plan_description == "面向工程实践与产业问题导向申请人，强调场景落地与交叉协同。"
    assert set(record.model_dump().keys()) == {
        "id",
        "plan_name",
        "academic_term",
        "academic_year",
        "semester",
        "application_count",
        "brochure_image_url",
        "plan_description",
    }


def test_recruitment_application_record_allows_nullable_intended_field() -> None:
    record = RecruitApplicationRecord(
        id=27,
        plan_id=1,
        business_key="ZSLQSP202605010027",
        student_name="注册学生",
        graduation_school="江南大学",
        highest_degree="硕士",
        intended_field=None,
        material_status="材料齐全",
        application_status="报名已提交",
    )

    assert record.intended_field is None


def test_get_recruitment_portal_application_detail_maps_only_student_filled_sections(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.get_recruitment_application_detail = lambda application_id: {
        "id": int(application_id),
        "plan_id": 9,
        "business_key": "SH20270018",
        "candidate_no": "SH20270018",
        "student_name": "张三",
        "graduation_school": "江南大学",
        "highest_degree": "本科毕业",
        "phone_number": "13800001111",
        "email": "zhangsan@example.com",
        "id_number": "32000019990101123X",
        "application_status": "报名已提交",
        "material_status": "待审核",
        "advisor_screening_status": "submitted",
        "advisor_screening_round": "first_choice",
        "advisor_screening_submitted_at": "2026-05-30 01:38:19",
        "advisor_signature_base64": "data:image/png;base64,TEST_SIGNATURE",
        "first_choice_screening_score": 88.0,
        "next_stage_name": "入营面试",
        "reviewer_name": "admin",
        "final_score": 92.5,
        "applied_at": "2026-05-08 10:20:30",
        "profile": {
            "full_name_pinyin": "ZHANG SAN",
            "profile_photo_url": "/api/v1/portal/attachments/profile.jpg",
            "id_card_collage_url": "/api/v1/portal/attachments/id-card.jpg",
            "gender": "男",
            "birth_date": "1999-01-01",
            "ethnic_group": "汉族",
            "political_status": "共青团员",
            "mailing_address": "上海市徐汇区",
            "emergency_contact_name": "李四",
            "emergency_contact_phone": "13900002222",
        },
        "source_channel": "其他",
        "source_channel_other": "老师宣讲",
        "preferences": [{"preference_order": 1, "research_center_name": "具身智能", "advisor_name": "刘亚", "is_optional": False}],
        "education_experiences": [{"sort_order": 1, "education_stage": "本科毕业", "school_name": "江南大学"}],
        "practice_experiences": [{"organization_name": "某研究院"}],
        "english_proficiencies": [{"exam_name": "IELTS", "score_text": "7.0"}],
        "family_members": [{"member_name": "张父", "relation_type": "父亲"}],
        "achievement_records": [{"achievement_type": "获奖经历", "award_name": "挑战杯"}],
        "personal_statement": {
            "personal_statement_text": "个人陈述",
            "growth_experience_text": "成长经历",
            "program_application_reason_text": "申报理由",
            "career_plan_text": "职业规划",
            "resume_attachment_url": "/api/v1/portal/attachments/resume.pdf",
        },
        "declaration": {"has_read_declaration": True, "declaration_text": "本人承诺材料真实"},
        "material_list_attachment": "/api/v1/portal/attachments/materials.zip",
        "material_list_attachment_name": "materials.zip",
        "self_evaluation": "旧字段，不应出现在新 DTO",
        "research_status_analysis": "旧字段，不应出现在新 DTO",
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    detail = store.get_recruitment_portal_application_detail(18)

    assert detail.application_id == 18
    assert detail.source_channel == "其他"
    assert detail.source_channel_other == "老师宣讲"
    assert detail.profile is not None
    assert detail.profile.emergency_contact_name == "李四"
    assert detail.english_proficiencies[0].exam_name == "IELTS"
    assert detail.achievement_records[0].award_name == "挑战杯"
    assert detail.advisor_screening_status == "已通过"
    assert detail.advisor_screening_submitted_at == "2026-05-30 01:38:19"
    assert detail.advisor_signature_base64 == "data:image/png;base64,TEST_SIGNATURE"
    assert detail.personal_statement.supporting_material_attachment_url == "/api/v1/portal/attachments/materials.zip"
    assert not hasattr(detail, "self_evaluation")



def test_get_portal_profile_options_places_masses_after_league_member(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.dict_options_map["student_political_status"] = [
        {"label": "民盟盟员", "value": "民盟盟员"},
        {"label": "群众", "value": "群众"},
        {"label": "共青团员", "value": "共青团员"},
        {"label": "中共党员", "value": "中共党员"},
    ]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    values = [item.value for item in store.get_portal_profile_options().political_status_options]

    assert values.index("群众") == values.index("共青团员") + 1


def test_get_portal_profile_options_includes_advisor_introduction(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    options = store.get_portal_profile_options()

    assert options.advisor_options[0].full_name == "刘亚"
    assert options.advisor_options[0].introduction == "长期从事具身智能与多模态学习研究。"


def test_get_student_options_registered_portal_advisor_filter_includes_disabled_advisors(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.system_user_rows[901] = {
        "id": 901,
        "username": "liu.ya",
        "full_name": "刘亚",
        "role_code": "advisor",
        "department_name": "智能制造学院",
        "account_status": "启用",
    }
    fake_postgres.system_user_rows[902] = {
        "id": 902,
        "username": "yuan.ye",
        "full_name": "袁野",
        "role_code": "advisor",
        "department_name": "工业软件学院",
        "account_status": "停用",
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    monkeypatch.setattr(store, "get_centers", lambda *args, **kwargs: CenterListResponse(items=[], total=0, page=1, page_size=1000))
    options = store.get_student_options()

    values = [item.value for item in options.registered_portal_advisor_filter_options]
    labels = [item.label for item in options.registered_portal_advisor_filter_options]

    assert "刘亚" in values
    assert "袁野" in values
    assert any(label == "袁野（停用）" for label in labels)


def test_get_student_options_registered_portal_advisor_filter_scopes_advisor_self(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.system_user_rows[901] = {
        "id": 901,
        "username": "liu.ya",
        "full_name": "刘亚",
        "role_code": "advisor",
        "department_name": "智能制造学院",
        "account_status": "启用",
    }
    fake_postgres.system_user_rows[902] = {
        "id": 902,
        "username": "yuan.ye",
        "full_name": "袁野",
        "role_code": "advisor",
        "department_name": "工业软件学院",
        "account_status": "启用",
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    monkeypatch.setattr(store, "get_centers", lambda *args, **kwargs: CenterListResponse(items=[], total=0, page=1, page_size=1000))
    options = store.get_student_options(principal={"username": "liu.ya", "full_name": "刘亚", "roles": ["advisor"]})

    assert [item.value for item in options.registered_portal_advisor_filter_options] == ["刘亚"]


def test_runtime_store_initializes_operation_log_counter_from_postgres(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.operation_log_max_id = 365
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    assert store._counters["operation_logs"] == 365


def test_get_registered_portal_students_scopes_advisor_to_self(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    captured: dict[str, Any] = {}

    def fake_list_registered_portal_students_page(**kwargs):
        captured.update(kwargs)
        return [], 0

    fake_postgres.list_registered_portal_students_page = fake_list_registered_portal_students_page
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_registered_portal_students(
        advisor_names=["袁野", "刘亚"],
        page=1,
        page_size=20,
        principal={"username": "liu.ya", "full_name": "刘亚", "roles": ["advisor"]},
    )

    assert response.total == 0
    assert captured["advisor_names"] == ["刘亚"]


def test_get_student_options_includes_registered_portal_application_status_options(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.dict_options_map["recruitment_application_status"] = [
        {"label": "报名已提交", "value": "报名已提交"},
        {"label": "待初筛确认", "value": "待初筛确认"},
        {"label": "材料评分中", "value": "材料评分中"},
    ]
    fake_postgres.system_user_rows[901] = {
        "id": 901,
        "username": "liu.ya",
        "full_name": "刘亚",
        "role_code": "advisor",
        "department_name": "智能制造学院",
        "account_status": "启用",
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state["recruitment_applications"] = [
        {"id": 1, "application_status": "待背景评估"},
        {"id": 2, "application_status": "待导师初筛-第一志愿"},
    ]
    monkeypatch.setattr(store, "get_centers", lambda *args, **kwargs: CenterListResponse(items=[], total=0, page=1, page_size=1000))
    options = store.get_student_options()

    assert any(item.value == "待初筛确认" for item in options.registered_portal_application_status_options)
    assert any(item.value == "待背景评估" for item in options.registered_portal_application_status_options)
    assert any(item.value == "待导师初筛-第一志愿" for item in options.registered_portal_application_status_options)
    assert all(item.value != "材料评分中" for item in options.registered_portal_application_status_options)


def test_get_recruitment_options_advisor_filter_uses_only_advisor_role_users(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state["system_users"] = [
        {"id": 11, "username": "liu.ya", "full_name": "刘亚", "role_code": "advisor", "account_status": "启用"},
        {"id": 12, "username": "wang.qing", "full_name": "王青", "role_code": "advisor", "account_status": "启用"},
        {"id": 13, "username": "center.manager", "full_name": "具身智能中心", "role_code": "academy_admin", "account_status": "启用"},
    ]
    store.state["recruitment_applications"] = [
        {"id": 1, "first_choice": "刘亚", "second_choice": "具身智能中心", "intended_advisor_name": "研究中心A"},
        {"id": 2, "first_choice": "王青", "second_choice": "研究中心B", "intended_advisor_name": "王青"},
    ]

    options = store.get_recruitment_options()

    assert [item.value for item in options.advisor_options] == ["刘亚", "王青"]
    assert all(item.value not in {"具身智能中心", "研究中心A", "研究中心B"} for item in options.advisor_options)


def test_get_recruitment_applications_scopes_advisor_screening_to_current_advisor(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.recruitment_application_rows = [
        {
            "id": 201,
            "plan_id": 5,
            "business_key": "SH20270511",
            "student_name": "待初筛学生甲",
            "graduation_school": "上海大学",
            "highest_degree": "硕士",
            "intended_field": "具身智能",
            "material_status": "通过",
            "application_status": "待导师初筛-第一志愿",
            "first_choice": "陈恺",
            "second_choice": None,
            "intended_advisor_name": "陈恺",
            "advisor_screening_round": "first_choice",
            "advisor_screening_status": "pending",
        },
        {
            "id": 202,
            "plan_id": 5,
            "business_key": "SH20270512",
            "student_name": "待初筛学生乙",
            "graduation_school": "复旦大学",
            "highest_degree": "硕士",
            "intended_field": "通用智能",
            "material_status": "通过",
            "application_status": "待导师初筛-第一志愿",
            "first_choice": "其他导师",
            "second_choice": None,
            "intended_advisor_name": "其他导师",
            "advisor_screening_round": "first_choice",
            "advisor_screening_status": "pending",
        },
    ]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_recruitment_applications(
        status="待导师初筛,待导师初筛-第一志愿,待导师初筛-第二志愿",
        plan_id=5,
        principal={"username": "chenkai", "full_name": "陈恺", "roles": ["advisor"]},
        page=1,
        page_size=20,
    )

    assert response.total == 1
    assert [item.business_key for item in response.items] == ["SH20270511"]


def test_get_recruitment_applications_filters_multiple_advisors(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.recruitment_application_rows = [
        {
            "id": 301,
            "plan_id": 5,
            "business_key": "SH20270521",
            "student_name": "待初筛学生甲",
            "graduation_school": "上海大学",
            "highest_degree": "硕士",
            "intended_field": "具身智能",
            "material_status": "通过",
            "application_status": "待初筛确认",
            "first_choice": "陈恺",
            "second_choice": None,
            "intended_advisor_name": "陈恺",
            "advisor_screening_round": "first_choice",
            "advisor_screening_status": "submitted",
        },
        {
            "id": 302,
            "plan_id": 5,
            "business_key": "SH20270522",
            "student_name": "待初筛学生乙",
            "graduation_school": "复旦大学",
            "highest_degree": "硕士",
            "intended_field": "通用智能",
            "material_status": "通过",
            "application_status": "待初筛确认",
            "first_choice": "其他导师",
            "second_choice": "王青",
            "intended_advisor_name": "其他导师",
            "advisor_screening_round": "second_choice",
            "advisor_screening_status": "submitted",
        },
        {
            "id": 303,
            "plan_id": 5,
            "business_key": "SH20270523",
            "student_name": "待初筛学生丙",
            "graduation_school": "浙江大学",
            "highest_degree": "硕士",
            "intended_field": "机器人",
            "material_status": "通过",
            "application_status": "待初筛确认",
            "first_choice": "刘亚",
            "second_choice": None,
            "intended_advisor_name": "刘亚",
            "advisor_screening_round": "first_choice",
            "advisor_screening_status": "submitted",
        },
    ]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_recruitment_applications(
        status="待初筛确认",
        plan_id=5,
        advisor_names=["陈恺", "王青"],
        principal={"username": "admin", "full_name": "管理员", "roles": ["AILABMGT"]},
        page=1,
        page_size=20,
    )

    assert response.total == 2
    assert {item.business_key for item in response.items} == {"SH20270521", "SH20270522"}


def test_resolve_registered_portal_student_export_ids_scopes_explicit_ids_for_advisor(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state["portal_students"] = [
        {"id": 11, "selected_advisor_name": "刘亚"},
        {"id": 12, "selected_advisor_name": "袁野"},
    ]

    scoped_ids = store._resolve_registered_portal_student_export_ids(
        [11, 12],
        keyword=None,
        application_form_status=None,
        recruitment_application_status=None,
        show_all_background_assessed=False,
        advisor_names=None,
        principal={"username": "liu.ya", "full_name": "刘亚", "roles": ["advisor"]},
    )

    assert scoped_ids == [11]


def test_resolve_registered_portal_student_export_ids_passes_recruitment_status_filter(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    captured_kwargs: dict[str, object] = {}

    def fake_get_registered_portal_students(**kwargs):
        captured_kwargs.update(kwargs)
        return SimpleNamespace(items=[SimpleNamespace(id=21)])

    monkeypatch.setattr(store, "get_registered_portal_students", fake_get_registered_portal_students)

    resolved_ids = store._resolve_registered_portal_student_export_ids(
        [],
        keyword="待导出",
        application_form_status="已填写报名",
        recruitment_application_status="待初筛确认",
        show_all_background_assessed=False,
        advisor_names=["陈恺"],
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    assert resolved_ids == [21]
    assert captured_kwargs["recruitment_application_status"] == "待初筛确认"


def test_resolve_registered_portal_student_export_ids_reads_all_pages(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    call_history: list[tuple[int, int]] = []

    def fake_get_registered_portal_students(*, page: int, page_size: int, **kwargs):
        call_history.append((page, page_size))
        if page == 1:
            return SimpleNamespace(items=[SimpleNamespace(id=21)], total=2)
        if page == 2:
            return SimpleNamespace(items=[SimpleNamespace(id=22)], total=2)
        return SimpleNamespace(items=[], total=2)

    monkeypatch.setattr(store, "get_registered_portal_students", fake_get_registered_portal_students)

    resolved_ids = store._resolve_registered_portal_student_export_ids(
        [],
        keyword="待导出",
        application_form_status="已填写报名",
        recruitment_application_status="待初筛确认",
        show_all_background_assessed=False,
        advisor_names=["陈恺"],
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    assert resolved_ids == [21, 22]
    assert [page for page, _ in call_history] == [1, 2]
    assert len({page_size for _, page_size in call_history}) == 1


def test_create_recruitment_plan_keeps_internal_defaults_while_returning_trimmed_record(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_recruitment_plan(_build_plan_payload("2026 新计划", "只保留五个字段"))

    stored_plan = next(item for item in store.state["recruitment_plans"] if item["id"] == created.id)
    assert stored_plan["current_stage"] == "报名配置"
    assert stored_plan["target_quota"] == 0
    assert stored_plan["interview_group_count"] == 0
    assert stored_plan["is_open"] is True
    assert stored_plan["plan_description"] == "只保留五个字段"


def test_create_role_uses_relational_sync(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    record = store.create_role(
        RoleUpsert(
            role_code="quality_reviewer",
            role_name="质量审核员",
            scope_name="系统治理",
            permissions=[],
        )
    )

    assert record.role_code == "quality_reviewer"
    assert len(fake_postgres.synced_roles) == 1
    assert fake_postgres.synced_roles[0]["role_payload"]["role_name"] == "质量审核员"
    assert fake_postgres.synced_roles[0]["counters"]["roles"] == store._counters["roles"]


def test_get_workflow_task_detail_prefers_postgres_snapshot(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.workflow_task_snapshot = {
        "id": 88,
        "workflow_name": "科研报告审批",
        "business_module": "培养管理",
        "business_key": "SR-20260401-0001",
        "title": "科研报告审批任务",
        "applicant_name": "张三",
        "current_handler": "导师组",
        "current_node": "导师审核",
        "priority": "高",
        "status": "待处理",
        "created_at": "2026-04-01 09:00:00",
        "due_at": "2026-04-03 18:00:00",
        "flow_code": "scientific_report",
        "node_key": "advisor_review",
        "entity_id": 12,
        "candidate_groups": ["advisors"],
        "history": [
            {
                "operated_at": "2026-04-01 09:00:00",
                "operator_username": "admin",
                "operator_full_name": "管理员",
                "action": "start",
                "action_label": "发起流程",
                "from_node": "开始",
                "to_node": "导师审核",
                "result_status": "待处理",
                "comment": "已提交",
            }
        ],
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    detail = store.get_workflow_task_detail(88, {"username": "advisor", "full_name": "导师", "roles": ["advisor"]})

    assert detail["task"].id == 88
    assert detail["task"].workflow_name == "科研报告审批"
    assert detail["history"][0]["action_label"] == "发起流程"


def test_get_workflow_task_detail_allows_custom_role_with_matching_permissions(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.business_key)
    fake_postgres.workflow_task_snapshot = dict(next(item for item in store.state["workflow_tasks"] if item["id"] == task_id))
    detail = store.get_workflow_task_detail(
        task_id,
        {
            "username": "lixiaoyu",
            "full_name": "李晓宇",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert [item.label for item in detail["task"].available_actions] == ["审核通过", "审核不通过"]


def test_get_workflow_task_detail_hides_background_assessment_actions_after_same_reviewer_completed(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.workflow_task_snapshot = {
        "id": 88,
        "workflow_name": "招生申请审批",
        "business_module": "招生管理",
        "business_key": "SH20270388",
        "title": "报名审核任务",
        "applicant_name": "张三",
        "current_handler": "书院管理员",
        "current_node": "背景评估",
        "priority": "高",
        "status": "处理中",
        "created_at": "2026-05-27 09:00:00",
        "due_at": "2026-05-29 18:00:00",
        "flow_code": "recruitment_application",
        "node_key": "background_assessment",
        "entity_id": 34,
        "candidate_groups": ["AILABMGT"],
        "history": [],
    }
    fake_postgres.background_assessments = [
        {
            "application_id": 34,
            "evaluator_username": "lixiaoyu",
            "evaluator_name": "李晓宇",
            "evaluator_role_code": "AILABMGT",
            "assessment_result": "通过",
            "assessment_comment": "已评估",
            "assessed_at": "2026-05-27 10:00:00",
        }
    ]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    detail = store.get_workflow_task_detail(
        88,
        {
            "username": "lixiaoyu",
            "full_name": "李晓宇",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert detail["task"].current_node == "背景评估"
    assert detail["task"].available_actions == []


def test_get_workflow_task_detail_hides_background_assessment_actions_for_advisor(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.workflow_task_snapshot = {
        "id": 89,
        "workflow_name": "招生申请审批",
        "business_module": "招生管理",
        "business_key": "SH20270389",
        "title": "报名审核任务",
        "applicant_name": "李四",
        "current_handler": "书院管理员",
        "current_node": "背景评估",
        "priority": "高",
        "status": "处理中",
        "created_at": "2026-05-27 09:00:00",
        "due_at": "2026-05-29 18:00:00",
        "flow_code": "recruitment_application",
        "node_key": "background_assessment",
        "entity_id": 35,
        "candidate_groups": ["AILABMGT"],
        "history": [],
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    detail = store.get_workflow_task_detail(
        89,
        {
            "username": "liu.ya",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert detail["task"].current_node == "背景评估"
    assert detail["task"].available_actions == []


def test_get_workflow_tasks_allows_custom_role_with_matching_permissions(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.business_key)
    fake_postgres.workflow_task_snapshot = dict(next(item for item in store.state["workflow_tasks"] if item["id"] == task_id))

    response = store.get_workflow_tasks(
        module="招生管理",
        keyword=application.business_key,
        principal={
            "username": "lixiaoyu",
            "full_name": "李晓宇",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert len(response.items) == 1
    assert [item.label for item in response.items[0].available_actions] == ["审核通过", "审核不通过"]


def test_get_principal_context_expands_read_permission_from_workflow_write(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    role_record = store.create_role(
        RoleUpsert(
            role_code="academy_admin",
            role_name="书院管理员",
            scope_name="书院治理",
            permissions=["students:read", "workflow:write", "recruitment:write"],
        )
    )
    store.create_system_user(
        SystemUserUpsert(
            username="lixiaoyu",
            full_name="李晓宇",
            role_code=role_record.role_code,
            department_name="书院",
            email="lixiaoyu@example.com",
            phone_number="13800001111",
            account_status="启用",
            password="Secret123!",
        )
    )

    principal = store.get_principal_context("lixiaoyu")

    assert "workflow:write" in principal["permissions"]
    assert "workflow:read" in principal["permissions"]
    assert "recruitment:read" in principal["permissions"]


def test_runtime_seed_advisor_and_academy_manager_include_recruitment_permissions(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    advisor_principal = store.get_principal_context("liu.ya")
    academy_manager_principal = store.get_principal_context("qin.rao")

    assert advisor_principal["roles"] == ["advisor"]
    assert "recruitment:read" in advisor_principal["permissions"]
    assert "recruitment:write" in advisor_principal["permissions"]
    assert "recruitment_plan:read" in advisor_principal["permissions"]
    assert "recruitment_registered_students:read" in advisor_principal["permissions"]
    assert "recruitment_advisor_screening:read" in advisor_principal["permissions"]
    assert "workflow:write" in advisor_principal["permissions"]

    assert academy_manager_principal["roles"] == ["AILABMGT"]
    assert "recruitment:read" in academy_manager_principal["permissions"]
    assert "recruitment:write" in academy_manager_principal["permissions"]
    assert "recruitment_plan:read" in academy_manager_principal["permissions"]
    assert "recruitment_registered_students:read" in academy_manager_principal["permissions"]
    assert "recruitment_initial_screening_confirmation:read" in academy_manager_principal["permissions"]
    assert "workflow:write" in academy_manager_principal["permissions"]


def test_update_profile_without_bound_system_user_still_persists_operation_log(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state.setdefault("profiles", {})["portal-only"] = {
        "username": "portal-only",
        "full_name": "门户访客",
        "role_name": "门户用户",
        "department_name": "校外",
        "phone_number": None,
        "email": "portal@example.com",
        "theme_color": "#0f4cbd",
    }

    updated = store.update_profile(
        "portal-only",
        UserProfileUpdate(full_name="门户访客-更新", department_name="校友", phone_number="13800001111", email="portal@example.com", theme_color="#123456"),
    )

    assert updated.full_name == "门户访客-更新"
    assert len(fake_postgres.synced_operation_logs) == 1
    assert fake_postgres.synced_operation_logs[0]["operation_log"]["entity_id"] == "portal-only"


def test_delete_recruitment_plan_removes_plan_and_syncs_incrementally(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_recruitment_plan(_build_plan_payload("2026 待删除计划"))

    store.delete_recruitment_plan(created.id)

    assert all(item["id"] != created.id for item in store.state["recruitment_plans"])
    assert fake_postgres.deleted_recruitment_plan_ids == [created.id]


def test_delete_recruitment_plan_rejects_when_applications_exist(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    with pytest.raises(ValueError, match="当前招生计划下仍有报名申请，不能删除"):
        store.delete_recruitment_plan(1)


def test_get_public_recruitment_plans_exposes_description_as_summary(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    response = store.get_public_recruitment_plans()

    public_plan = response.items[0]
    assert public_plan.plan_name == "2026 智能制造联合培养"
    assert public_plan.summary == "聚焦智能制造方向联合培养，适合具备制造业、自动化、AI 背景的申请人。"
    assert set(public_plan.model_dump().keys()) == {
        "id",
        "plan_name",
        "academic_term",
        "brochure_image_url",
        "summary",
    }


def test_get_public_recruitment_plans_returns_latest_plan_first(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state["recruitment_plans"] = [
        {"id": 1, "plan_name": "2025 春季补录", "academic_year": "2025", "semester": "春", "plan_description": "春季补录说明"},
        {"id": 2, "plan_name": "2026 秋季博士招生", "academic_year": "2026", "semester": "秋", "plan_description": "秋季计划说明"},
        {"id": 3, "plan_name": "2026 春季计划", "academic_year": "2026", "semester": "春", "plan_description": "春季计划说明"},
    ]

    response = store.get_public_recruitment_plans()

    assert [item.plan_name for item in response.items] == ["2026 秋季博士招生", "2026 春季计划", "2025 春季补录"]


def test_register_portal_student_sends_email_notification(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800002222",
            email="new-student@example.com",
            full_name="新注册学生",
            id_number="320000199902022221",
            password="Secret123!",
        )
    )

    assert fake_mailer.portal_registration_calls == [{"full_name": "新注册学生", "email": "new-student@example.com"}]


def test_reset_portal_student_password_sends_email_notification(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalPasswordResetRequest, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800003333",
            email="password-reset@example.com",
            full_name="密码重置学生",
            id_number="320000199903033336",
            password="Secret123!",
        )
    )

    fake_mailer.portal_registration_calls.clear()
    store.reset_portal_student_password(
        PortalPasswordResetRequest(
            account="password-reset@example.com",
            id_number="320000199903033336",
            new_password="NewSecret123!",
        )
    )

    assert fake_mailer.portal_password_reset_calls == [
        {"full_name": "密码重置学生", "email": "password-reset@example.com", "account": "password-reset@example.com"}
    ]


def test_execute_workflow_action_sends_recruitment_pass_email(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    store.execute_workflow_action(
        task_id,
        "approve",
        None,
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    assert fake_mailer.recruitment_status_calls == []




def test_execute_workflow_action_background_assessment_rejects_advisor_even_with_write_permissions(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    store.execute_workflow_action(
        task_id,
        "approve",
        "资料审核通过",
        principal={
            "username": "academy.admin",
            "full_name": "书院管理员",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    with pytest.raises(PermissionError, match="当前账号无权执行该流程活动"):
        store.execute_workflow_action(
            task_id,
            "approve",
            "导师越权评估",
            principal={
                "username": "liu.ya",
                "full_name": "刘亚",
                "roles": ["advisor"],
                "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
            },
        )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert current_task["node_key"] == "background_assessment"
    assert current_application["application_status"] == "待背景评估"
    assert fake_postgres.list_background_assessments(application.id) == []
def test_execute_workflow_action_background_assessment_two_passes_advances_to_advisor_screening(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    store.execute_workflow_action(
        task_id,
        "approve",
        "资料审核通过",
        principal={
            "username": "academy.admin",
            "full_name": "书院管理员",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert current_task["node_key"] == "background_assessment"
    assert current_application["application_status"] == "待背景评估"

    store.execute_workflow_action(
        task_id,
        "approve",
        "第一位书院管理员通过",
        principal={
            "username": "reviewer.a",
            "full_name": "评估人甲",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    mid_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    mid_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert mid_task["node_key"] == "background_assessment"
    assert mid_application["application_status"] == "待背景评估"
    assert len(fake_postgres.list_background_assessments(application.id)) == 1

    store.execute_workflow_action(
        task_id,
        "approve",
        "第二位书院管理员通过",
        principal={
            "username": "reviewer.b",
            "full_name": "评估人乙",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    completed_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    completed_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert completed_task["node_key"] == "advisor_screening"
    assert completed_application["application_status"] == "待导师初筛-第一志愿"
    assert completed_application["advisor_screening_status"] == "pending"
    assert completed_application["advisor_screening_round"] == "first_choice"
    assert len(fake_postgres.list_background_assessments(application.id)) == 2
    assert fake_mailer.recruitment_status_calls[-1]["application_status"] == "待导师初筛-第一志愿"


def test_execute_workflow_action_background_assessment_two_rejects_terminates_application(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(update={"email": "background-reject@example.com"}),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    store.execute_workflow_action(
        task_id,
        "approve",
        "资料审核通过",
        principal={"username": "academy.admin", "full_name": "书院管理员", "roles": ["academy_admin"]},
    )

    store.execute_workflow_action(
        task_id,
        "reject",
        "第一位书院管理员不通过",
        principal={"username": "reviewer.a", "full_name": "评估人甲", "roles": ["academy_admin"]},
    )
    store.execute_workflow_action(
        task_id,
        "reject",
        "第二位书院管理员不通过",
        principal={"username": "reviewer.b", "full_name": "评估人乙", "roles": ["academy_admin"]},
    )

    completed_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    completed_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert completed_task["node_key"] is None
    assert completed_task["status"] == "已驳回"
    assert completed_application["application_status"] == "报名终止"
    assert fake_mailer.recruitment_status_calls[-1]["application_status"] == "报名终止"
    assert fake_mailer.recruitment_status_calls[-1]["review_comment"] == "第二位书院管理员不通过"


def test_submit_advisor_screening_batch_first_choice_pass_moves_to_confirmation(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "screening-pass@example.com",
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)

    response = store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[
                {
                    "application_id": application.id,
                    "advisor_score": 91,
                }
            ],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert response.batch_id == 1
    assert response.screening_round == "first_choice"
    assert response.submitted_count == 1
    assert current_task["node_key"] == "initial_screening_confirmation"
    assert current_application["application_status"] == "待初筛确认"
    assert current_application["initial_screening_status"] == "pending"
    assert current_application["first_choice_screening_score"] == pytest.approx(91)
    assert current_task["latest_comment"] == "导师初筛自动通过，分数 91.00，系统按 80 分阈值自动判定"
    assert len(fake_postgres.advisor_screening_batches) == 1
    assert fake_postgres.advisor_screening_batches[0]["batch_payload"]["advisor_name"] == "刘亚"
    assert fake_mailer.recruitment_status_calls[-1]["application_status"] == "待导师初筛-第一志愿"


def test_submit_advisor_screening_batch_rejects_duplicate_round_submission(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore.__new__(RuntimeManagementStore)
    store._lock = RLock()
    store._postgres_store = fake_postgres
    store._email_service = NotificationEmailService()
    store._loaded_state_from_postgres = False
    store._hydrated_state_from_postgres = False
    store._migrated_workflow_task_ids = set()
    store._counters = {"operation_logs": 0}
    store.state = {
        "counters": store._counters,
        "recruitment_applications": [
            {
                "id": 101,
                "plan_id": 1,
                "business_key": "BK-101",
                "student_name": "张三",
                "graduation_school": "江南大学",
                "highest_degree": "硕士",
                "material_status": "材料齐全",
                "application_status": "待初筛确认",
                "advisor_screening_status": "submitted",
                "advisor_screening_round": "first_choice",
                "portal_student_id": 7,
                "first_choice": "刘亚",
                "second_choice": "王青",
                "intended_advisor_name": "刘亚",
                "first_choice_screening_batch_id": 1,
                "first_choice_screening_submitted_at": "2026-06-09 10:00:00",
            }
        ],
        "workflow_tasks": [
            {
                "id": 201,
                "business_key": "BK-101",
                "node_key": "advisor_screening",
                "status": "处理中",
            }
        ],
    }
    store._find_required = lambda collection_name, entity_id: (
        0,
        store.state[collection_name][0],
    )
    store._workflow_task_index_by_business_key = lambda business_key: (
        0,
        store.state["workflow_tasks"][0],
    )

    with pytest.raises(ValueError, match="当前申请该轮导师初筛已提交，不能重复提交"):
        store.submit_advisor_screening_batch(
            AdvisorScreeningBatchSubmitRequest(
                signature_base64="data:image/png;base64,abc123",
                items=[
                    {
                        "application_id": 101,
                        "advisor_score": 88,
                    }
                ],
            ),
            principal={
                "username": "advisor.liu",
                "full_name": "刘亚",
                "roles": ["advisor"],
                "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
            },
        )


def test_background_assessment_pass_twice_persists_choice_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "screening-flow@example.com",
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    current_task = store.state["workflow_tasks"][0]
    current_task.update(
        {
            "flow_code": "recruitment_application",
            "node_key": "background_assessment",
            "status": "处理中",
            "current_node": "背景评估",
            "current_handler": "书院管理员",
        }
    )
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    current_application.update(
        {
            "application_status": "待背景评估",
            "advisor_screening_status": "pending",
            "advisor_screening_round": None,
            "first_choice_id": 11,
            "second_choice_id": 12,
        }
    )
    fake_postgres.updated_recruitment_applications.clear()

    store.execute_workflow_action(
        task_id,
        "approve",
        "第一位书院管理员通过",
        principal={"username": "reviewer.a", "full_name": "评估人甲", "roles": ["academy_admin"]},
    )
    store.execute_workflow_action(
        task_id,
        "approve",
        "第二位书院管理员通过",
        principal={"username": "reviewer.b", "full_name": "评估人乙", "roles": ["academy_admin"]},
    )

    synced_payload = fake_postgres.updated_recruitment_applications[-1]["payload"]
    assert synced_payload["advisor_screening_round"] == "first_choice"
    assert synced_payload["first_choice"] == "刘亚"
    assert synced_payload["second_choice"] == "王青"
    assert synced_payload["first_choice_id"] == 11
    assert synced_payload["second_choice_id"] == 12


def test_submit_advisor_screening_batch_allows_missing_signature(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore.__new__(RuntimeManagementStore)
    store._lock = RLock()
    store._postgres_store = fake_postgres
    store._email_service = fake_mailer
    store._loaded_state_from_postgres = False
    store._hydrated_state_from_postgres = False
    store._migrated_workflow_task_ids = set()
    store._counters = {"operation_logs": 0}
    store.state = {
        "counters": store._counters,
        "recruitment_applications": [
            {
                "id": 101,
                "plan_id": 1,
                "business_key": "BK-101",
                "student_name": "张三",
                "graduation_school": "江南大学",
                "highest_degree": "硕士",
                "material_status": "材料齐全",
                "application_status": "待导师初筛-第一志愿",
                "advisor_screening_status": "pending",
                "advisor_screening_round": "first_choice",
                "portal_student_id": 7,
                "first_choice": "刘亚",
                "second_choice": "王青",
                "intended_advisor_name": "刘亚",
            }
        ],
        "workflow_tasks": [
            {
                "id": 201,
                "business_key": "BK-101",
                "node_key": "advisor_screening",
                "status": "处理中",
            }
        ],
    }
    store._find_required = lambda collection_name, entity_id: (
        0,
        store.state[collection_name][0],
    )
    store._workflow_task_index_by_business_key = lambda business_key: (
        0,
        store.state["workflow_tasks"][0],
    )
    store._resolve_advisor_screening_round = lambda entity: "first_choice"
    store._principal_matches_screening_advisor = lambda entity, screening_round, principal_summary: True
    store._build_workflow_transition_record = lambda task, updated_entity, **kwargs: {
        **task,
        "node_key": kwargs.get("next_node"),
        "status": kwargs.get("task_status"),
        "latest_comment": kwargs.get("comment"),
    }
    store._record_operation = lambda *args, **kwargs: {
        "category": args[0],
        "object_type": args[1],
        "target_key": args[2],
        "action_label": args[3],
        "description": args[4],
        "operator_username": kwargs.get("operator_username"),
        "target_name": kwargs.get("target_name"),
    }

    response = store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64=None,
            items=[
                {
                    "application_id": 101,
                    "advisor_score": 88,
                }
            ],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert response.submitted_count == 1
    assert response.screening_round == "first_choice"
    assert fake_postgres.advisor_screening_batches[0]["batch_payload"]["signature_base64"] == ""


def test_portal_workflow_progress_summary_uses_scoring_based_screening_copy(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    summary = store._build_portal_workflow_progress_summary("2026-05-28 09:00:00", "待导师初筛-第一志愿")

    current_stage = next(item for item in summary.stages if item.key == "initial_screening")
    assert current_stage.description == "等待初筛完成"


def test_portal_workflow_progress_summary_marks_background_and_later_terminated_stages(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    summary = store._build_portal_workflow_progress_summary(
        "2026-05-28 09:00:00",
        "报名终止",
        background_assessments=[
            {"assessment_result": "通过"},
            {"assessment_result": "不通过"},
        ],
    )

    terminated_stage_keys = ["background_review", "initial_screening", "camp_interview", "result_publish", "pre_admission"]
    terminated_stages = {item.key: item for item in summary.stages if item.key in terminated_stage_keys}

    assert summary.result_label == "终止"
    assert set(terminated_stages) == set(terminated_stage_keys)
    assert all(item.status == "terminated" for item in terminated_stages.values())
    assert all(item.description == "终止" for item in terminated_stages.values())


def test_portal_workflow_progress_summary_keeps_background_review_completed_when_initial_screening_failed(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    summary = store._build_portal_workflow_progress_summary(
        "2026-05-28 09:00:00",
        "报名终止",
        advisor_screening_status="rejected",
        initial_screening_result="rejected",
        background_assessments=[
            {"assessment_result": "通过"},
            {"assessment_result": "通过"},
        ],
    )

    stage_by_key = {item.key: item for item in summary.stages}

    assert stage_by_key["background_review"].status == "completed"
    assert stage_by_key["initial_screening"].status == "terminated"
    assert stage_by_key["initial_screening"].description == "终止"
    assert stage_by_key["camp_interview"].status == "terminated"
    assert stage_by_key["result_publish"].status == "terminated"
    assert stage_by_key["pre_admission"].status == "terminated"


def test_confirm_initial_screening_pass_moves_to_camp_interview(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "confirmation-pass@example.com",
                "portal_student_id": 88,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[
                {
                    "application_id": application.id,
                    "advisor_score": 95,
                }
            ],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    result = store.confirm_initial_screening(
        application.id,
        InitialScreeningConfirmationRequest(result="passed", comment="进入入营面试"),
        principal={
            "username": "academy.confirm",
            "full_name": "初筛确认人",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert result.application_status == "入营面试"
    assert current_task["node_key"] == "camp_interview"
    assert current_application["initial_screening_result"] == "passed"
    assert current_application["next_stage_name"] == "入营面试"
    assert len(fake_postgres.initial_screening_confirmations) == 1
    assert len(fake_postgres.initial_screening_notifications) == 2
    assert fake_mailer.recruitment_status_calls[-1]["application_status"] == "入营面试"


def test_confirm_initial_screening_reject_email_contains_review_comment(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "confirmation-reject@example.com",
                "portal_student_id": 88,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[{"application_id": application.id, "advisor_score": 95}],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    store.confirm_initial_screening(
        application.id,
        InitialScreeningConfirmationRequest(result="rejected", comment="科研背景与项目方向匹配不足"),
        principal={
            "username": "academy.confirm",
            "full_name": "初筛确认人",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    assert fake_mailer.recruitment_status_calls[-1]["application_status"] == "报名终止"
    assert fake_mailer.recruitment_status_calls[-1]["review_comment"] == "科研背景与项目方向匹配不足"


def test_rollback_registered_portal_student_stage_moves_confirmation_back_to_advisor_screening(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    store.state.setdefault("portal_students", []).insert(
        0,
        {
            "id": 88,
            "full_name": "回退学生",
            "email": "rollback-stage@example.com",
            "account_status": "启用",
            "phone_number": "13800008888",
        },
    )
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "rollback-stage@example.com",
                "portal_student_id": 88,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[{"application_id": application.id, "advisor_score": 92}],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    response = store.rollback_registered_portal_student_stage(
        88,
        RegisteredPortalStudentRollbackStageRequest(target_stage="advisor_screening_first", comment="补充复核"),
        principal={"username": "admin", "full_name": "平台管理员", "roles": ["platform_admin"]},
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert response.message == "已退回至导师初筛-第一志愿"
    assert response.email_sent is True
    assert current_task["node_key"] == "advisor_screening"
    assert current_application["application_status"] == "待导师初筛-第一志愿"
    assert current_application["advisor_screening_status"] == "pending"
    assert current_application["first_choice_screening_score"] is None
    assert current_application["initial_screening_status"] is None
    assert fake_postgres.rollback_recruitment_applications[-1]["clear_initial_screening_confirmation"] is True
    assert fake_mailer.recruitment_stage_rollback_calls[-1]["target_stage_label"] == "导师初筛-第一志愿"


def test_rollback_registered_portal_student_stage_moves_camp_interview_back_to_qualification_review(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    store.state.setdefault("portal_students", []).insert(
        0,
        {
            "id": 89,
            "full_name": "退回资格学生",
            "email": "rollback-qualification@example.com",
            "account_status": "启用",
            "phone_number": "13800008889",
        },
    )
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "rollback-qualification@example.com",
                "portal_student_id": 89,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[AdvisorScreeningSubmitItem(application_id=application.id, advisor_score=94)],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )
    store.confirm_initial_screening(
        application.id,
        InitialScreeningConfirmationRequest(result="passed", comment="进入入营面试"),
        principal={
            "username": "academy.confirm",
            "full_name": "初筛确认人",
            "roles": ["academy_admin"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    response = store.rollback_registered_portal_student_stage(
        89,
        RegisteredPortalStudentRollbackStageRequest(target_stage="qualification_review", comment="重新核验材料"),
        principal={"username": "admin", "full_name": "平台管理员", "roles": ["platform_admin"]},
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    rollback_payload = fake_postgres.rollback_recruitment_applications[-1]
    assert response.message == "已退回至资格审核"
    assert current_task["node_key"] == "qualification_review"
    assert current_task["latest_comment"] == "重新核验材料"
    assert current_application["application_status"] == "报名已提交"
    assert current_application["advisor_screening_status"] is None
    assert current_application["advisor_screening_round"] is None
    assert current_application["first_choice_screening_score"] is None
    assert current_application["second_choice_screening_score"] is None
    assert current_application["initial_screening_status"] is None
    assert current_application["initial_screening_result"] is None
    assert rollback_payload["clear_background_assessments"] is True
    assert rollback_payload["clear_initial_screening_confirmation"] is True
    assert fake_postgres.background_assessments == []
    assert fake_postgres.initial_screening_confirmations == []
    assert fake_postgres.initial_screening_notifications == []
    assert fake_mailer.recruitment_stage_rollback_calls[-1]["target_stage_label"] == "资格审核"


def test_rollback_registered_portal_student_stage_rejects_forward_stage(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    store.state.setdefault("portal_students", []).insert(
        0,
        {
            "id": 90,
            "full_name": "禁止前进学生",
            "email": "rollback-forward@example.com",
            "account_status": "启用",
            "phone_number": "13800008890",
        },
    )
    store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={"email": "rollback-forward@example.com", "portal_student_id": 90}
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    with pytest.raises(ValueError, match="仅可退回到当前环节或更前的环节"):
        store.rollback_registered_portal_student_stage(
            90,
            RegisteredPortalStudentRollbackStageRequest(target_stage="background_assessment"),
            principal={"username": "admin", "full_name": "平台管理员", "roles": ["platform_admin"]},
        )


def test_rollback_registered_portal_student_stage_does_not_fallback_to_full_save(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    store.state.setdefault("portal_students", []).insert(
        0,
        {
            "id": 91,
            "full_name": "持久化失败学生",
            "email": "rollback-persist-fail@example.com",
            "account_status": "启用",
            "phone_number": "13800008891",
        },
    )
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "rollback-persist-fail@example.com",
                "portal_student_id": 91,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)

    def fail_rollback_persist(*args, **kwargs):
        raise RuntimeError("database unavailable")

    def fail_full_save() -> None:
        raise AssertionError("full save should not be used as rollback fallback")

    fake_postgres.rollback_recruitment_application_stage = fail_rollback_persist
    monkeypatch.setattr(store, "_save", fail_full_save)

    with pytest.raises(RuntimeError, match="退回环节持久化失败"):
        store.rollback_registered_portal_student_stage(
            91,
            RegisteredPortalStudentRollbackStageRequest(target_stage="qualification_review", comment="重新核验材料"),
            principal={"username": "admin", "full_name": "平台管理员", "roles": ["platform_admin"]},
        )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert current_task["node_key"] == "advisor_screening"
    assert current_application["application_status"] == "待导师初筛-第一志愿"
    assert current_application["advisor_screening_status"] == "pending"


def test_rollback_registered_portal_student_stage_rejects_non_platform_admin(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    with pytest.raises(PermissionError, match="仅平台管理员或书院管理员可退回注册学生报名环节"):
        store.rollback_registered_portal_student_stage(
            88,
            RegisteredPortalStudentRollbackStageRequest(target_stage="background_assessment"),
            principal={
                "username": "academy.admin",
                "full_name": "书院管理员",
                "roles": ["advisor"],
                "permissions": [],
            },
        )


def test_rollback_registered_portal_student_stage_allows_academy_admin(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    store.state.setdefault("portal_students", []).insert(
        0,
        {
            "id": 92,
            "full_name": "书院退回学生",
            "email": "academy-rollback@example.com",
            "account_status": "启用",
            "phone_number": "13800008892",
        },
    )
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "academy-rollback@example.com",
                "portal_student_id": 92,
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    current_task = store.state["workflow_tasks"][0]
    current_task["node_key"] = "background_assessment"
    current_task["current_node"] = "背景评估"
    current_task["current_handler"] = "书院管理员"
    current_task["status"] = "处理中"
    current_task["business_key"] = str(application.business_key)
    application.application_status = "待背景评估"

    response = store.rollback_registered_portal_student_stage(
        92,
        RegisteredPortalStudentRollbackStageRequest(target_stage="qualification_review", comment="补充核验"),
        principal={
            "username": "academy.admin",
            "full_name": "书院管理员",
            "roles": ["academy_admin"],
            "permissions": ["recruitment_registered_students:write"],
        },
    )

    current_task = store.state["workflow_tasks"][0]
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert response.message == "已退回至资格审核"
    assert current_task["node_key"] == "qualification_review"
    assert current_application["application_status"] == "报名已提交"


def test_submit_advisor_screening_batch_below_threshold_moves_to_second_choice(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "screening-threshold@example.com",
                "intended_advisor_name": "刘亚",
                "first_choice": "刘亚",
                "second_choice": "王青",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = store.state["workflow_tasks"][0]["id"]
    _advance_recruitment_application_to_advisor_screening(store, task_id)

    response = store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,abc123",
            items=[
                {
                    "application_id": application.id,
                    "advisor_score": 79,
                }
            ],
        ),
        principal={
            "username": "advisor.liu",
            "full_name": "刘亚",
            "roles": ["advisor"],
            "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
        },
    )

    current_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)
    current_application = next(item for item in store.state["recruitment_applications"] if item["id"] == application.id)
    assert response.screening_round == "first_choice"
    assert current_task["node_key"] == "advisor_screening"
    assert current_application["application_status"] == "待导师初筛-第二志愿"
    assert current_application["advisor_screening_status"] == "pending"
    assert current_application["advisor_screening_round"] == "second_choice"
    assert current_application["first_choice_screening_score"] == pytest.approx(79)
    assert fake_postgres.advisor_screening_batches[0]["application_payloads"][0]["is_passed"] is False


def test_end_to_end_screening_flow_passes_across_student_advisor_and_academy_manager(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]
    advisor_principal = store.get_principal_context("liu.ya")
    academy_manager_principal = store.get_principal_context("qin.rao")

    registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007701",
            email="screening-pass-e2e@example.com",
            full_name="初筛通过联调学生",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    submission = store.submit_portal_application(
        registration.student.id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == submission.application_business_key)
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    _sync_recruitment_application_rows_from_state(store, fake_postgres)

    advisor_list = store.get_recruitment_applications(
        status="待导师初筛,待导师初筛-第一志愿,待导师初筛-第二志愿",
        plan_id=available_plan.id,
        principal=advisor_principal,
        page=1,
        page_size=20,
    )

    assert advisor_list.total == 1
    assert advisor_list.items[0].business_key == submission.application_business_key

    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,advisorpass",
            items=[{"application_id": advisor_list.items[0].id, "advisor_score": 92}],
        ),
        principal=advisor_principal,
    )
    _sync_recruitment_application_rows_from_state(store, fake_postgres)

    academy_list = store.get_recruitment_applications(
        status="待初筛确认",
        plan_id=available_plan.id,
        principal=academy_manager_principal,
        page=1,
        page_size=20,
    )

    assert academy_list.total == 1
    assert academy_list.items[0].business_key == submission.application_business_key

    store.confirm_initial_screening(
        academy_list.items[0].id,
        InitialScreeningConfirmationRequest(result="passed", comment="联调通过，进入入营面试"),
        principal=academy_manager_principal,
    )

    portal_student = store.get_portal_student(registration.student.id)
    workflow_progress = portal_student.workflow_progress
    assert portal_student.recruitment_application_status == "入营面试"
    assert workflow_progress is not None
    stage_by_key = {item.key: item for item in workflow_progress.stages}
    assert stage_by_key["initial_screening"].status == "completed"
    assert stage_by_key["camp_interview"].status == "current"
    assert stage_by_key["camp_interview"].description == "初筛已通过，当前进入入营面试环节"
    assert len(fake_postgres.advisor_screening_batches) == 1
    assert len(fake_postgres.initial_screening_confirmations) == 1


def test_end_to_end_screening_flow_rejection_updates_second_choice_and_student_progress(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]
    first_advisor_principal = store.get_principal_context("liu.ya")
    second_advisor_principal = {
        "username": "advisor.wang",
        "full_name": "王青",
        "roles": ["advisor"],
        "permissions": ["recruitment:read", "recruitment:write", "workflow:read", "workflow:write"],
    }

    registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007702",
            email="screening-reject-e2e@example.com",
            full_name="初筛终止联调学生",
            id_number="320000199909099934",
            password="Secret123!",
        )
    )
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    payload = payload.model_copy(
        update={
            "preferences": [
                payload.preferences[0].model_copy(update={"advisor_name": "刘亚", "advisor_user_id": 11}),
                payload.preferences[0].model_copy(update={"preference_order": 2, "advisor_name": "王青", "advisor_user_id": 12}),
            ],
            "selected_advisor_user_id": 11,
            "selected_advisor_name": "刘亚",
        }
    )
    submission = store.submit_portal_application(registration.student.id, payload)

    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == submission.application_business_key)
    _advance_recruitment_application_to_advisor_screening(store, task_id)
    _sync_recruitment_application_rows_from_state(store, fake_postgres)

    first_advisor_list = store.get_recruitment_applications(
        status="待导师初筛,待导师初筛-第一志愿,待导师初筛-第二志愿",
        plan_id=available_plan.id,
        principal=first_advisor_principal,
        page=1,
        page_size=20,
    )
    assert first_advisor_list.total == 1

    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,advisorrejectfirst",
            items=[{"application_id": first_advisor_list.items[0].id, "advisor_score": 79}],
        ),
        principal=first_advisor_principal,
    )
    _sync_recruitment_application_rows_from_state(store, fake_postgres)

    second_advisor_list = store.get_recruitment_applications(
        status="待导师初筛,待导师初筛-第一志愿,待导师初筛-第二志愿",
        plan_id=available_plan.id,
        principal=second_advisor_principal,
        page=1,
        page_size=20,
    )
    assert second_advisor_list.total == 1
    assert second_advisor_list.items[0].business_key == submission.application_business_key

    store.submit_advisor_screening_batch(
        AdvisorScreeningBatchSubmitRequest(
            signature_base64="data:image/png;base64,advisorrejectsecond",
            items=[{"application_id": second_advisor_list.items[0].id, "advisor_score": 60}],
        ),
        principal=second_advisor_principal,
    )

    portal_student = store.get_portal_student(registration.student.id)
    workflow_progress = portal_student.workflow_progress
    assert portal_student.recruitment_application_status == "报名终止"
    assert workflow_progress is not None
    assert workflow_progress.result_label == "终止"
    stage_by_key = {item.key: item for item in workflow_progress.stages}
    assert stage_by_key["background_review"].status == "completed"
    assert stage_by_key["initial_screening"].status == "terminated"
    assert stage_by_key["camp_interview"].status == "terminated"
    assert len(fake_postgres.advisor_screening_batches) == 2


def test_execute_workflow_action_rejects_role_without_workflow_permissions(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    application = store.create_recruitment_application(
        _build_recruitment_application_payload(),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.business_key)

    with pytest.raises(PermissionError, match="当前账号无权执行该流程活动"):
        store.execute_workflow_action(
            task_id,
            "approve",
            None,
            principal={
                "username": "admin",
                "full_name": "管理员",
                "roles": ["platform_admin"],
                "permissions": ["recruitment:read", "recruitment:write", "workflow:read"],
            },
        )


def test_execute_workflow_action_qualification_approve_enters_background_assessment_and_reject_still_sends_email(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009999",
            email="candidate-email@example.com",
            full_name="邮件联调考生",
            id_number="32000019990101123X",
            password="Secret123!",
        )
    )
    applicant_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "candidate-email@example.com")
    application = store.submit_portal_application(
        applicant_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    approve_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.application_business_key)
    store.execute_workflow_action(
        approve_task_id,
        "approve",
        None,
        principal={"username": "academy.admin", "full_name": "书院管理员", "roles": ["academy_admin"]},
    )

    approved_application = next(item for item in store.state["recruitment_applications"] if item["business_key"] == application.application_business_key)
    approved_task = next(item for item in store.state["workflow_tasks"] if item["business_key"] == application.application_business_key)
    assert approved_application["application_status"] == "待背景评估"
    assert approved_task["node_key"] == "background_assessment"
    assert not any(item["full_name"] == "邮件联调考生" for item in store.state["students"])

    rejected_application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(update={"email": "candidate-reject@example.com", "phone_number": "13800009998", "id_number": "32000019990101124X", "student_name": "驳回联调考生"}),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    reject_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == rejected_application.business_key)
    store.execute_workflow_action(
        reject_task_id,
        "reject",
        "材料信息不足",
        principal={"username": "academy.admin", "full_name": "书院管理员", "roles": ["academy_admin"]},
    )

    assert fake_mailer.recruitment_status_calls[-1] == {
        "student_name": "驳回联调考生",
        "email": "candidate-reject@example.com",
        "business_key": rejected_application.business_key,
        "application_status": "驳回重填",
        "plan_name": "2026 秋季博士招生",
        "review_comment": "材料信息不足",
    }


def test_execute_workflow_action_qualification_reject_ignores_student_notification_switch(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_mailer.workflow_notifications_is_enabled = False
    fake_mailer.is_enabled = True
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007777",
            email="reject-switch@example.com",
            full_name="开关联调考生",
            id_number="32000019990101123X",
            password="Secret123!",
        )
    )
    applicant_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "reject-switch@example.com")
    application = store.submit_portal_application(
        applicant_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    approve_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.application_business_key)
    store.execute_workflow_action(
        approve_task_id,
        "reject",
        "资料信息不足，需要补充",
        principal={"username": "academy.admin", "full_name": "书院管理员", "roles": ["academy_admin"]},
    )

    assert fake_mailer.recruitment_material_review_rejection_calls[-1] == {
        "student_name": "开关联调考生",
        "email": "reject-switch@example.com",
        "business_key": application.application_business_key,
        "application_status": "驳回重填",
        "plan_name": "2026 智能制造联合培养",
        "review_comment": "资料信息不足，需要补充",
    }


def test_execute_workflow_action_qualification_approve_does_not_create_student_master_yet(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009999",
            email="candidate-email@example.com",
            full_name="邮件联调考生",
            id_number="32000019990101123X",
            password="Secret123!",
        )
    )

    applicant_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "candidate-email@example.com")
    application = store.submit_portal_application(
        applicant_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    approve_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == application.application_business_key)
    store.execute_workflow_action(
        approve_task_id,
        "approve",
        None,
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    approved_application = next(item for item in store.state["recruitment_applications"] if item["business_key"] == application.application_business_key)
    assert approved_application["portal_student_id"] == applicant_id
    assert fake_postgres.created_students == []


def test_rejected_portal_application_resets_submission_and_allows_resubmit(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009966",
            email="resubmit-student@example.com",
            full_name="驳回重填学生",
            id_number="320000199909099934",
            password="Secret123!",
        )
    )
    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "resubmit-student@example.com")
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    first_submit = store.submit_portal_application(student_id, payload)

    reject_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == first_submit.application_business_key)
    store.execute_workflow_action(
        reject_task_id,
        "reject",
        "请补充材料后重新提交",
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    portal_student = next(item for item in store.state["portal_students"] if item["id"] == student_id)
    latest_application = next(item for item in store.state["recruitment_applications"] if item["business_key"] == first_submit.application_business_key)
    assert portal_student["submitted_at"] is None
    assert latest_application["application_status"] == "驳回重填"
    assert fake_postgres.updated_recruitment_applications[-1]["application_id"] == latest_application["id"]
    assert fake_postgres.updated_recruitment_applications[-1]["payload"]["application_status"] == "驳回重填"
    assert fake_postgres.updated_portal_students[-1]["student_payload"]["id"] == student_id
    assert fake_postgres.updated_portal_students[-1]["student_payload"]["submitted_at"] is None

    second_submit = store.submit_portal_application(student_id, payload)
    assert second_submit.application_status == "报名已提交"


def test_execute_workflow_action_returns_current_detail_for_duplicate_terminal_reject(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    rejected_application = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "email": "candidate-duplicate-reject@example.com",
                "phone_number": "13800009977",
                "id_number": "32000019990101127X",
                "student_name": "重复驳回联调考生",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    reject_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == rejected_application.business_key)

    first_result = store.execute_workflow_action(
        reject_task_id,
        "reject",
        "材料信息不足",
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )
    second_result = store.execute_workflow_action(
        reject_task_id,
        "reject",
        "材料信息不足",
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    assert first_result["task"].status == "已驳回"
    assert second_result["task"].status == "已驳回"
    assert second_result["task"].id == reject_task_id
    assert len(second_result["history"]) == len(first_result["history"])
    assert second_result["history"][-1]["action"] == "reject"


def test_get_registered_portal_students_marks_returned_forms(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009965",
            email="returned-student@example.com",
            full_name="状态联调学生",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "returned-student@example.com")
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    first_submit = store.submit_portal_application(student_id, payload)
    reject_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == first_submit.application_business_key)
    store.execute_workflow_action(
        reject_task_id,
        "reject",
        "退回补充",
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    response = store.get_registered_portal_students(page=1, page_size=20)
    returned_record = next(item for item in response.items if item.email == "returned-student@example.com")
    assert returned_record.application_form_status == "驳回重填"
    assert returned_record.recruitment_application_status == "驳回重填"
    assert returned_record.submitted_at is None


def test_update_recruitment_application_persists_full_state(monkeypatch) -> None:
    from app.schemas.recruitment import RecruitApplicationUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_recruitment_application(_build_recruitment_application_payload())

    updated = store.update_recruitment_application(
        created.id,
        RecruitApplicationUpsert(
            plan_id=created.plan_id,
            student_name="已更新考生",
            graduation_school="复旦大学",
            highest_degree="博士",
            intended_field="具身智能",
            email="candidate-email@example.com",
            phone_number="13800009999",
            id_number="32000019990101123X",
            material_status="材料齐全",
            application_status="报名已提交",
            source_channel="高校老师推荐",
            source_channel_other="校友引荐",
        ),
    )

    assert updated.student_name == "已更新考生"
    assert updated.graduation_school == "复旦大学"
    assert updated.intended_field == "具身智能"
    assert updated.material_status == "材料齐全"
    assert fake_postgres.saved_states
    saved_application = next(item for item in fake_postgres.saved_states[-1]["recruitment_applications"] if item["id"] == created.id)
    assert saved_application["student_name"] == "已更新考生"
    assert saved_application["graduation_school"] == "复旦大学"
    assert saved_application["intended_field"] == "具身智能"
    assert saved_application["source_channel"] == "高校老师推荐"
    assert saved_application["source_channel_other"] == "校友引荐"


def test_register_portal_student_does_not_trigger_email_when_smtp_disabled(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_mailer.is_enabled = False
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800004444",
            email="no-smtp@example.com",
            full_name="未配置邮箱学生",
            id_number="320000199904044440",
            password="Secret123!",
        )
    )

    assert fake_mailer.portal_registration_calls == []


def test_send_portal_registration_email_code_writes_cache_and_sends_email(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()

    response = store.send_portal_registration_email_code("verify@example.com")

    code_key = store._portal_registration_email_code_key("verify@example.com")
    cooldown_key = store._portal_registration_email_cooldown_key("verify@example.com")
    assert response.cooldown_seconds == 60
    assert response.expires_in_seconds == 600
    assert len(fake_cache.get(code_key) or "") == 6
    assert fake_cache.ttl(cooldown_key) == 60
    assert fake_mailer.custom_message_calls[-1]["to_email"] == "verify@example.com"
    assert fake_cache.get(code_key) in fake_mailer.custom_message_calls[-1]["text_body"]


def test_send_portal_login_email_code_writes_cache_and_sends_email(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001234",
            email="login-verify@example.com",
            full_name="验证码登录学生",
            id_number="320000199909099993",
            password="Secret123!",
        )
    )

    response = store.send_portal_login_email_code("login-verify@example.com")

    code_key = store._portal_login_email_code_key("login-verify@example.com")
    cooldown_key = store._portal_login_email_cooldown_key("login-verify@example.com")
    assert response.cooldown_seconds == 60
    assert response.expires_in_seconds == 600
    assert len(fake_cache.get(code_key) or "") == 6
    assert fake_cache.ttl(cooldown_key) == 60
    assert fake_mailer.custom_message_calls[-1]["to_email"] == "login-verify@example.com"
    assert fake_cache.get(code_key) in fake_mailer.custom_message_calls[-1]["text_body"]


def test_login_portal_student_by_email_code_succeeds_and_clears_code(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001235",
            email="email-code-login@example.com",
            full_name="邮箱验证码学生",
            id_number="320000199909099969",
            password="Secret123!",
        )
    )
    fake_cache.set(store._portal_login_email_code_key("email-code-login@example.com"), "123456", ex=600)
    fake_cache.set(store._portal_login_email_cooldown_key("email-code-login@example.com"), "1", ex=60)

    record = store.login_portal_student_by_email_code("email-code-login@example.com", "123456")

    assert record.email == "email-code-login@example.com"
    assert fake_cache.get(store._portal_login_email_code_key("email-code-login@example.com")) is None


def test_login_portal_student_normalizes_datetime_submitted_at(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalLoginRequest, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001237",
            email="datetime-login@example.com",
            full_name="时间登录学生",
            id_number="320000199909099969",
            password="Secret123!",
        )
    )

    student = next(item for item in store.state["portal_students"] if item["email"] == "datetime-login@example.com")
    student["submitted_at"] = datetime(2026, 5, 1, 10, 20, 30)
    student["application_draft"] = {"submitted_at": datetime(2026, 5, 1, 10, 20, 30)}

    record = store.login_portal_student(PortalLoginRequest(account="datetime-login@example.com", password="Secret123!"))

    assert record.submitted_at == "2026-05-01 10:20:30"
    assert record.application_draft is not None
    assert record.application_draft.submitted_at == "2026-05-01 10:20:30"


def test_get_portal_student_refreshes_stale_runtime_submission_state(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    response = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001238",
            email="portal-refresh@example.com",
            full_name="刷新状态学生",
            id_number="320000199909099985",
            password="Secret123!",
        )
    )

    student_id = response.student.id
    runtime_student = next(item for item in store.state["portal_students"] if item["id"] == student_id)
    runtime_student["submitted_at"] = "2026-04-20 10:00:00"
    runtime_student["application_draft"] = {"submitted_at": "2026-04-20 10:00:00"}

    def _get_portal_student_detail(_: int) -> dict:
        return {**runtime_student, "submitted_at": None, "application_draft": {"submitted_at": None}}

    fake_postgres.get_portal_student_detail = _get_portal_student_detail

    record = store.get_portal_student(student_id)

    assert record.submitted_at is None
    assert record.application_draft is not None
    assert record.application_draft.submitted_at is None
    assert runtime_student["submitted_at"] is None


def test_save_portal_application_draft_uses_refreshed_portal_student_state(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationDraftUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    response = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001239",
            email="portal-draft-refresh@example.com",
            full_name="刷新草稿学生",
            id_number="320000199909099993",
            password="Secret123!",
        )
    )

    student_id = response.student.id
    runtime_student = next(item for item in store.state["portal_students"] if item["id"] == student_id)
    runtime_student["submitted_at"] = "2026-04-20 10:00:00"
    runtime_student["application_draft"] = {"submitted_at": "2026-04-20 10:00:00"}

    def _get_portal_student_detail(_: int) -> dict:
        return {**runtime_student, "submitted_at": None, "application_draft": {"submitted_at": None}}

    fake_postgres.get_portal_student_detail = _get_portal_student_detail

    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    record = store.save_portal_application_draft(student_id, PortalApplicationDraftUpsert.model_validate(payload.model_dump(mode="python")))

    assert record.submitted_at is None
    assert fake_postgres.updated_portal_students[-1]["student_payload"]["submitted_at"] is None


def test_save_portal_application_draft_preserves_cleared_nested_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationDraftUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    response = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001240",
            email="portal-draft-clear@example.com",
            full_name="清空字段学生",
            id_number="320000199909099993",
            password="Secret123!",
        )
    )

    student_id = response.student.id
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    payload_data = payload.model_dump(mode="python")
    payload_data["validation_section_id"] = "application-section"
    payload_data["personal_statement"] = {
        "personal_statement_text": None,
        "growth_experience_text": None,
        "program_application_reason_text": None,
        "career_plan_text": None,
        "resume_attachment_url": None,
        "resume_attachment_name": None,
        "supporting_material_attachment_url": None,
        "supporting_material_attachment_name": None,
        "ai_problem_statement": None,
        "ai_industry_opinion": None,
    }

    store.save_portal_application_draft(student_id, PortalApplicationDraftUpsert.model_validate(payload_data))

    saved_draft = fake_postgres.updated_portal_students[-1]["student_payload"]["application_draft"]
    saved_statement = saved_draft["personal_statement"]
    assert "growth_experience_text" in saved_statement and saved_statement["growth_experience_text"] is None
    assert "resume_attachment_url" in saved_statement and saved_statement["resume_attachment_url"] is None
    assert "supporting_material_attachment_url" in saved_statement and saved_statement["supporting_material_attachment_url"] is None


def test_login_portal_student_by_email_code_rejects_wrong_code(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001236",
            email="email-code-login-wrong@example.com",
            full_name="验证码错误学生",
            id_number="320000199909099977",
            password="Secret123!",
        )
    )
    fake_cache.set(store._portal_login_email_code_key("email-code-login-wrong@example.com"), "654321", ex=600)

    with pytest.raises(ValueError, match="邮件验证码不正确"):
        store.login_portal_student_by_email_code("email-code-login-wrong@example.com", "123456")


def test_validate_portal_registration_email_code_rejects_wrong_code(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()
    fake_cache.set(store._portal_registration_email_code_key("verify@example.com"), "654321", ex=600)

    with pytest.raises(ValueError, match="邮件验证码不正确"):
        store.validate_portal_registration_email_code("verify@example.com", "123456")


def test_clear_portal_registration_email_code_removes_cached_code(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()
    code_key = store._portal_registration_email_code_key("verify@example.com")
    cooldown_key = store._portal_registration_email_cooldown_key("verify@example.com")
    fake_cache.set(code_key, "123456", ex=600)
    fake_cache.set(cooldown_key, "1", ex=60)

    store.clear_portal_registration_email_code("verify@example.com")

    assert fake_cache.get(code_key) is None
    assert fake_cache.get(cooldown_key) is None


def test_register_portal_student_rejects_duplicate_phone_number(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800001111",
            email="original-phone-check@example.com",
            full_name="基准学生",
            id_number="320000199905055555",
            password="Secret123!",
        )
    )
    with pytest.raises(ValueError, match="该手机号已注册，请直接登录"):
        store.register_portal_student(
            PortalRegistrationRequest(
                phone_number="13800001111",
                email="different-phone-check@example.com",
                full_name="同手机号学生",
                id_number="320000199905055563",
                password="Secret123!",
            )
        )


def test_register_portal_student_rejects_duplicate_email(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800005555",
            email="portal-student@example.com",
            full_name="基准学生",
            id_number="32000019990606666X",
            password="Secret123!",
        )
    )
    with pytest.raises(ValueError, match="该邮箱已注册，请直接登录"):
        store.register_portal_student(
            PortalRegistrationRequest(
                phone_number="13800005556",
                email="portal-student@example.com",
                full_name="同邮箱学生",
                id_number="320000199906066678",
                password="Secret123!",
            )
        )


def test_register_portal_student_rejects_duplicate_id_number(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800006666",
            email="original-id-check@example.com",
            full_name="基准学生",
            id_number="32000019990101123X",
            password="Secret123!",
        )
    )
    with pytest.raises(ValueError, match="该身份证号已注册，请使用找回密码"):
        store.register_portal_student(
            PortalRegistrationRequest(
                phone_number="13800006667",
                email="different-id-check@example.com",
                full_name="同身份证学生",
                id_number="32000019990101123X",
                password="Secret123!",
            )
        )


def test_register_portal_student_allows_same_full_name(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    response = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007777",
            email="same-name@example.com",
            full_name="张三",
            id_number="320000199907077774",
            password="Secret123!",
        )
    )

    assert response.student.full_name == "张三"
    assert response.student.phone_number == "13800007777"


def test_register_portal_student_auto_binds_latest_recruitment_plan(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    response = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007778",
            email="auto-bind-plan@example.com",
            full_name="自动绑定计划学生",
            id_number="320000199907077782",
            password="Secret123!",
        )
    )

    latest_plan = max(store.state["recruitment_plans"], key=store._portal_plan_sort_key)
    persisted_student = next(item for item in store.state["portal_students"] if item["id"] == response.student.id)

    assert response.student.selected_plan_id == latest_plan["id"]
    assert persisted_student["selected_plan_id"] == latest_plan["id"]


def test_get_registered_portal_students_marks_submission_status(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800008888",
            email="submitted-student@example.com",
            full_name="已报名学生",
            id_number="320000199908088889",
            password="Secret123!",
        )
    )
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009998",
            email="registered-only@example.com",
            full_name="未报名学生",
            id_number="320000199909099985",
            password="Secret123!",
        )
    )

    submitted_student_id = next(
        item["id"] for item in store.state["portal_students"] if item["email"] == "submitted-student@example.com"
    )
    store.submit_portal_application(
        submitted_student_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    response = store.get_registered_portal_students(page=1, page_size=20)
    status_map = {item.email: item.application_form_status for item in response.items}
    submitted_record = next(item for item in response.items if item.email == "submitted-student@example.com")

    assert status_map["submitted-student@example.com"] == "已填写报名"
    assert status_map["registered-only@example.com"] == "未填写报名"
    assert submitted_record.recruitment_application_id is not None
    assert submitted_record.recruitment_application_business_key


def test_export_registered_portal_students_includes_undergraduate_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800006688",
            email="export-undergraduate@example.com",
            full_name="导出本科学生",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    payload_data = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name).model_dump(mode="python")
    payload_data.update(
        {
            "preferences": [
                {
                    "preference_order": 1,
                    "research_center_name": available_team.team_name,
                    "team_id": None,
                    "advisor_name": available_team.lead_advisor_name,
                    "advisor_user_id": None,
                    "is_optional": False,
                }
            ],
            "selected_team_id": None,
            "selected_team_name": available_team.team_name,
            "selected_advisor_user_id": None,
            "selected_advisor_name": available_team.lead_advisor_name,
            "education_experiences": [
                {
                    "sort_order": 1,
                    "education_stage": "高中毕业",
                    "start_month": "2018-09",
                    "end_month": "2021-06",
                    "school_name": "上海市第一中学",
                    "verifier_name": "高中老师",
                    "verifier_phone": "13800001111",
                },
                {
                    "sort_order": 2,
                    "education_stage": "本科在读",
                    "start_month": "2021-09",
                    "end_month": "2025-06",
                    "school_name": "东南大学",
                    "major_name": "人工智能",
                    "average_score": "91.5",
                    "gpa": "3.8",
                    "ranking": "5/120",
                    "verifier_name": "本科辅导员",
                    "verifier_phone": "13800002222",
                    "transcript_attachment_url": "/api/v1/portal/attachments/transcript-undergraduate.pdf",
                },
            ],
        }
    )
    payload = PortalApplicationUpsert.model_validate(payload_data)
    store.submit_portal_application(registration.student.id, payload)

    content = store.export_registered_portal_students([registration.student.id])

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    exported_row = dict(zip(rows[0], rows[1], strict=False))

    assert exported_row["姓名"] == "导出本科学生"
    assert exported_row["报名状态"] == "已填写报名"
    assert exported_row["本科教育阶段"] == "本科在读"
    assert exported_row["本科学校"] == "东南大学"
    assert exported_row["本科专业"] == "人工智能"
    assert exported_row["本科平均成绩"] == "91.5"
    assert exported_row["本科绩点"] == "3.8"
    assert exported_row["本科成绩排名"] == "5/120"
    assert exported_row["本科证明人姓名"] == "本科辅导员"
    assert exported_row["本科证明人手机"] == "13800002222"
    assert exported_row["本科成绩单附件"] == "https://admissions.pjlab.org.cn/api/v1/portal/attachments/transcript-undergraduate.pdf"
    assert exported_row["本科学位证附件"] is None
    assert exported_row["本科毕业证附件"] is None


def test_export_registered_portal_students_supports_filtered_full_export_with_saved_draft(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationDraftUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    draft_registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007788",
            email="draft-export@example.com",
            full_name="草稿导出学生",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800007789",
            email="other-export@example.com",
            full_name="其他学生",
            id_number="320000199909099985",
            password="Secret123!",
        )
    )

    payload_data = _build_portal_application_payload(
        available_plan.id,
        available_team.team_name,
        available_team.lead_advisor_name,
    ).model_dump(mode="python")
    payload_data.update(
        {
            "preferences": [
                {
                    "preference_order": 1,
                    "research_center_name": available_team.team_name,
                    "team_id": None,
                    "advisor_name": available_team.lead_advisor_name,
                    "advisor_user_id": None,
                    "is_optional": False,
                }
            ],
            "selected_team_id": None,
            "selected_team_name": available_team.team_name,
            "selected_advisor_user_id": None,
            "selected_advisor_name": available_team.lead_advisor_name,
            "source_channel_other": "老师推荐",
            "personal_statement": {
                **payload_data["personal_statement"],
                "supporting_material_attachment_url": "/portal-attachments/uploads/student-7/material/support-a.pdf",
                "supporting_material_attachment_name": "support-a.pdf",
            },
            "declaration": {
                "has_read_declaration": True,
                "declaration_text": "我已知悉报名要求",
                "progress_snapshot": {"profile": True, "statement": True},
            },
        }
    )
    store.save_portal_application_draft(draft_registration.student.id, PortalApplicationDraftUpsert.model_validate(payload_data))

    runtime_student = next(item for item in store.state["portal_students"] if item["id"] == draft_registration.student.id)

    def _get_portal_student_detail(_: int) -> dict:
        return runtime_student

    fake_postgres.get_portal_student_detail = _get_portal_student_detail

    monkeypatch.setattr(
        store,
        "get_registered_portal_students",
        lambda **kwargs: SimpleNamespace(items=[SimpleNamespace(id=draft_registration.student.id)], total=1),
    )

    content = store.export_registered_portal_students([], keyword="草稿导出", application_form_status="未填写报名")

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))

    assert len(rows) == 2

    exported_row = dict(zip(rows[0], rows[1], strict=False))
    assert exported_row["姓名"] == "草稿导出学生"
    assert exported_row["报名状态"] == "未填写报名"
    assert str(exported_row["证件照附件"]).startswith("https://admissions.pjlab.org.cn/")
    assert str(exported_row["证件照附件"]).endswith("profile_photo/photo-a.jpg")
    assert str(exported_row["个人陈述简历附件"]).startswith("https://admissions.pjlab.org.cn/")
    assert str(exported_row["个人陈述简历附件"]).endswith("resume/resume-a.pdf")
    assert str(exported_row["个人陈述支撑材料附件"]).startswith("https://admissions.pjlab.org.cn/")
    assert str(exported_row["个人陈述支撑材料附件"]).endswith("material/support-a.pdf")
    assert exported_row["信息来源渠道"] == "上海人工智能实验室官网"
    assert exported_row["其他信息来源"] == "老师推荐"
    assert exported_row["声明内容"] == "我已知悉报名要求"
    assert '"profile": true' in str(exported_row["声明进度快照JSON"])
    assert '"source_channel_other": "老师推荐"' in str(exported_row["报名草稿JSON"])


def test_export_registered_portal_students_maps_returned_to_rejected_status(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009988",
            email="returned-export@example.com",
            full_name="退回状态学生",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    payload = _build_portal_application_payload(
        available_plan.id,
        available_team.team_name,
        available_team.lead_advisor_name,
    )
    store.submit_portal_application(registration.student.id, PortalApplicationUpsert.model_validate(payload.model_dump(mode="python")))

    runtime_student = next(item for item in store.state["portal_students"] if item["id"] == registration.student.id)
    runtime_student["submitted_at"] = None

    for application in store.state["recruitment_applications"]:
        if int(application.get("portal_student_id") or 0) == registration.student.id:
            application["application_status"] = "returned"

    def _get_portal_student_detail(_: int) -> dict:
        return runtime_student

    fake_postgres.get_portal_student_detail = _get_portal_student_detail

    monkeypatch.setattr(
        store,
        "get_registered_portal_students",
        lambda **kwargs: SimpleNamespace(items=[SimpleNamespace(id=registration.student.id)], total=1),
    )

    content = store.export_registered_portal_students([registration.student.id])

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    exported_row = dict(zip(rows[0], rows[1], strict=False))

    assert exported_row["姓名"] == "退回状态学生"
    assert exported_row["报名状态"] == "驳回重填"


def test_build_registered_portal_student_export_row_does_not_truncate_repeated_groups(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    first_student = PortalStudentRecord.model_validate(
        {
            "id": 1,
            "full_name": "多经历学生",
            "phone_number": "13800001234",
            "email": "multi@example.com",
            "id_number": "320000199909099918",
            "application_draft": {
                "preferences": [
                    {
                        "preference_order": 1,
                        "research_center_name": "方向一",
                        "advisor_name": "导师甲",
                        "is_optional": False,
                    },
                    {
                        "preference_order": 2,
                        "research_center_name": "方向二",
                        "advisor_name": "导师乙",
                        "is_optional": True,
                    },
                ],
                "education_experiences": [
                    {"sort_order": 1, "education_stage": "高中毕业", "school_name": "高中A"},
                    {"sort_order": 2, "education_stage": "本科毕业", "school_name": "本科A"},
                    {"sort_order": 3, "education_stage": "硕士毕业", "school_name": "硕士A"},
                    {"sort_order": 4, "education_stage": "博士在读", "school_name": "博士A"},
                ],
            },
        }
    )
    second_student = PortalStudentRecord.model_validate(
        {
            "id": 2,
            "full_name": "少经历学生",
            "phone_number": "13800001235",
            "email": "single@example.com",
            "id_number": "320000199909099985",
            "application_draft": {
                "preferences": [
                    {
                        "preference_order": 1,
                        "research_center_name": "唯一方向",
                        "advisor_name": "导师丙",
                        "is_optional": False,
                    }
                ],
                "education_experiences": [
                    {"sort_order": 1, "education_stage": "本科毕业", "school_name": "本科B"},
                ],
            },
        }
    )

    records = [
        store._build_registered_portal_student_export_row(first_student, None, None, None, None, None),
        store._build_registered_portal_student_export_row(second_student, None, None, None, None, None),
    ]

    workbook = load_workbook(BytesIO(build_registered_portal_students_template(records)), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    first_row = dict(zip(rows[0], rows[1], strict=False))
    second_row = dict(zip(rows[0], rows[2], strict=False))

    assert first_row["志愿2导师"] == "导师乙"
    assert second_row["志愿2导师"] is None
    assert first_row["教育经历4学校名称"] == "博士A"
    assert second_row["教育经历4学校名称"] is None


def test_create_registered_portal_student_export_job_completes_and_can_download(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.auth import Principal
    from app.schemas.portal import PortalRegistrationRequest
    from app.schemas.student import RegisteredPortalStudentExportRequest

    class ImmediateExecutor:
        def submit(self, fn, *args, **kwargs):
            fn(*args, **kwargs)
            return None

    monkeypatch.setattr(RuntimeManagementStoreStudentsMixin, "_registered_portal_export_executor", ImmediateExecutor())

    store = RuntimeManagementStore()
    registration = store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800006689",
            email="async-export@example.com",
            full_name="异步导出学生",
            id_number="320000199909099985",
            password="Secret123!",
        )
    )
    principal = Principal(username="student-admin", full_name="学生管理员", roles=["student_admin"], permissions=["students:read"])

    response = store.create_registered_portal_student_export_job(
        RegisteredPortalStudentExportRequest(ids=[registration.student.id]),
        principal=principal,
    )
    jobs = store.list_registered_portal_student_export_jobs(principal=principal)

    assert response.job.job_id
    assert jobs.items[0].status == "completed"
    assert jobs.items[0].download_url

    file_name, content = store.get_registered_portal_student_export_job_download(response.job.job_id, principal=principal)

    assert file_name.endswith(".xlsx")
    assert content


def test_get_student_stats_includes_registered_portal_student_counts(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009991",
            email="stats-submitted@example.com",
            full_name="统计已报名",
            id_number="320000199909099918",
            password="Secret123!",
        )
    )
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009992",
            email="stats-unsubmitted@example.com",
            full_name="统计未报名",
            id_number="320000199909099926",
            password="Secret123!",
        )
    )

    submitted_student_id = next(
        item["id"] for item in store.state["portal_students"] if item["email"] == "stats-submitted@example.com"
    )
    store.submit_portal_application(
        submitted_student_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name).model_copy(update={"graduation_school": "东南大学"}),
    )

    stats = store.get_student_stats()

    assert stats.registered_portal_total == 2
    assert stats.portal_submitted_total == 1
    assert stats.portal_unsubmitted_total == 1


def test_get_student_stats_uses_empty_formal_student_state_without_reviving_seed_centers(monkeypatch) -> None:
    class EmptyFormalStudentStateStore(FakePostgresStateStore):
        def load_student_state(self):
            return []

        def load_team_state(self):
            return [
                {
                    "id": 7,
                    "team_code": "CENTER-007",
                    "team_name": "前沿探索中心",
                    "department_name": "",
                    "discipline_name": "",
                    "lead_user_id": None,
                    "lead_advisor_name": None,
                    "advisor_names": [],
                    "advisor_ids": [],
                    "advisor_relation_ids": [],
                    "research_directions": [],
                    "status": "启用",
                    "established_on": "2026-04-30",
                    "description": None,
                },
                {
                    "id": 8,
                    "team_code": "CENTER-008",
                    "team_name": "AI For Science中心",
                    "department_name": "",
                    "discipline_name": "",
                    "lead_user_id": None,
                    "lead_advisor_name": None,
                    "advisor_names": [],
                    "advisor_ids": [],
                    "advisor_relation_ids": [],
                    "research_directions": [],
                    "status": "启用",
                    "established_on": "2026-04-30",
                    "description": None,
                },
            ]

        def load_recruitment_plan_state(self):
            return []

        def load_portal_student_state(self):
            return []

        def load_recruitment_application_state(self):
            return []

        def load_workflow_task_state(self):
            return []

    fake_postgres = EmptyFormalStudentStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    stats = store.get_student_stats()

    assert stats.total_students == 0
    assert stats.active_students == 0
    assert stats.center_total == 2
    assert stats.enabled_center_total == 2


def test_submit_portal_application_uses_incremental_portal_sync(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009990",
            email="incremental-submit@example.com",
            full_name="增量提交学生",
            id_number="32000019990909990X",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "incremental-submit@example.com")
    save_count_before_submit = len(fake_postgres.saved_states)
    response = store.submit_portal_application(
        student_id,
        _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name),
    )

    expected_year = f"{datetime.now().year + 1:04d}"
    assert response.application_business_key.startswith("SH")
    assert response.application_business_key[2:6] == expected_year
    assert response.application_business_key[6:].isdigit()
    assert len(response.application_business_key) == 10
    assert len(fake_postgres.portal_application_submissions) == 1
    assert len(fake_postgres.saved_states) == save_count_before_submit


def test_submit_portal_application_allows_advisor_only_preference_without_team(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009991",
            email="advisor-only-submit@example.com",
            full_name="导师直选学生",
            id_number="320000199909099934",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "advisor-only-submit@example.com")
    payload_data = _build_portal_application_payload(available_plan.id, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload_data["preferences"][0]["team_id"] = None
    payload_data["preferences"][0]["research_center_name"] = None
    payload_data["selected_team_id"] = None
    payload_data["selected_team_name"] = None

    response = store.submit_portal_application(student_id, PortalApplicationUpsert.model_validate(payload_data))

    assert response.application_status == "报名已提交"
    assert fake_postgres.portal_application_submissions[-1]["portal_student_payload"]["selected_team_name"] is None
    assert fake_postgres.portal_application_submissions[-1]["application_payload"]["first_choice"] == "刘亚"


def test_migrate_workflow_runtime_normalizes_legacy_recruitment_business_key(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    created = store.create_recruitment_application(
        _build_recruitment_application_payload().model_copy(
            update={
                "student_name": "历史报名号考生",
                "phone_number": "13800006666",
                "email": "legacy-business-key@example.com",
                "id_number": "320000199901016666",
            }
        ),
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    legacy_business_key = "REC-20270301-0001"
    application = next(item for item in store.state["recruitment_applications"] if item["id"] == created.id)
    task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == created.business_key)
    workflow_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)

    application["business_key"] = legacy_business_key
    application["candidate_no"] = legacy_business_key
    application["applied_at"] = "2027-03-01 10:00:00"
    application["created_at"] = "2027-03-01 10:00:00"
    workflow_task["business_key"] = legacy_business_key
    workflow_task["created_at"] = "2027-03-01 10:00:00"
    workflow_task["form_summary"] = f"业务编号：{legacy_business_key}；研究方向：{application['first_choice']}"

    changed = store._migrate_workflow_runtime()

    normalized_application = next(item for item in store.state["recruitment_applications"] if item["id"] == created.id)
    normalized_task = next(item for item in store.state["workflow_tasks"] if item["id"] == task_id)

    assert changed is True
    assert normalized_application["business_key"].startswith("SH2028")
    assert len(normalized_application["business_key"]) == 10
    assert normalized_application["candidate_no"] == normalized_application["business_key"]
    assert normalized_task["business_key"] == normalized_application["business_key"]
    assert normalized_task["form_summary"].startswith(f"业务编号：{normalized_application['business_key']}；")


def test_submit_portal_application_allows_resubmit_when_submitted_at_is_stale(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009967",
            email="stale-submit@example.com",
            full_name="脏提交时间学生",
            id_number="320000199909099934",
            password="Secret123!",
        )
    )
    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "stale-submit@example.com")
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    first_submit = store.submit_portal_application(student_id, payload)

    reject_task_id = next(item["id"] for item in store.state["workflow_tasks"] if item["business_key"] == first_submit.application_business_key)
    store.execute_workflow_action(
        reject_task_id,
        "reject",
        "请补充材料后重新提交",
        principal={"username": "admin", "full_name": "管理员", "roles": ["platform_admin"]},
    )

    portal_student = next(item for item in store.state["portal_students"] if item["id"] == student_id)
    portal_student["submitted_at"] = "2026-05-06 20:00:00"

    second_submit = store.submit_portal_application(student_id, payload)

    reopened_task = next(item for item in store.state["workflow_tasks"] if item["business_key"] == second_submit.application_business_key)
    assert second_submit.application_status == "报名已提交"
    assert reopened_task["status"] == "待处理"
    assert reopened_task["node_key"] == "qualification_review"
    assert fake_postgres.portal_application_submissions[-1]["workflow_task"]["node_key"] == "qualification_review"


def test_init_persists_migrated_workflow_runtime_loaded_from_postgres(monkeypatch) -> None:
    class StaleWorkflowStateStore(FakePostgresStateStore):
        def load_state(self):
            return None

        def load_student_state(self):
            return []

        def load_team_state(self):
            return []

        def load_recruitment_plan_state(self):
            return []

        def load_portal_student_state(self):
            return []

        def load_recruitment_application_state(self):
            return [
                {
                    "id": 27,
                    "plan_id": 5,
                    "portal_student_id": 1,
                    "student_name": "罗凯",
                    "business_key": "SH20260008",
                    "candidate_no": "SH20260008",
                    "first_choice": "前沿探索中心",
                    "intended_field": "前沿探索中心",
                    "material_status": "待补材料",
                    "application_status": "报名已提交",
                    "applied_at": "2026-05-01 04:45:27",
                }
            ]

        def load_workflow_task_state(self):
            return [
                {
                    "id": 1,
                    "workflow_name": "招生报名审批流程",
                    "business_module": "招生管理",
                    "business_key": "SH20260008",
                    "title": "罗凯报名审核",
                    "applicant_name": "罗凯",
                    "current_handler": "流程结束",
                    "current_node": "流程结束",
                    "priority": "中",
                    "status": "已驳回",
                    "created_at": "2026-05-01 04:26:31",
                    "due_at": "2026-05-03 04:26:31",
                    "form_summary": "业务编号：SH20260008；研究方向：前沿探索中心；材料状态：待补材料",
                    "latest_comment": "审核不通过",
                    "flow_code": "recruitment_application",
                    "node_key": "流程结束",
                    "entity_id": 27,
                    "history": [],
                }
            ]

    fake_postgres = StaleWorkflowStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr(
        "app.services.management_service_core.build_runtime_seed_state",
        lambda: {
            "counters": {},
            "roles": [],
            "profiles": {},
            "teams": [],
            "students": [],
            "recruitment_plans": [],
            "portal_students": [],
            "recruitment_applications": [],
            "workflow_tasks": [],
            "system_users": [],
            "audit_policies": [],
            "scientific_reports": [],
            "outbound_studies": [],
            "theses": [],
            "operation_logs": [],
            "dict_items": [],
            "integrations": [],
        },
    )

    store = RuntimeManagementStore()

    repaired_task = next(item for item in store.state["workflow_tasks"] if item["id"] == 1)
    assert repaired_task["status"] == "待处理"
    assert repaired_task["node_key"] == "qualification_review"
    assert fake_postgres.synced_workflow_tasks
    saved_task = fake_postgres.synced_workflow_tasks[-1]["task_payload"]
    assert saved_task["id"] == 1
    assert saved_task["status"] == "待处理"
    assert saved_task["node_key"] == "qualification_review"


def test_init_rebases_workflow_task_counter_before_creating_missing_runtime_task(monkeypatch) -> None:
    class MissingWorkflowTaskStateStore(FakePostgresStateStore):
        def load_state(self):
            return {
                "counters": {"workflow_tasks": 25},
                "roles": [],
                "profiles": {},
                "teams": [],
                "students": [],
                "recruitment_plans": [],
                "portal_students": [],
                "recruitment_applications": [],
                "workflow_tasks": [],
                "system_users": [],
                "audit_policies": [],
                "scientific_reports": [],
                "outbound_studies": [],
                "theses": [],
                "operation_logs": [],
                "dict_items": [],
                "integrations": [],
            }

        def load_student_state(self):
            return []

        def load_team_state(self):
            return []

        def load_recruitment_plan_state(self):
            return []

        def load_portal_student_state(self):
            return []

        def load_recruitment_application_state(self):
            return [
                {
                    "id": 27,
                    "plan_id": 5,
                    "portal_student_id": 1,
                    "student_name": "罗凯",
                    "business_key": "SH20260007",
                    "candidate_no": "SH20260007",
                    "first_choice": "智能制造团队",
                    "intended_field": "智能制造团队",
                    "material_status": "待补材料",
                    "application_status": "报名已提交",
                    "applied_at": "2026-05-01 04:45:27",
                }
            ]

        def load_workflow_task_state(self):
            return [
                {
                    "id": 26,
                    "workflow_name": "授位审批流程",
                    "business_module": "学位管理",
                    "business_key": "SWSQSP202604070002",
                    "title": "重复 ID 占位任务",
                    "applicant_name": "测试学生",
                    "current_handler": "研究生秘书",
                    "current_node": "材料复核",
                    "priority": "中",
                    "status": "处理中",
                    "created_at": "2026-05-01 04:26:31",
                    "due_at": "2026-05-03 04:26:31",
                    "form_summary": "占位任务",
                    "latest_comment": None,
                    "flow_code": "thesis",
                    "node_key": "secretary_review",
                    "entity_id": 26,
                    "history": [],
                }
            ]

    fake_postgres = MissingWorkflowTaskStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    created_task = next(item for item in store.state["workflow_tasks"] if item["business_key"] == "SH20260007")

    assert created_task["id"] == 27
    assert len([item for item in store.state["workflow_tasks"] if int(item["id"]) == 26]) == 1
    assert fake_postgres.synced_workflow_tasks
    saved_task = fake_postgres.synced_workflow_tasks[-1]["task_payload"]
    assert saved_task["id"] == 27
    assert saved_task["business_key"] == "SH20260007"


def test_submit_portal_application_blocks_modification_after_submission(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009989",
            email="reapply-student@example.com",
            full_name="重报学生",
            id_number="320000199909099942",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "reapply-student@example.com")
    first_submit = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    store.submit_portal_application(student_id, first_submit)

    with pytest.raises(ValueError, match="当前仅支持只读浏览"):
        store.submit_portal_application(student_id, first_submit)


def test_save_portal_application_draft_blocks_modification_after_submission(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalApplicationDraftUpsert, PortalRegistrationRequest

    store = RuntimeManagementStore()
    available_plan = store.get_public_recruitment_plans().items[0]
    available_team = store.get_public_teams().items[0]

    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009987",
            email="readonly-student@example.com",
            full_name="只读学生",
            id_number="320000199909099934",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "readonly-student@example.com")
    payload = _build_portal_application_payload(available_plan.id, available_team.team_name, available_team.lead_advisor_name)
    store.submit_portal_application(student_id, payload)

    with pytest.raises(ValueError, match="当前仅支持只读浏览"):
        store.save_portal_application_draft(student_id, PortalApplicationDraftUpsert.model_validate(payload.model_dump(mode="python")))


def test_deactivate_registered_portal_student_blocks_login(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalLoginRequest, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009993",
            email="deactivate-student@example.com",
            full_name="注销学生",
            id_number="320000199909099993",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "deactivate-student@example.com")
    response = store.deactivate_registered_portal_student(student_id)

    assert response.account_status == "停用"
    with pytest.raises(ValueError, match="账号已停用，请联系管理员"):
        store.login_portal_student(PortalLoginRequest(account="deactivate-student@example.com", password="Secret123!"))


def test_activate_registered_portal_student_restores_login(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalLoginRequest, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009996",
            email="reactivate-student@example.com",
            full_name="重新启用学生",
            id_number="320000199909099969",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "reactivate-student@example.com")
    store.deactivate_registered_portal_student(student_id)

    response = store.activate_registered_portal_student(student_id)

    assert response.account_status == "启用"
    record = store.login_portal_student(PortalLoginRequest(account="reactivate-student@example.com", password="Secret123!"))
    assert record.email == "reactivate-student@example.com"


def test_change_portal_student_password_requires_current_password(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    from app.schemas.portal import PortalLoginRequest, PortalPasswordChangeRequest, PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009997",
            email="profile-password@example.com",
            full_name="个人空间学生",
            id_number="320000199909099977",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "profile-password@example.com")

    with pytest.raises(ValueError, match="当前密码不正确"):
        store.change_portal_student_password(
            student_id,
            PortalPasswordChangeRequest(current_password="Wrong123!", new_password="NewSecret123!"),
        )

    store.change_portal_student_password(
        student_id,
        PortalPasswordChangeRequest(current_password="Secret123!", new_password="NewSecret123!"),
    )

    record = store.login_portal_student(PortalLoginRequest(account="profile-password@example.com", password="NewSecret123!"))
    assert record.email == "profile-password@example.com"


def test_reset_registered_portal_student_password_sends_email_when_smtp_enabled(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009994",
            email="admin-reset-student@example.com",
            full_name="重置密码学生",
            id_number="320000199909099942",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "admin-reset-student@example.com")
    response = store.reset_registered_portal_student_password(student_id)

    assert response.email_sent is True
    assert response.temporary_password
    assert fake_mailer.portal_admin_password_reset_calls == [
        {
            "full_name": "重置密码学生",
            "email": "admin-reset-student@example.com",
            "temporary_password": response.temporary_password,
        }
    ]


def test_send_registered_portal_student_email_skips_when_smtp_disabled(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_mailer.is_enabled = False
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest
    from app.schemas.student import RegisteredPortalStudentEmailRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009995",
            email="custom-email-student@example.com",
            full_name="邮件学生",
            id_number="320000199909099950",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "custom-email-student@example.com")
    response = store.send_registered_portal_student_email(
        student_id,
        RegisteredPortalStudentEmailRequest(subject="自定义通知", content="请尽快完善报名信息。"),
    )

    assert response.email_sent is False
    assert fake_mailer.custom_message_calls == []


def test_send_registered_portal_student_email_skips_when_workflow_notifications_disabled(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_mailer.is_enabled = True
    fake_mailer.workflow_notifications_is_enabled = False
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest
    from app.schemas.student import RegisteredPortalStudentEmailRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009996",
            email="workflow-disabled-student@example.com",
            full_name="流程邮件关闭学生",
            id_number="320000199909099951",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "workflow-disabled-student@example.com")
    response = store.send_registered_portal_student_email(
        student_id,
        RegisteredPortalStudentEmailRequest(subject="自定义通知", content="请尽快完善报名信息。"),
    )

    assert response.email_sent is False
    assert fake_mailer.custom_message_calls == []


def test_reset_registered_portal_student_password_still_sends_when_workflow_notifications_disabled(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_mailer = FakeNotificationEmailService()
    fake_mailer.is_enabled = True
    fake_mailer.workflow_notifications_is_enabled = False
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.NotificationEmailService", lambda **kwargs: fake_mailer)

    from app.schemas.portal import PortalRegistrationRequest

    store = RuntimeManagementStore()
    store.register_portal_student(
        PortalRegistrationRequest(
            phone_number="13800009997",
            email="workflow-reset-disabled@example.com",
            full_name="流程重置关闭学生",
            id_number="320000199909099952",
            password="Secret123!",
        )
    )

    student_id = next(item["id"] for item in store.state["portal_students"] if item["email"] == "workflow-reset-disabled@example.com")
    response = store.reset_registered_portal_student_password(student_id)

    assert response.email_sent is True
    assert fake_mailer.portal_admin_password_reset_calls == [
        {
            "full_name": "流程重置关闭学生",
            "email": "workflow-reset-disabled@example.com",
            "temporary_password": response.temporary_password,
        }
    ]


def test_recruitment_application_upsert_rejects_invalid_resident_id_number() -> None:
    from pydantic import ValidationError

    from app.schemas.recruitment import RecruitApplicationUpsert

    with pytest.raises(ValidationError, match="居民身份证号码格式不正确"):
        RecruitApplicationUpsert(
            plan_id=1,
            student_name="身份证校验考生",
            graduation_school="江南大学",
            highest_degree="硕士",
            intended_field="智能制造",
            id_type="居民身份证",
            id_number="320000199901011234",
            material_status="待审核",
            application_status="报名已提交",
        )


def test_portal_registration_request_rejects_invalid_id_number() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalRegistrationRequest

    with pytest.raises(ValidationError, match="身份证号格式不正确"):
        PortalRegistrationRequest(
            full_name="无效身份证学生",
            phone_number="13800001234",
            email="invalid-id@example.com",
            id_number="320000199901011234",
            password="Secret123!",
        )


def test_portal_registration_request_rejects_invalid_phone_or_email() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalRegistrationRequest

    with pytest.raises(ValidationError, match="手机号格式不正确|邮箱格式不正确"):
        PortalRegistrationRequest(
            full_name="联系方式无效学生",
            phone_number="12345",
            email="invalid-email",
            id_number="32000019990101123X",
            password="Secret123!",
        )


def test_recruitment_application_upsert_rejects_invalid_phone_or_email() -> None:
    from pydantic import ValidationError

    from app.schemas.recruitment import RecruitApplicationUpsert

    with pytest.raises(ValidationError, match="手机号格式不正确|邮箱格式不正确"):
        RecruitApplicationUpsert(
            plan_id=1,
            student_name="联系方式校验考生",
            graduation_school="江南大学",
            highest_degree="硕士",
            intended_field="智能制造",
            phone_number="12345",
            email="invalid-email",
            material_status="待审核",
            application_status="报名已提交",
        )


def test_user_profile_update_rejects_invalid_phone_or_email() -> None:
    from pydantic import ValidationError

    from app.schemas.auth import UserProfileUpdate

    with pytest.raises(ValidationError, match="手机号格式不正确|邮箱格式不正确"):
        UserProfileUpdate(
            full_name="测试用户",
            phone_number="12345",
            email="invalid-email",
            theme_color="#409eff",
        )


def test_student_upsert_rejects_invalid_phone_number() -> None:
    from pydantic import ValidationError

    from app.schemas.student import StudentUpsert

    with pytest.raises(ValidationError, match="手机号格式不正确"):
        StudentUpsert(
            student_no="20260001",
            full_name="测试学生",
            status="在校",
            advisor_name="刘亚",
            center_name="智能制造联合团队",
            degree_type="工程博士",
            enrollment_year=2026,
            phone_number="12345",
        )


def test_system_user_upsert_rejects_invalid_phone_number() -> None:
    from pydantic import ValidationError

    from app.schemas.system import SystemUserUpsert

    with pytest.raises(ValidationError, match="手机号格式不正确"):
        SystemUserUpsert(
            username="invalid.phone",
            full_name="测试账号",
            role_code="admin",
            department_name="信息中心",
            phone_number="12345",
            account_status="启用",
            password="Secret123!",
        )


def test_system_user_upsert_rejects_invalid_email() -> None:
    from pydantic import ValidationError

    from app.schemas.system import SystemUserUpsert

    with pytest.raises(ValidationError, match="邮箱格式不正确"):
        SystemUserUpsert(
            username="invalid.email",
            full_name="测试账号",
            role_code="admin",
            department_name="信息中心",
            email="invalid-email",
            account_status="启用",
            password="Secret123!",
        )


def test_system_user_upsert_requires_introduction_for_advisor() -> None:
    from pydantic import ValidationError

    from app.schemas.system import SystemUserUpsert

    with pytest.raises(ValidationError, match="导师角色必须填写介绍"):
        SystemUserUpsert(
            username="advisor.nointro",
            full_name="测试导师",
            role_code="advisor",
            department_name="智能中心",
            introduction="   ",
            email="advisor@example.com",
            phone_number="13800001111",
            account_status="启用",
            password="Secret123!",
        )


def test_update_system_user_requires_email_and_phone(monkeypatch) -> None:
    from app.schemas.system import SystemUserUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")

    with pytest.raises(ValueError, match="邮箱不能为空"):
        store.update_system_user(
            int(target_user["id"]),
            SystemUserUpsert(
                username=target_user["username"],
                full_name=target_user["full_name"],
                role_code=target_user["role_code"],
                department_name=target_user.get("department_name") or "",
                introduction="导师简介占位",
                email=None,
                phone_number="13800001111",
                account_status=target_user["account_status"],
                password=None,
            ),
        )

    with pytest.raises(ValueError, match="手机号不能为空"):
        store.update_system_user(
            int(target_user["id"]),
            SystemUserUpsert(
                username=target_user["username"],
                full_name=target_user["full_name"],
                role_code=target_user["role_code"],
                department_name=target_user.get("department_name") or "",
                introduction="导师简介占位",
                email="user@example.com",
                phone_number=None,
                account_status=target_user["account_status"],
                password=None,
            ),
        )


def test_update_system_user_uses_postgres_row_when_runtime_user_missing(monkeypatch) -> None:
    from app.schemas.system import SystemUserUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")
    target_role = next(role for role in store.state["roles"] if role["role_code"] == target_user["role_code"])
    user_id = int(target_user["id"])

    fake_postgres.system_user_rows[user_id] = {
        "id": user_id,
        "username": target_user["username"],
        "full_name": target_user["full_name"],
        "role_code": target_user["role_code"],
        "role_name": target_role["role_name"],
        "department_name": target_user.get("department_name") or "",
        "introduction": "原始导师介绍",
        "email": "existing@example.com",
        "phone_number": "13800001111",
        "account_status": target_user["account_status"],
        "last_login_at": target_user.get("last_login_at"),
        "password_hash": target_user.get("password_hash"),
    }
    store.state["system_users"] = [item for item in store.state["system_users"] if int(item["id"]) != user_id]

    updated = store.update_system_user(
        user_id,
        SystemUserUpsert(
            username=target_user["username"],
            full_name="关系库回填用户",
            role_code=target_user["role_code"],
            department_name="研究生院",
            introduction="更新后的导师介绍",
            email="patched@example.com",
            phone_number="13800002222",
            account_status=target_user["account_status"],
            password=None,
        ),
    )

    assert updated.full_name == "关系库回填用户"
    assert fake_postgres.saved_states[-1]["user_payload"]["id"] == user_id
    assert any(int(item["id"]) == user_id and item["full_name"] == "关系库回填用户" for item in store.state["system_users"])


def test_create_system_user_persists_advisor_introduction(monkeypatch) -> None:
    from app.schemas.system import SystemUserUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    created = store.create_system_user(
        SystemUserUpsert(
            username="mentor.intro",
            full_name="导师介绍测试",
            role_code="advisor",
            department_name="智能中心",
            introduction="长期从事智能制造与工业软件方向研究。",
            email="mentor.intro@example.com",
            phone_number="13800009999",
            account_status="启用",
            password="Secret123!",
        )
    )

    assert created.introduction == "长期从事智能制造与工业软件方向研究。"
    assert store.state["profiles"]["mentor.intro"]["introduction"] == "长期从事智能制造与工业软件方向研究。"
    assert fake_postgres.saved_states[-1]["profile_payload"]["introduction"] == "长期从事智能制造与工业软件方向研究。"


def test_get_system_users_does_not_fallback_to_runtime_seed_when_postgres_query_fails(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    with pytest.raises(DatabaseUnavailableError, match="系统用户数据当前仅允许从数据库读取"):
        store.get_system_users(role_code="advisor", page=1, page_size=10)


def test_export_system_users_supports_filtered_results(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["role_code"] == "advisor")
    user_id = int(target_user["id"])
    fake_postgres.system_user_rows[user_id] = {
        "id": user_id,
        "username": target_user["username"],
        "full_name": target_user["full_name"],
        "role_code": target_user["role_code"],
        "department_name": "智能中心",
        "introduction": "导师导出测试",
        "email": "advisor.export@example.com",
        "phone_number": "13800005555",
        "account_status": "启用",
        "last_login_at": "2026-05-25 10:00:00",
    }

    content = store.export_system_users(role_code="advisor", operator_username="admin")

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    assert worksheet.title == "系统用户"
    assert worksheet.cell(1, 2).value == "账号"
    exported_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert any(row[1] == target_user["username"] for row in exported_rows)
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["action"] == "导出账号"


def test_export_system_users_supports_selected_ids(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    candidates = [item for item in store.state["system_users"] if item["username"] != "admin"][:2]
    for index, item in enumerate(candidates, start=1):
        fake_postgres.system_user_rows[int(item["id"])] = {
            "id": int(item["id"]),
            "username": item["username"],
            "full_name": item["full_name"],
            "role_code": item["role_code"],
            "department_name": f"部门{index}",
            "introduction": f"介绍{index}",
            "email": f"user{index}@example.com",
            "phone_number": f"1380000555{index}",
            "account_status": item["account_status"],
            "last_login_at": None,
        }

    selected_ids = [int(candidates[1]["id"]), int(candidates[0]["id"])]
    content = store.export_system_users(selected_ids, operator_username="auditor")

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    exported_usernames = [worksheet.cell(row, 2).value for row in range(2, worksheet.max_row + 1)]
    assert exported_usernames == [candidates[1]["username"], candidates[0]["username"]]
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["operator_username"] == "auditor"


def test_import_system_users_generates_unique_username_and_allows_empty_optional_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    fake_postgres.system_user_rows[901] = {
        "id": 901,
        "username": "wangli",
        "full_name": "王力",
        "role_code": "admin",
        "department_name": "基础平台",
        "email": "wangli@example.com",
        "phone_number": "13800008888",
        "account_status": "启用",
    }

    result = store.import_system_users(
        [
            {
                "row_number": 2,
                "username": None,
                "full_name": "王力",
                "role_code": "advisor",
                "department_name": None,
                "introduction": None,
                "email": None,
                "phone_number": None,
                "account_status": "启用",
                "password": None,
            }
        ],
        operator_username="admin",
    )

    assert result.created_count == 1
    assert result.updated_count == 0
    assert result.failed_count == 0
    created_row = next(item for item in store.state["system_users"] if item["username"] == "wangli2")
    assert created_row["role_code"] == "advisor"
    assert PASSWORD_CONTEXT.verify("wangli2@123", created_row["password_hash"])
    assert store.state["profiles"]["wangli2"]["introduction"] is None
    assert store.state["profiles"]["wangli2"]["email"] is None
    assert store.state["profiles"]["wangli2"]["phone_number"] is None


def test_import_system_users_updates_existing_user_and_can_clear_optional_fields(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")
    original_password_hash = target_user.get("password_hash")
    fake_postgres.system_user_rows[int(target_user["id"])] = {
        "id": int(target_user["id"]),
        "username": target_user["username"],
        "full_name": target_user["full_name"],
        "role_code": target_user["role_code"],
        "department_name": target_user.get("department_name") or "智能中心",
        "introduction": "旧介绍",
        "email": "old@example.com",
        "phone_number": "13800001234",
        "account_status": "启用",
    }

    result = store.import_system_users(
        [
            {
                "row_number": 2,
                "username": target_user["username"],
                "full_name": "导入覆盖用户",
                "role_code": target_user["role_code"],
                "department_name": None,
                "introduction": None,
                "email": None,
                "phone_number": None,
                "account_status": "停用",
                "password": None,
            }
        ],
        operator_username="admin",
    )

    assert result.created_count == 0
    assert result.updated_count == 1
    assert result.failed_count == 0
    updated_row = next(item for item in store.state["system_users"] if item["username"] == target_user["username"])
    assert updated_row["full_name"] == "导入覆盖用户"
    assert updated_row["account_status"] == "停用"
    assert updated_row["password_hash"] == original_password_hash
    assert store.state["profiles"][target_user["username"]]["email"] is None
    assert store.state["profiles"][target_user["username"]]["phone_number"] is None


def test_import_system_users_updates_existing_user_password_when_provided(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")
    original_password_hash = target_user.get("password_hash")
    fake_postgres.system_user_rows[int(target_user["id"])] = {
        "id": int(target_user["id"]),
        "username": target_user["username"],
        "full_name": target_user["full_name"],
        "role_code": target_user["role_code"],
        "department_name": target_user.get("department_name") or "智能中心",
        "introduction": "旧介绍",
        "email": "old@example.com",
        "phone_number": "13800001234",
        "account_status": "启用",
    }

    result = store.import_system_users(
        [
            {
                "row_number": 2,
                "username": target_user["username"],
                "full_name": target_user["full_name"],
                "role_code": target_user["role_code"],
                "department_name": target_user.get("department_name"),
                "introduction": "修改密码",
                "email": "reset@example.com",
                "phone_number": "13800001234",
                "account_status": "启用",
                "password": "ResetPass@456",
            }
        ],
        operator_username="admin",
    )

    assert result.created_count == 0
    assert result.updated_count == 1
    assert result.failed_count == 0
    updated_row = next(item for item in store.state["system_users"] if item["username"] == target_user["username"])
    assert updated_row["password_hash"] != original_password_hash
    assert PASSWORD_CONTEXT.verify("ResetPass@456", updated_row["password_hash"])


def test_import_system_users_can_enable_existing_user(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")
    target_user["account_status"] = "停用"
    fake_postgres.system_user_rows[int(target_user["id"])] = {
        "id": int(target_user["id"]),
        "username": target_user["username"],
        "full_name": target_user["full_name"],
        "role_code": target_user["role_code"],
        "department_name": target_user.get("department_name") or "智能中心",
        "introduction": "旧介绍",
        "email": "old@example.com",
        "phone_number": "13800001234",
        "account_status": "停用",
    }

    result = store.import_system_users(
        [
            {
                "row_number": 2,
                "username": target_user["username"],
                "full_name": target_user["full_name"],
                "role_code": target_user["role_code"],
                "department_name": target_user.get("department_name"),
                "introduction": "恢复启用",
                "email": "restore@example.com",
                "phone_number": "13800001234",
                "account_status": "启用",
                "password": None,
            }
        ],
        operator_username="admin",
    )

    assert result.created_count == 0
    assert result.updated_count == 1
    assert result.failed_count == 0
    updated_row = next(item for item in store.state["system_users"] if item["username"] == target_user["username"])
    assert updated_row["account_status"] == "启用"



def test_get_profile_uses_redis_cache_and_update_profile_refreshes_cache(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    fake_postgres.profile_rows = {
        "cache.user": {
            "username": "cache.user",
            "full_name": "缓存用户",
            "role_name": "导师",
            "department_name": "智能制造学院",
            "introduction": "最初介绍",
            "phone_number": "13800009998",
            "email": "cache.user@example.com",
            "theme_color": "#0f4cbd",
        }
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()

    first = store.get_profile("cache.user")
    fake_postgres.profile_rows["cache.user"]["full_name"] = "数据库新名字"
    second = store.get_profile("cache.user")
    updated = store.update_profile(
        "cache.user",
        UserProfileUpdate(full_name="缓存已刷新", phone_number="13800009997", email="cache.user@example.com", theme_color="#123456"),
    )
    third = store.get_profile("cache.user")

    assert first.full_name == "缓存用户"
    assert second.full_name == "缓存用户"
    assert updated.full_name == "缓存已刷新"
    assert third.full_name == "缓存已刷新"


def test_get_profile_recovers_from_cached_null_when_postgres_has_profile(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    expected_profile = {
        "username": "cache.null.user",
        "full_name": "空缓存恢复用户",
        "role_name": "平台管理员",
        "department_name": "研究生院",
        "introduction": None,
        "phone_number": "13800008888",
        "email": "recover@example.com",
        "theme_color": "#0f4cbd",
    }
    call_count = {"value": 0}

    def fake_get_user_profile(username: str):
        call_count["value"] += 1
        return dict(expected_profile) if username == expected_profile["username"] else None

    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()
    fake_postgres.get_user_profile = fake_get_user_profile
    store._write_json_cache(
        store._user_profile_cache_key(expected_profile["username"]),
        {CACHE_NULL_SENTINEL_KEY: True},
        60,
        with_jitter=False,
    )

    profile = store.get_profile(expected_profile["username"])
    cache_hit, cached_payload = store._read_json_cache(store._user_profile_cache_key(expected_profile["username"]))

    assert profile.full_name == expected_profile["full_name"]
    assert call_count["value"] == 1
    assert cache_hit is True
    assert cached_payload == expected_profile


def test_authenticate_system_user_uses_cached_postgres_user_context(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_cache = FakeCacheClient()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)
    monkeypatch.setattr("app.services.management_service.get_cache_client", lambda: fake_cache)

    store = RuntimeManagementStore()
    role_record = store.create_role(
        RoleUpsert(
            role_code="cache_role",
            role_name="缓存角色",
            scope_name="系统治理",
            permissions=["system:write"],
        )
    )
    store.create_system_user(
        SystemUserUpsert(
            username="cache.auth",
            full_name="缓存认证用户",
            role_code=role_record.role_code,
            department_name="信息化办公室",
            email="cache.auth@example.com",
            phone_number="13800006666",
            account_status="启用",
            password="Secret123!",
        )
    )

    first = store.authenticate_system_user("cache.auth", "Secret123!")
    fake_postgres.system_user_rows = {}
    second = store.authenticate_system_user("cache.auth", "Secret123!")

    assert first is not None
    assert second is not None
    assert "system:read" in second["permissions"]


def test_governance_state_hydrates_from_postgres_on_startup(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.role_rows = [
        {
            "id": 11,
            "role_code": "platform_admin",
            "role_name": "平台管理员",
            "scope_name": "系统治理",
            "permissions": ["system:write"],
            "user_count": 1,
        }
    ]
    fake_postgres.system_user_state_rows = [
        {
            "id": 21,
            "username": "prod.admin",
            "full_name": "生产管理员",
            "role_code": "platform_admin",
            "role_name": "平台管理员",
            "department_name": "信息中心",
            "introduction": "负责系统治理与平台运维。",
            "email": "prod.admin@example.com",
            "phone_number": "13800001111",
            "account_status": "启用",
            "last_login_at": None,
            "password_hash": "hashed-password",
        }
    ]
    fake_postgres.profile_rows = {
        "prod.admin": {
            "username": "prod.admin",
            "full_name": "生产管理员",
            "role_name": "平台管理员",
            "department_name": "信息中心",
            "introduction": "负责系统治理与平台运维。",
            "phone_number": "13800001111",
            "email": "prod.admin@example.com",
            "theme_color": "#0f4cbd",
        }
    }
    fake_postgres.audit_policy_rows = [{"id": 31, "item": "登录审计", "policy": "保留 180 天", "status": "启用"}]
    fake_postgres.integration_rows = [{"id": 41, "name": "统一认证", "direction": "双向", "cadence": "实时", "status": "正常", "owner": "信息中心"}]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    assert [role["role_code"] for role in store.state["roles"]] == ["platform_admin"]
    assert [user["username"] for user in store.state["system_users"]] == ["prod.admin"]
    assert store.state["profiles"]["prod.admin"]["email"] == "prod.admin@example.com"
    assert store.state["profiles"]["prod.admin"]["introduction"] == "负责系统治理与平台运维。"
    assert [item["item"] for item in store.state["audit_policies"]] == ["登录审计"]
    assert [item["name"] for item in store.state["integrations"]] == ["统一认证"]
    assert store._counters["roles"] == 11
    assert store._counters["system_users"] == 21


def test_get_system_stats_prefers_postgres_snapshot(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.system_stats_snapshot = {
        "integration_total": 6,
        "active_integration_total": 5,
        "operation_log_total": 120,
        "sync_failure_total": 3,
        "user_total": 18,
        "role_total": 7,
    }
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    stats = store.get_system_stats()

    assert stats.integration_total == 6
    assert stats.operation_log_total == 120
    assert stats.user_total == 18
    assert stats.role_total == 7


def test_update_role_uses_postgres_row_when_runtime_role_missing(monkeypatch) -> None:
    from app.schemas.system import RoleUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_role = store.state["roles"][0]
    role_id = int(target_role["id"])
    fake_postgres.role_rows = [
        {
            "id": role_id,
            "role_code": target_role["role_code"],
            "role_name": target_role["role_name"],
            "scope_name": target_role.get("scope_name") or "系统治理",
            "permissions": [],
            "user_count": 1,
        }
    ]
    store.state["roles"] = [item for item in store.state["roles"] if int(item["id"]) != role_id]

    updated = store.update_role(
        role_id,
        RoleUpsert(
            role_code=target_role["role_code"],
            role_name="平台治理管理员",
            scope_name=target_role.get("scope_name") or "系统治理",
                permissions=[],
        ),
    )

    assert updated.role_name == "平台治理管理员"
    assert fake_postgres.synced_roles[-1]["role_payload"]["id"] == role_id


def test_update_role_rejects_permissions_missing_in_database(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    fake_postgres.missing_permission_codes_result = ["recruitment_plan:read"]
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_role = next(item for item in store.state["roles"] if str(item.get("role_code") or "") == "AILABMGT")

    with pytest.raises(ValueError, match="Permissions not configured in database: recruitment_plan:read"):
        store.update_role(
            int(target_role["id"]),
            RoleUpsert(
                role_code=str(target_role["role_code"]),
                role_name=str(target_role["role_name"]),
                scope_name=str(target_role.get("scope_name") or "招生管理"),
                permissions=["dashboard:read", "recruitment_plan:read"],
            ),
        )


def test_update_audit_policy_uses_postgres_row_when_runtime_policy_missing(monkeypatch) -> None:
    from app.schemas.system import AuditPolicyUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_policy = store.state["audit_policies"][0]
    policy_id = int(target_policy["id"])
    fake_postgres.audit_policy_rows = [dict(target_policy)]
    store.state["audit_policies"] = [item for item in store.state["audit_policies"] if int(item["id"]) != policy_id]

    updated = store.update_audit_policy(
        policy_id,
        AuditPolicyUpsert(item=target_policy["item"], policy="保留 365 天", status="启用"),
    )

    assert updated.policy == "保留 365 天"
    assert fake_postgres.saved_states[-1]["operation_log"]["entity_name"] == "审计策略"


def test_update_integration_uses_postgres_row_when_runtime_item_missing(monkeypatch) -> None:
    from app.schemas.system import IntegrationUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_item = store.state["integrations"][0]
    integration_id = int(target_item["id"])
    fake_postgres.integration_rows = [dict(target_item)]
    store.state["integrations"] = [item for item in store.state["integrations"] if int(item["id"]) != integration_id]

    updated = store.update_integration(
        integration_id,
        IntegrationUpsert(
            name=target_item["name"],
            direction=target_item["direction"],
            cadence="每小时",
            status=target_item["status"],
            owner=target_item["owner"],
        ),
    )

    assert updated.cadence == "每小时"
    assert fake_postgres.saved_states[-1]["operation_log"]["entity_name"] == "集成链路"


def test_create_system_user_failure_writes_operation_log(monkeypatch) -> None:
    from app.schemas.system import SystemUserUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    with pytest.raises(ValueError, match="Username already exists"):
        store.create_system_user(
            SystemUserUpsert(
                username="admin",
                full_name="重复账号",
                role_code="admin",
                department_name="信息中心",
                email="duplicate@example.com",
                phone_number="13800001111",
                account_status="启用",
                password="Secret123!",
            )
        )

    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["entity_name"] == "系统用户"
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["action"] == "新建账号"
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["result"] == "failed"
    assert "Username already exists" in fake_postgres.synced_operation_logs[-1]["operation_log"]["summary"]


def test_update_system_user_failure_writes_operation_log(monkeypatch) -> None:
    from app.schemas.system import SystemUserUpsert

    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()
    target_user = next(item for item in store.state["system_users"] if item["username"] != "admin")

    with pytest.raises(ValueError, match="Username already exists"):
        store.update_system_user(
            int(target_user["id"]),
            SystemUserUpsert(
                username="admin",
                full_name=target_user["full_name"],
                role_code=target_user["role_code"],
                department_name=target_user.get("department_name") or "",
                email="user@example.com",
                phone_number="13800001111",
                account_status=target_user["account_status"],
                password=None,
            ),
        )

    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["entity_name"] == "系统用户"
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["action"] == "维护账号"
    assert fake_postgres.synced_operation_logs[-1]["operation_log"]["result"] == "failed"
    assert "Username already exists" in fake_postgres.synced_operation_logs[-1]["operation_log"]["summary"]


def test_portal_application_upsert_rejects_missing_basic_required_fields() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["profile"]["full_name_pinyin"] = ""
    payload["profile"]["emergency_contact_name"] = None

    with pytest.raises(ValidationError, match="基本信息以下字段必填：姓名拼音、紧急联系人姓名"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_rejects_invalid_emergency_contact_phone() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["profile"]["emergency_contact_phone"] = "12345"

    with pytest.raises(ValidationError, match="紧急联系人手机格式不正确"):
        PortalApplicationUpsert.model_validate(payload)


def test_build_portal_student_record_tolerates_invalid_legacy_emergency_contact_phone(monkeypatch) -> None:
    fake_postgres = FakePostgresStateStore()
    monkeypatch.setattr("app.services.management_service.PostgresStateStore", lambda: fake_postgres)

    store = RuntimeManagementStore()

    record = store._build_portal_student_record(
        {
            "id": 7,
            "full_name": "历史数据学生",
            "phone_number": "13800001111",
            "email": "legacy-portal@example.com",
            "id_number": "32000019990101123X",
            "account_status": "启用",
            "profile": {
                "full_name_pinyin": "LI SHI",
                "emergency_contact_name": "家长",
                "emergency_contact_phone": "1",
            },
        }
    )

    assert record.profile is not None
    assert record.profile.emergency_contact_phone is None


def test_portal_application_upsert_requires_source_channel_other_when_needed() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["source_channel"] = "其他"
    payload["source_channel_other"] = ""

    with pytest.raises(ValidationError, match="选择“其他”来源时，请补充说明"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_requires_source_channel() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["source_channel"] = ""

    with pytest.raises(ValidationError, match="请选择了解项目方式"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_rejects_legacy_source_channel_value() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["source_channel"] = "实验室官网"

    with pytest.raises(ValidationError, match="请选择了解项目方式"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_requires_primary_preference_advisor() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["preferences"] = [
        {
            "preference_order": 1,
            "research_center_name": "智能制造联合团队",
            "team_id": 1,
            "advisor_name": "",
            "advisor_user_id": None,
            "is_optional": False,
        }
    ]

    with pytest.raises(ValidationError, match="请选择第一志愿导师"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_requires_submission_declaration() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["signed_agreement"] = False
    payload["declaration"]["has_read_declaration"] = False

    with pytest.raises(ValidationError, match="请先确认提交声明"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_rejects_missing_english_certificate_attachment() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["english_proficiencies"][0]["certificate_attachment_url"] = ""

    with pytest.raises(ValidationError, match="英语能力1必须上传英语证明附件"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_upsert_rejects_personal_statement_over_1200_chars() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["personal_statement"]["personal_statement_text"] = "甲" * 1201

    with pytest.raises(ValidationError, match="个人陈述需控制在 1200 字以内"):
        PortalApplicationUpsert.model_validate(payload)


def test_portal_application_draft_upsert_rejects_missing_required_fields() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["profile"]["full_name_pinyin"] = ""

    with pytest.raises(ValidationError, match="基本信息以下字段必填：姓名拼音"):
        PortalApplicationDraftUpsert.model_validate(payload)


def test_portal_application_draft_upsert_validates_only_current_section_when_scope_provided() -> None:
    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "basic-section"
    payload["profile"]["full_name_pinyin"] = "测试学生"
    payload["signed_agreement"] = False
    payload["declaration"]["has_read_declaration"] = False
    payload["english_proficiencies"] = []

    validated = PortalApplicationDraftUpsert.model_validate(payload)

    assert validated.validation_section_id == "basic-section"


def test_portal_application_draft_upsert_treats_blank_preference_ids_as_unselected() -> None:
    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "application-section"
    payload["preferences"] = [
        {
            "preference_order": 1,
            "research_center_name": "智能制造联合团队",
            "team_id": "",
            "advisor_name": "刘亚",
            "advisor_user_id": "",
            "is_optional": False,
        }
    ]
    payload["selected_team_id"] = ""
    payload["selected_advisor_user_id"] = ""

    validated = PortalApplicationDraftUpsert.model_validate(payload)

    assert validated.preferences[0].team_id is None
    assert validated.preferences[0].advisor_user_id is None
    assert validated.selected_team_id is None
    assert validated.selected_advisor_user_id is None


def test_portal_application_draft_upsert_rejects_legacy_source_channel_value_for_application_section() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "application-section"
    payload["source_channel"] = "实验室官网"

    with pytest.raises(ValidationError, match="请选择了解项目方式"):
        PortalApplicationDraftUpsert.model_validate(payload)


def test_portal_application_draft_upsert_allows_blank_second_preference_without_team() -> None:
    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "application-section"
    payload["preferences"] = [
        {
            "preference_order": 1,
            "research_center_name": "智能制造联合团队",
            "team_id": 1,
            "advisor_name": "刘亚",
            "advisor_user_id": 11,
            "is_optional": False,
        },
        {
            "preference_order": 2,
            "advisor_name": "",
            "advisor_user_id": None,
            "is_optional": True,
        },
    ]

    validated = PortalApplicationDraftUpsert.model_validate(payload)

    assert validated.preferences[1].advisor_name in {"", None}


def test_portal_application_draft_upsert_allows_current_bachelor_section_without_degree_and_graduation_attachments() -> None:
    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "education-section"
    payload["education_experiences"] = [
        {
            "sort_order": 1,
            "education_stage": "高中毕业",
            "start_month": "2016-09",
            "end_month": "2019-06",
            "school_name": "无锡市第一中学",
            "verifier_name": "王老师",
            "verifier_phone": "13800002222",
        },
        {
            "sort_order": 2,
            "education_stage": "本科在读",
            "start_month": "2019-09",
            "school_name": "江南大学",
            "major_name": "自动化",
            "average_score": "89",
            "gpa": "3.8",
            "ranking": "12/120",
            "transcript_attachment_url": "/portal-attachments/uploads/student-7/education_transcript/transcript-a.pdf",
            "degree_certificate_attachment_url": "",
            "graduation_certificate_attachment_url": "",
            "verifier_name": "李老师",
            "verifier_phone": "13800003333",
        },
    ]

    validated = PortalApplicationDraftUpsert.model_validate(payload)

    assert validated.validation_section_id == "education-section"


def test_portal_application_draft_upsert_rejects_missing_degree_and_graduation_attachments_for_graduated_bachelor_section() -> None:
    from pydantic import ValidationError

    from app.schemas.portal import PortalApplicationDraftUpsert

    payload = _build_portal_application_payload(1, "智能制造联合团队", "刘亚").model_dump(mode="python")
    payload["validation_section_id"] = "education-section"
    payload["education_experiences"] = [
        {
            "sort_order": 1,
            "education_stage": "高中毕业",
            "start_month": "2016-09",
            "end_month": "2019-06",
            "school_name": "无锡市第一中学",
            "verifier_name": "王老师",
            "verifier_phone": "13800002222",
        },
        {
            "sort_order": 2,
            "education_stage": "本科毕业",
            "start_month": "2019-09",
            "end_month": "2023-06",
            "school_name": "江南大学",
            "major_name": "自动化",
            "average_score": "89",
            "gpa": "3.8",
            "ranking": "12/120",
            "transcript_attachment_url": "/portal-attachments/uploads/student-7/education_transcript/transcript-a.pdf",
            "degree_certificate_attachment_url": "",
            "graduation_certificate_attachment_url": "",
            "verifier_name": "李老师",
            "verifier_phone": "13800003333",
        },
        {
            "sort_order": 3,
            "education_stage": "硕士在读",
            "start_month": "2023-09",
            "school_name": "江南大学",
            "major_name": "控制科学与工程",
            "average_score": "91",
            "gpa": "3.9",
            "ranking": "8/60",
            "transcript_attachment_url": "/portal-attachments/uploads/student-7/education_transcript/transcript-b.pdf",
            "verifier_name": "周老师",
            "verifier_phone": "13800005555",
        },
    ]

    with pytest.raises(ValidationError, match="教育经历2以下字段必填：学位证附件、毕业证附件"):
        PortalApplicationDraftUpsert.model_validate(payload)
