from io import BytesIO

from openpyxl import load_workbook

from app.services.recruitment_excel_service import (
    build_recruitment_template,
    build_registered_portal_students_template,
    parse_recruitment_template,
)


def test_parse_recruitment_template_reads_template_rows() -> None:
    content = build_recruitment_template(
        [
            {
                "review_round": "第一轮",
                "student_name": "张三",
                "first_choice": "人工智能",
                "second_choice": "计算机视觉",
                "gender": "男",
                "phone_number": "13900001111",
                "email": "zhangsan@example.com",
                "graduation_school": "复旦大学",
                "material_status": "材料齐全",
            }
        ]
    )

    rows = parse_recruitment_template(content)

    assert len(rows) == 1
    assert rows[0]["student_name"] == "张三"
    assert rows[0]["first_choice"] == "人工智能"
    assert rows[0]["intended_field"] == "人工智能"
    assert rows[0]["undergraduate_school"] == "复旦大学"


def test_build_recruitment_template_preserves_duplicate_profile_columns() -> None:
    content = build_recruitment_template(
        [
            {
                "student_name": "李四",
                "personal_statement_text": "第一段个人简介",
                "supplementary_profile": "第二段个人简介",
            }
        ]
    )

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    values = [cell.value for cell in worksheet[2]]
    profile_indexes = [index for index, header in enumerate(headers) if header == "个人简介"]

    assert len(profile_indexes) == 2
    assert values[profile_indexes[0]] == "第一段个人简介"
    assert values[profile_indexes[1]] == "第二段个人简介"


def test_build_registered_portal_students_template_strips_illegal_excel_characters() -> None:
    content = build_registered_portal_students_template(
        [
            {
                "full_name": "测试学生",
                "achievement_1_paper_title": "FED-DUET:\x0b DUAL EXPERT-ORCHESTRATED FRAMEWORK FOR CONTINUAL FEDERATED VISION LANGUAGE LEARNING",
            }
        ]
    )

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    values = [cell.value for cell in worksheet[2]]
    exported_row = dict(zip(headers, values, strict=False))

    assert exported_row["姓名"] == "测试学生"
    assert exported_row["科研成果1论文标题"] == "FED-DUET: DUAL EXPERT-ORCHESTRATED FRAMEWORK FOR CONTINUAL FEDERATED VISION LANGUAGE LEARNING"


def test_build_registered_portal_students_template_strips_xml_invalid_noncharacters() -> None:
    content = build_registered_portal_students_template(
        [
            {
                "full_name": "测试学生",
                "personal_profile": "含有非法字符\uffff的简介",
            }
        ]
    )

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    values = [cell.value for cell in worksheet[2]]
    exported_row = dict(zip(headers, values, strict=False))

    assert exported_row["姓名"] == "测试学生"
    assert exported_row["个人简介"] == "含有非法字符的简介"


def test_build_registered_portal_students_template_uses_max_repeated_group_count() -> None:
    content = build_registered_portal_students_template(
        [
            {
                "full_name": "最多列学生",
                "preference_1_advisor_name": "导师甲",
                "preference_2_advisor_name": "导师乙",
                "education_1_stage": "高中毕业",
                "education_4_stage": "博士在读",
            },
            {
                "full_name": "较少列学生",
                "preference_1_advisor_name": "导师丙",
                "education_1_stage": "本科毕业",
            },
        ]
    )

    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    first_row = dict(zip(headers, [cell.value for cell in worksheet[2]], strict=False))
    second_row = dict(zip(headers, [cell.value for cell in worksheet[3]], strict=False))

    assert "志愿2导师" in headers
    assert "教育经历4教育阶段" in headers
    assert "志愿3导师" not in headers
    assert first_row["志愿2导师"] == "导师乙"
    assert second_row["志愿2导师"] is None
    assert first_row["教育经历4教育阶段"] == "博士在读"
    assert second_row["教育经历4教育阶段"] is None