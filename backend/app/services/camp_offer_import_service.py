from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

import psycopg
from openpyxl import load_workbook

from app.core.config import settings
from app.schemas.recruitment import (
    AdmissionOfferedSchoolImportIssue,
    AdmissionOfferedSchoolImportResult,
    CampOfferImportIssue,
    CampOfferImportResult,
    HackathonScoreImportIssue,
    HackathonScoreImportResult,
    IsInCampSelectionImportIssue,
    IsInCampSelectionImportResult,
)
from app.services.recruitment_excel_service import (
    parse_admission_offered_school_template,
    parse_hackathon_score_template,
    parse_is_in_camp_selection_template,
)


def _conninfo() -> str:
    return (
        f"host={settings.postgres_host} "
        f"port={settings.postgres_port} "
        f"dbname={settings.postgres_db} "
        f"user={settings.postgres_user} "
        f"password={settings.postgres_password}"
    )


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    return re.sub(r"[\s\-_—–()（）\[\]【】{}<>《》:：,，.。/\\]+", "", text)


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def _parse_bool(value: Any) -> bool | None:
    normalized = _normalize_cell(value)
    if normalized is None:
        return None
    if isinstance(normalized, bool):
        return normalized
    if isinstance(normalized, (int, float)):
        if normalized == 1:
            return True
        if normalized == 0:
            return False
        return None
    text = str(normalized).strip().lower()
    if text in {"true", "1", "yes", "y", "是", "同意"}:
        return True
    if text in {"false", "0", "no", "n", "否", "不同意"}:
        return False
    return None


def _parse_datetime(value: Any) -> datetime | date | None:
    value = _normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        text = value.replace("T", " ").strip()
        for parser in (
            datetime.fromisoformat,
            lambda item: datetime.strptime(item, "%Y-%m-%d %H:%M:%S"),
            lambda item: datetime.strptime(item, "%Y-%m-%d %H:%M"),
            lambda item: datetime.strptime(item, "%Y/%m/%d %H:%M:%S"),
            lambda item: datetime.strptime(item, "%Y/%m/%d %H:%M"),
            lambda item: datetime.strptime(item, "%Y-%m-%d"),
        ):
            try:
                parsed = parser(text)
            except Exception:
                continue
            if isinstance(parsed, datetime):
                return parsed
            return datetime.combine(parsed, datetime.min.time())
    return None


def _resolve_header_field(header: str) -> str | None:
    if not header:
        return None
    if header == "candidate_no" or ("报名号" in header and "candidate_no" in header):
        return "candidate_no"
    if header in {"报名号", "candidateno"}:
        return "candidate_no"
    if header in {"planid", "plan_id", "计划id"}:
        return "plan_id"
    if header in {"isagree", "is_agree", "是否同意"}:
        return "is_agree"
    if header in {"reason", "原因", "reson"}:
        return "reason"
    if header in {"issentmail", "is_sent_mail", "是否已发邮件"}:
        return "is_sent_mail"
    if header in {"studentoffersubmittedat", "student_offer_submitted_at", "学生提交日期", "submited"}:
        return "student_offer_submitted_at"
    return None


def _load_import_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return []
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    normalized_header_row = [_normalize_header(value) for value in rows[0]]
    header_index: dict[str, int] = {}
    for index, header in enumerate(normalized_header_row):
        field_name = _resolve_header_field(header)
        if field_name:
            header_index[field_name] = index

    if "candidate_no" not in header_index:
        raise ValueError("导入文件必须包含报名号（candidate_no）列")

    result: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue
        item: dict[str, Any] = {"row_number": row_number}
        for field_name, index in header_index.items():
            item[field_name] = _normalize_cell(row[index] if index < len(row) else None)
        result.append(item)
    return result


def _resolve_plan_id(
    raw_plan_id: Any,
    fallback_plan_id: int | None,
    latest_plan_id: int | None,
) -> int | None:
    text = _normalize_cell(raw_plan_id)
    if text not in (None, ""):
        try:
            return int(text)
        except (TypeError, ValueError):
            return None
    if fallback_plan_id is not None:
        return int(fallback_plan_id)
    if latest_plan_id is not None:
        return int(latest_plan_id)
    return None


def _load_latest_plan_id(cur: Any) -> int | None:
    cur.execute(
        """
        SELECT id
        FROM dtlms_recruitment_plans
        ORDER BY id DESC
        LIMIT 1
        """
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def _load_existing_candidates(cur: Any, candidate_nos: list[str]) -> set[str]:
    if not candidate_nos:
        return set()
    cur.execute(
        """
        SELECT candidate_no
        FROM dtlms_recruitment_applications
        WHERE is_deleted = FALSE
          AND candidate_no = ANY(%s)
        """,
        (candidate_nos,),
    )
    return {str(row[0]).strip() for row in cur.fetchall() if row and row[0]}


def _load_existing_plan_ids(cur: Any, plan_ids: list[int]) -> set[int]:
    if not plan_ids:
        return set()
    cur.execute(
        """
        SELECT id
        FROM dtlms_recruitment_plans
        WHERE id = ANY(%s)
        """,
        (plan_ids,),
    )
    return {int(row[0]) for row in cur.fetchall() if row and row[0] is not None}


def _load_existing_offer_keys(cur: Any, candidate_nos: list[str], plan_ids: list[int]) -> set[tuple[str, int]]:
    if not candidate_nos or not plan_ids:
        return set()
    cur.execute(
        """
        SELECT candidate_no, plan_id
        FROM dtlms_plan_offer
        WHERE candidate_no = ANY(%s)
          AND plan_id = ANY(%s)
        """,
        (candidate_nos, plan_ids),
    )
    return {(str(row[0]).strip(), int(row[1])) for row in cur.fetchall() if row and row[0] is not None and row[1] is not None}


def import_camp_offers_from_excel(file_bytes: bytes, *, plan_id: int | None = None) -> CampOfferImportResult:
    rows = _load_import_rows(file_bytes)
    if not rows:
        return CampOfferImportResult(
            imported_count=0,
            skipped_count=0,
            plan_id=int(plan_id or 0),
            imported_ids=[],
            issues=[],
        )

    issues: list[CampOfferImportIssue] = []
    latest_plan_id = int(plan_id or 0)
    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cur:
            latest_plan_id = _load_latest_plan_id(cur)
            prepared_rows: list[dict[str, Any]] = []
            candidate_nos: list[str] = []
            plan_ids: list[int] = []
            seen_keys: set[tuple[str, int]] = set()

            for row in rows:
                row_number = int(row["row_number"])
                candidate_no = str(row.get("candidate_no") or "").strip()
                if not candidate_no:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            reason="报名号不能为空",
                        )
                    )
                    continue

                resolved_plan_id = _resolve_plan_id(row.get("plan_id"), plan_id, latest_plan_id)
                if resolved_plan_id is None:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="未找到可用的计划编号",
                        )
                    )
                    continue

                is_sent_mail = _parse_bool(row.get("is_sent_mail"))
                if row.get("is_sent_mail") not in (None, "") and is_sent_mail is None:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="是否已发邮件格式无法识别",
                        )
                    )
                    continue

                is_agree = _parse_bool(row.get("is_agree"))
                if row.get("is_agree") not in (None, "") and is_agree is None:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="是否同意格式无法识别",
                        )
                    )
                    continue

                submitted_at = _parse_datetime(row.get("student_offer_submitted_at"))
                if row.get("student_offer_submitted_at") not in (None, "") and submitted_at is None:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="学生提交日期格式无法识别",
                        )
                    )
                    continue

                key = (candidate_no, int(resolved_plan_id))
                if key in seen_keys:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="文件中重复的报名号和计划编号组合",
                        )
                    )
                    continue
                seen_keys.add(key)
                prepared_rows.append(
                    {
                        "row_number": row_number,
                        "candidate_no": candidate_no,
                        "plan_id": int(resolved_plan_id),
                        "is_sent_mail": is_sent_mail,
                        "is_agree": is_agree,
                        "reason": _normalize_cell(row.get("reason")),
                        "student_offer_submitted_at": submitted_at,
                    }
                )
                candidate_nos.append(candidate_no)
                plan_ids.append(int(resolved_plan_id))

            if not prepared_rows:
                return CampOfferImportResult(
                    imported_count=0,
                    skipped_count=len(rows),
                    plan_id=int(plan_id or latest_plan_id or 0),
                    imported_ids=[],
                    issues=issues,
                )

            existing_candidates = _load_existing_candidates(cur, candidate_nos)
            existing_plan_ids = _load_existing_plan_ids(cur, sorted(set(plan_ids)))
            existing_offer_keys = _load_existing_offer_keys(cur, candidate_nos, sorted(set(plan_ids)))

            valid_rows: list[dict[str, Any]] = []
            for row in prepared_rows:
                row_number = int(row["row_number"])
                candidate_no = str(row["candidate_no"])
                resolved_plan_id = int(row["plan_id"])
                key = (candidate_no, resolved_plan_id)
                if resolved_plan_id not in existing_plan_ids:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason=f"计划编号 {resolved_plan_id} 不存在",
                        )
                    )
                    continue
                if candidate_no not in existing_candidates:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="报名号在招生报名表中不存在",
                        )
                    )
                    continue
                if key in existing_offer_keys:
                    issues.append(
                        CampOfferImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            reason="该报名号和计划编号的入营名单已存在",
                        )
                    )
                    continue
                valid_rows.append(row)

            if not valid_rows:
                conn.commit()
                return CampOfferImportResult(
                    imported_count=0,
                    skipped_count=len(rows),
                    plan_id=int(plan_id or latest_plan_id or 0),
                    imported_ids=[],
                    issues=issues,
                )

            cur.execute(
                """
                INSERT INTO dtlms_plan_offer (
                    candidate_no,
                    plan_id,
                    is_sent_mail,
                    is_agree,
                    reson,
                    submitted_at,
                    created_at,
                    updated_at
                )
                SELECT
                    data.candidate_no,
                    data.plan_id,
                    data.is_sent_mail,
                    data.is_agree,
                    data.reson,
                    data.submitted_at,
                    CURRENT_TIMESTAMP,
                    CURRENT_TIMESTAMP
                FROM UNNEST(
                    %s::text[],
                    %s::bigint[],
                    %s::boolean[],
                    %s::boolean[],
                    %s::text[],
                    %s::timestamptz[]
                ) AS data(candidate_no, plan_id, is_sent_mail, is_agree, reson, submitted_at)
                RETURNING id
                """,
                (
                    [str(row["candidate_no"]) for row in valid_rows],
                    [int(row["plan_id"]) for row in valid_rows],
                    [bool(row["is_sent_mail"]) for row in valid_rows],
                    [row["is_agree"] for row in valid_rows],
                    [row["reason"] for row in valid_rows],
                    [row["student_offer_submitted_at"] for row in valid_rows],
                ),
            )
            imported_ids = [int(item[0]) for item in cur.fetchall()]
            conn.commit()

    skipped_count = len(rows) - len(imported_ids)
    return CampOfferImportResult(
        imported_count=len(imported_ids),
        skipped_count=skipped_count,
        plan_id=int(plan_id or latest_plan_id or 0),
        imported_ids=imported_ids,
        issues=issues,
    )



# --------------------------------------------------------------------------
# 2026-07-03: 黑客松夏令营「评分导入」专用入口
# 业务规则(Q1~Q4 全部确认 A):
#   Q1 匹配规则:  手机号 AND 邮箱 同时相等 (联合匹配)
#   Q2 数据源:    dtlms_recruitment_applications.student_phone / student_email
#   Q3 异常策略:  匹配不到入营名单 -> 跳过, 在 issues 中报告(不抛错)
#   Q4 重复策略:  无条件覆盖 hackathon_score / hackathon_comments
# 仅更新这 2 个字段 + updated_at, 不动其他列(避免误覆盖业务流转状态)。
# --------------------------------------------------------------------------
def import_hackathon_scores_from_excel(file_bytes: bytes) -> HackathonScoreImportResult:
    """解析 + 导入 黑客松夏令营评分 Excel。

    入参: file_bytes (前端 multipart/form-data 上传的 .xlsx 字节)
    出参: HackathonScoreImportResult
    """
    rows = parse_hackathon_score_template(file_bytes)
    total_rows = len(rows)
    if total_rows == 0:
        return HackathonScoreImportResult(
            total_rows=0,
            matched_count=0,
            unmatched_count=0,
            updated_ids=[],
            issues=[],
        )

    issues: list[HackathonScoreImportIssue] = []
    updated_ids: list[int] = []

    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cur:
            # 1) 逐行匹配 + UPDATE; 单行失败不影响其他行 (Q3 跳过策略)
            for row in rows:
                row_number = int(row["row_number"])
                phone = row.get("phone")
                email = row.get("email")
                score = row.get("hackathon_score")
                comment = row.get("hackathon_comments")

                # 行内校验
                if not phone or not email:
                    issues.append(
                        HackathonScoreImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            reason="学生手机号/邮箱不能为空",
                        )
                    )
                    continue
                if score is None:
                    issues.append(
                        HackathonScoreImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            reason="夏令营评分不能为空或格式无法识别",
                        )
                    )
                    continue

                # 2) 联合匹配: 手机号 AND 邮箱 同时相等 (Q1: A)
                #    dtlms_plan_offer 与 dtlms_recruitment_applications 通过 candidate_no 关联
                #    实际字段: dtlms_recruitment_applications.phone_number / email
                #    一次 UPDATE, 不影响其他字段 (Q4: 无条件覆盖)
                cur.execute(
                    """
                    UPDATE dtlms_plan_offer offer
                    SET hackathon_score = %s,
                        hackathon_comments = %s,
                        updated_at = CURRENT_TIMESTAMP
                    FROM dtlms_recruitment_applications app
                    WHERE offer.candidate_no = app.candidate_no
                      AND offer.plan_id = app.plan_id
                      AND app.phone_number = %s
                      AND app.email = %s
                    RETURNING offer.id
                    """,
                    (score, comment, phone, email),
                )
                results = cur.fetchall()
                if not results:
                    # Q3: 匹配不到 -> 跳过, 在 issues 中报告
                    issues.append(
                        HackathonScoreImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            reason="未匹配到入营名单记录(手机号+邮箱联合查询无结果)",
                        )
                    )
                    continue
                # 一位学生若在多个 plan 下有入营名单, 全部更新
                for r in results:
                    updated_ids.append(int(r[0]))

            conn.commit()

    # matched: 至少有一行被 UPDATE 成功的原始 Excel 行(去重)
    # updated_ids 中可能有多条 (一个学生在多个 plan 下有入营名单), 用行号去重更合理
    matched_row_count = total_rows - sum(
        1 for it in issues if it.reason == "未匹配到入营名单记录(手机号+邮箱联合查询无结果)"
    )
    return HackathonScoreImportResult(
        total_rows=total_rows,
        matched_count=matched_row_count,
        unmatched_count=total_rows - matched_row_count,
        updated_ids=updated_ids,
        issues=issues,
    )


# --------------------------------------------------------------------------
# 2026-07-06: 黑客松夏令营 “上传录取学校” 专用导入
# 区别于 /camp-offers/import-hackathon-scores:
#   - 通过 dtlms_recruitment_applications.phone_number + email 联合匹配入营名单
#   - 仅更新 dtlms_plan_offer.admission_offered_school (varchar(64))
#   - 匹配不到入营名单的行跳过, 在 issues 中报告 (Q3: 不报错)
def import_admission_offered_schools_from_excel(file_bytes: bytes) -> AdmissionOfferedSchoolImportResult:
    """解析 + 导入 录取学校 Excel.

    入参: file_bytes (前端 multipart/form-data 上传的 .xlsx 字节)
    出参: AdmissionOfferedSchoolImportResult
    """
    rows = parse_admission_offered_school_template(file_bytes)
    total_rows = len(rows)
    if total_rows == 0:
        return AdmissionOfferedSchoolImportResult(
            total_rows=0,
            matched_count=0,
            unmatched_count=0,
            updated_ids=[],
            issues=[],
        )

    issues: list[AdmissionOfferedSchoolImportIssue] = []
    updated_ids: list[int] = []

    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cur:
            for row in rows:
                row_number = int(row["row_number"])
                phone = row.get("phone")
                email = row.get("email")
                school = row.get("admission_offered_school")

                if not phone or not email:
                    issues.append(
                        AdmissionOfferedSchoolImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            school=school,
                            reason="学生手机号/邮箱不能为空",
                        )
                    )
                    continue
                if school is None:
                    issues.append(
                        AdmissionOfferedSchoolImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            school=school,
                            reason="录取学校不能为空",
                        )
                    )
                    continue

                cur.execute(
                    """
                    UPDATE dtlms_plan_offer offer
                    SET admission_offered_school = %s,
                        updated_at = CURRENT_TIMESTAMP
                    FROM dtlms_recruitment_applications app
                    WHERE offer.candidate_no = app.candidate_no
                      AND offer.plan_id = app.plan_id
                      AND app.phone_number = %s
                      AND app.email = %s
                    RETURNING offer.id
                    """,
                    (school, phone, email),
                )
                results = cur.fetchall()
                if not results:
                    issues.append(
                        AdmissionOfferedSchoolImportIssue(
                            row_number=row_number,
                            phone=phone,
                            email=email,
                            school=school,
                            reason="未匹配到入营名单记录(手机号+邮箱联合查询无结果)",
                        )
                    )
                    continue
                for r in results:
                    updated_ids.append(int(r[0]))

            conn.commit()

    unmatched_reasons = {
        "学生手机号/邮箱不能为空",
        "录取学校不能为空",
        "未匹配到入营名单记录(手机号+邮箱联合查询无结果)",
    }
    matched_row_count = total_rows - sum(1 for it in issues if it.reason in unmatched_reasons)
    return AdmissionOfferedSchoolImportResult(
        total_rows=total_rows,
        matched_count=matched_row_count,
        unmatched_count=total_rows - matched_row_count,
        updated_ids=updated_ids,
        issues=issues,
    )


# --------------------------------------------------------------------------
# 2026-07-06: 黑客松夏令营 “导入夏令营选拔的学生” 专用导入
# 区别于 /camp-offers/import-admission-offered-schools:
#   - 通过 dtlms_plan_offer.candidate_no (报名号) 匹配入营名单
#   - 仅更新 dtlms_plan_offer.is_in_camp_selection (boolean)
#   - 表头: 报名号 / 夏令营选拔 (内容: 是 / 否)
#   - 匹配不到入营名单的行跳过, 在 issues 中报告 (Q3: 不报错)
#
# 允许的“夏令营选拔”值 (大小写不敏感):
#   - True  (是 / yes / y / true / 1)
#   - False (否 / no  / n / false / 0)
#   - 其他 -> 记录 issue (reason=夏令营选拔值无法识别)
_TRUE_TOKENS = {"是", "yes", "y", "true", "1"}
_FALSE_TOKENS = {"否", "no", "n", "false", "0"}


def _parse_selection_value(raw: str | None) -> bool | None:
    """将表格中的“夏令营选拔”文本转为 bool。

    返回:
        - True / False: 识别成功
        - None:         文本为空 / 无法识别 (service 会记录 issue)
    """
    if raw is None:
        return None
    token = str(raw).strip().lower()
    if not token:
        return None
    if token in _TRUE_TOKENS:
        return True
    if token in _FALSE_TOKENS:
        return False
    return None


def import_is_in_camp_selection_from_excel(file_bytes: bytes) -> IsInCampSelectionImportResult:
    """解析 + 导入 夏令营选拔 Excel。

    入参: file_bytes (前端 multipart/form-data 上传的 .xlsx 字节)
    出参: IsInCampSelectionImportResult
    """
    rows = parse_is_in_camp_selection_template(file_bytes)
    total_rows = len(rows)
    if total_rows == 0:
        return IsInCampSelectionImportResult(
            total_rows=0,
            matched_count=0,
            unmatched_count=0,
            updated_ids=[],
            issues=[]
        )

    issues: list[IsInCampSelectionImportIssue] = []
    updated_ids: list[int] = []

    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cur:
            for row in rows:
                row_number = int(row["row_number"])
                candidate_no = row.get("candidate_no")
                raw_value = row.get("is_in_camp_selection_raw")

                if not candidate_no:
                    issues.append(
                        IsInCampSelectionImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            raw_value=raw_value,
                            reason="报名号不能为空"
                        )
                    )
                    continue

                selection = _parse_selection_value(raw_value)
                if selection is None:
                    issues.append(
                        IsInCampSelectionImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            raw_value=raw_value,
                            reason="夏令营选拔值无法识别(期待: 是 / 否 / yes / no / true / false / 1 / 0)"
                        )
                    )
                    continue

                cur.execute(
                    """
                    UPDATE dtlms_plan_offer
                    SET is_in_camp_selection = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE candidate_no = %s
                    RETURNING id
                    """
                    , (selection, candidate_no))
                results = cur.fetchall()
                if not results:
                    issues.append(
                        IsInCampSelectionImportIssue(
                            row_number=row_number,
                            candidate_no=candidate_no,
                            raw_value=raw_value,
                            reason="未匹配到入营名单记录(报名号查询无结果)"
                        )
                    )
                    continue
                for r in results:
                    updated_ids.append(int(r[0]))

            conn.commit()

    unmatched_reasons = {
        "报名号不能为空",
        "夏令营选拔值无法识别(期待: 是 / 否 / yes / no / true / false / 1 / 0)",
        "未匹配到入营名单记录(报名号查询无结果)",
    }
    matched_row_count = total_rows - sum(1 for it in issues if it.reason in unmatched_reasons)
    return IsInCampSelectionImportResult(
        total_rows=total_rows,
        matched_count=matched_row_count,
        unmatched_count=total_rows - matched_row_count,
        updated_ids=updated_ids,
        issues=issues,
    )
