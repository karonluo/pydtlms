from __future__ import annotations

from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.worksheet import Worksheet


RECRUITMENT_TEMPLATE_COLUMNS: list[tuple[str, str]] = [
    ("review_round", "轮次"),
    ("student_name", "全名"),
    ("first_choice", "第一志愿"),
    ("second_choice", "第二志愿"),
    ("gender", "性别"),
    ("political_status", "政治面貌"),
    ("marital_status", "婚否"),
    ("religious_belief", "宗教信仰"),
    ("native_place", "籍贯"),
    ("phone_number", "电话"),
    ("email", "邮箱"),
    ("mailing_address", "联系地址"),
    ("id_type", "证件类型"),
    ("id_number", "证件号码"),
    ("material_status", "资料审核"),
    ("graduation_school", "本科院校"),
    ("accept_adjustment", "是否接受调剂"),
    ("undergraduate_average_score", "本科期间平均成绩"),
    ("undergraduate_gpa", "本科期间绩点"),
    ("undergraduate_rank", "本科平均成绩排名"),
    ("undergraduate_major", "本科专业"),
    ("graduate_average_score", "硕士期间平均成绩"),
    ("graduate_gpa", "硕士期间绩点"),
    ("graduate_rank", "硕士平均成绩排名"),
    ("graduate_major", "硕士专业"),
    ("intended_advisor_name", "意向导师姓名"),
    ("discovery_channel", "了解上海人工智能实验室联合培养博士生的方式"),
    ("graduate_school", "硕士院校"),
    ("overseas_university_name", "就读境外大学名称"),
    ("overseas_master_university_name", "就读境外硕士大学名称"),
    ("self_evaluation", "本人自我评价"),
    ("applied_at", "申请时间"),
    ("research_problem", "你认为目前AI技术发展过程中还未被解决的，且你未来希望去作为科研目标解决的最重要问题是什么？"),
    ("research_status_analysis", "上述问题现在全球科研界完成到了什么阶段以及有什么局限性？你觉得是否有新方法可以解决？"),
    ("research_impact", "上述问题如果解决了，对于技术进步、技术普及（e.g.成本大幅降低）、以及日常生活会带来什么影响？"),
    ("ai_society_impact", "你认为AI未来最能在什么环节或者场景影响/改变/优化/威胁人类社会？"),
    ("dissenting_view", "请陈述一个目前AI行业基本形成共识，但你不同意的观点，可以适当展开（选填）。"),
    ("family_info", "家庭情况"),
    ("education_experience", "教育经历"),
    ("practice_experience", "实践经历"),
    ("personal_statement_text", "个人简介"),
    ("student_activity_experience", "学生活动经历"),
    ("personal_statement_attachment", "个人陈述附件"),
    ("material_list_attachment", "材料清单附件"),
    ("supplementary_profile", "个人简介"),
]

REGISTERED_PORTAL_STUDENT_EXPORT_BASE_COLUMNS: list[tuple[str, str]] = [
    ("full_name", "姓名"),
    ("phone_number", "手机号"),
    ("email", "邮箱"),
    ("id_number", "证件号码"),
    ("portal_business_key", "门户业务编号"),
    ("candidate_no", "考生编号"),
    ("account_status", "账号状态"),
    ("application_form_status", "报名状态"),
    ("selected_plan_id", "招生计划ID"),
    ("selected_plan_name", "招生计划"),
    ("first_choice_advisor_name", "第一志愿导师"),
    ("first_choice_screening_score", "第一志愿导师评分"),
    ("first_choice_center_name", "第一志愿导师所属研究中心"),
    ("second_choice_advisor_name", "第二志愿导师"),
    ("second_choice_screening_score", "第二志愿导师评分"),
    ("second_choice_center_name", "第二志愿导师所属研究中心"),
    ("recruitment_application_business_key", "报名业务编号"),
    ("recruitment_application_id", "报名记录ID"),
    ("recruitment_application_status", "申请流转状态"),
    ("registered_at", "注册时间"),
    ("submitted_at", "报名提交时间"),
    ("full_name_pinyin", "姓名拼音"),
    ("profile_photo_url", "证件照附件"),
    ("id_card_collage_url", "身份证拼图附件"),
    ("gender", "性别"),
    ("birth_date", "出生日期"),
    ("ethnic_group", "民族"),
    ("native_place", "籍贯"),
    ("political_status", "政治面貌"),
    ("marital_status", "婚姻状况"),
    ("religious_belief", "宗教信仰"),
    ("id_type", "证件类型"),
    ("mailing_address", "通讯地址"),
    ("emergency_contact_name", "紧急联系人"),
    ("emergency_contact_phone", "紧急联系电话"),
    ("graduation_school", "毕业院校/就读学校"),
    ("highest_degree", "最高学历/教育阶段"),
    ("intended_field", "意向研究方向"),
    ("source_channel", "信息来源渠道"),
    ("source_channel_other", "其他信息来源"),
    ("english_level", "英语水平"),
    ("family_info", "家庭情况摘要"),
    ("education_experience", "教育经历摘要"),
    ("practice_experience", "实践经历摘要"),
    ("personal_profile", "个人简介"),
    ("recommendation_notes", "科研成果摘要"),
    ("personal_statement_text", "个人陈述摘要"),
    ("self_evaluation", "自我评价"),
    ("signed_agreement", "已阅读声明"),
]

REGISTERED_PORTAL_STUDENT_EXPORT_DYNAMIC_COLUMN_DEFINITIONS: list[tuple[str, str, list[tuple[str, str]]]] = [
    (
        "education",
        "教育经历",
        [
            ("sort_order", "排序"),
            ("stage", "教育阶段"),
            ("start_month", "开始年月"),
            ("end_month", "结束年月"),
            ("school_name", "学校名称"),
            ("major_name", "专业名称"),
            ("average_score", "平均成绩"),
            ("gpa", "绩点"),
            ("ranking", "成绩排名"),
            ("verifier_name", "证明人姓名"),
            ("verifier_phone", "证明人手机"),
            ("transcript_attachment", "成绩单附件"),
            ("degree_certificate_attachment", "学位证附件"),
            ("graduation_certificate_attachment", "毕业证附件"),
        ],
    ),
    (
        "practice",
        "实践经历",
        [
            ("sort_order", "排序"),
            ("start_month", "开始年月"),
            ("end_month", "结束年月"),
            ("organization_name", "单位名称"),
            ("position_name", "岗位名称"),
            ("responsibility", "职责"),
            ("verifier_name", "证明人姓名"),
            ("verifier_phone", "证明人手机"),
        ],
    ),
    (
        "english",
        "英语成绩",
        [
            ("sort_order", "排序"),
            ("exam_name", "考试名称"),
            ("score_text", "成绩"),
            ("certificate_attachment", "证书附件"),
        ],
    ),
    (
        "family",
        "家庭成员",
        [
            ("sort_order", "排序"),
            ("member_name", "姓名"),
            ("relation_type", "关系"),
            ("employer_name", "工作单位"),
            ("job_title", "职务"),
            ("contact_phone", "联系电话"),
        ],
    ),
    (
        "achievement",
        "科研成果",
        [
            ("sort_order", "排序"),
            ("achievement_type", "类型"),
            ("achievement_month", "成果时间"),
            ("paper_title", "论文标题"),
            ("journal_or_conference", "期刊或会议"),
            ("publish_or_index_month", "发表或检索时间"),
            ("author_order", "作者排序"),
            ("award_name", "奖项名称"),
            ("awarding_org", "授奖单位"),
            ("award_level", "奖项级别"),
            ("award_year", "获奖年份"),
            ("award_rank", "获奖排名"),
            ("description_text", "描述"),
            ("responsibility_text", "个人贡献"),
            ("award_certificate_attachment", "证明附件"),
        ],
    ),
]

REGISTERED_PORTAL_STUDENT_EXPORT_TAIL_COLUMNS: list[tuple[str, str]] = [
    ("personal_statement_resume_attachment", "个人陈述简历附件"),
    ("personal_statement_supporting_material_attachment", "个人陈述支撑材料附件"),
    ("personal_statement_growth_experience_text", "个人陈述成长经历"),
    ("personal_statement_why_apply_text", "个人陈述报考原因"),
    ("personal_statement_career_plan_text", "个人陈述未来规划"),
    ("personal_statement_research_interest_text", "个人陈述研究兴趣"),
    ("personal_statement_personal_statement_text", "个人陈述正文"),
    ("declaration_has_read", "声明已阅读"),
    ("declaration_text", "声明内容"),
    ("undergraduate_stage", "本科教育阶段"),
    ("undergraduate_start_month", "本科开始年月"),
    ("undergraduate_end_month", "本科结束年月"),
    ("undergraduate_school_name", "本科学校"),
    ("undergraduate_major_name", "本科专业"),
    ("undergraduate_average_score", "本科平均成绩"),
    ("undergraduate_gpa", "本科绩点"),
    ("undergraduate_ranking", "本科成绩排名"),
    ("undergraduate_verifier_name", "本科证明人姓名"),
    ("undergraduate_verifier_phone", "本科证明人手机"),
    ("undergraduate_transcript_attachment", "本科成绩单附件"),
    ("undergraduate_degree_certificate_attachment", "本科学位证附件"),
    ("undergraduate_graduation_certificate_attachment", "本科毕业证附件"),
]

ADVISOR_SCREENING_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("row_no", "序号"),
    ("candidate_no", "报名号"),
    ("business_key", "业务编号"),
    ("student_name", "姓名"),
    ("choice_name", "志愿"),
    ("advisor_screening_round", "初筛轮次"),
    ("first_choice", "第一志愿"),
    ("second_choice", "第二志愿"),
    ("first_choice_screening_score", "第一志愿分数"),
    ("second_choice_screening_score", "第二志愿分数"),
    ("application_status", "申请状态"),
    ("advisor_screening_status", "提交状态"),
    ("applied_at", "申请时间"),
    ("first_choice_screening_submitted_at", "第一志愿提交时间"),
    ("second_choice_screening_submitted_at", "第二志愿提交时间"),
]

CAMP_OFFER_TEMPLATE_COLUMNS: list[tuple[str, str]] = [
    ("candidate_no", "candidate_no"),
    ("plan_id", "plan_id"),
    ("is_agree", "is_agree"),
    ("reason", "reason"),
    ("is_sent_mail", "is_sent_mail"),
    ("student_offer_submitted_at", "student_offer_submitted_at"),
]


def _normalize_header(value: Any) -> str:
    return str(value or "").replace(" ", "").replace("\n", "").strip()


def _normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _sanitize_excel_cell_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    sanitized = ILLEGAL_CHARACTERS_RE.sub("", value)
    return "".join(char for char in sanitized if _is_valid_excel_xml_char(char))


def _is_valid_excel_xml_char(char: str) -> bool:
    code_point = ord(char)
    return (
        code_point in {0x9, 0xA, 0xD}
        or 0x20 <= code_point <= 0xD7FF
        or 0xE000 <= code_point <= 0xFFFD
        or 0x10000 <= code_point <= 0x10FFFF
    )


def _resolve_registered_portal_repeated_group_max_count(records: list[dict[str, Any]], prefix: str) -> int:
    max_count = 0
    prefix_text = f"{prefix}_"
    for record in records:
        for field_name in record:
            if not field_name.startswith(prefix_text):
                continue
            index_text, _, _ = field_name[len(prefix_text) :].partition("_")
            if index_text.isdigit():
                max_count = max(max_count, int(index_text))
    return max_count


def _resolve_registered_portal_student_export_columns(records: list[dict[str, Any]]) -> list[tuple[str, str]]:
    columns = list(REGISTERED_PORTAL_STUDENT_EXPORT_BASE_COLUMNS)
    for prefix, label_prefix, field_definitions in REGISTERED_PORTAL_STUDENT_EXPORT_DYNAMIC_COLUMN_DEFINITIONS:
        max_count = _resolve_registered_portal_repeated_group_max_count(records, prefix)
        for index in range(1, max_count + 1):
            for field_suffix, label_suffix in field_definitions:
                columns.append((f"{prefix}_{index}_{field_suffix}", f"{label_prefix}{index}{label_suffix}"))
    columns.extend(REGISTERED_PORTAL_STUDENT_EXPORT_TAIL_COLUMNS)
    return columns


def _build_excel_row(record: dict[str, Any], columns: list[tuple[str, str]]) -> list[Any]:
    return [_sanitize_excel_cell_value(record.get(field)) for field, _ in columns]


def parse_recruitment_template(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = cast(Worksheet, workbook.active)
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    normalized_headers = [_normalize_header(item) for item in header_row]
    expected_headers = [_normalize_header(label) for _, label in RECRUITMENT_TEMPLATE_COLUMNS]
    if normalized_headers[: len(expected_headers)] != expected_headers:
        raise ValueError("导入文件表头与资料审核名单模板不一致")

    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        values = list(row[: len(RECRUITMENT_TEMPLATE_COLUMNS)])
        if not any(value not in (None, "") for value in values):
            continue
        item = {
            field: _normalize_cell(value)
            for (field, _), value in zip(RECRUITMENT_TEMPLATE_COLUMNS, values, strict=False)
        }
        item["intended_field"] = item.get("first_choice") or item.get("second_choice") or "待分配方向"
        item["undergraduate_school"] = item.get("graduation_school") or item.get("undergraduate_school")
        result.append(item)
    return result


def parse_camp_offer_template(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = cast(Worksheet, workbook.active)
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    normalized_headers = [_normalize_header(item).lower() for item in header_row]
    if not normalized_headers:
        raise ValueError("导入文件缺少表头")

    header_index: dict[str, int] = {}
    for idx, header in enumerate(normalized_headers):
        if not header:
            continue
        if header in {"报名号", "candidate_no"}:
            header_index["candidate_no"] = idx
        elif header in {"plan_id", "planid", "计划id"}:
            header_index["plan_id"] = idx
        elif header in {"is_agree", "isagree", "是否同意"}:
            header_index["is_agree"] = idx
        elif header in {"reason", "原因", "reson"}:
            header_index["reason"] = idx
        elif header in {"is_sent_mail", "issentmail", "是否已发邮件"}:
            header_index["is_sent_mail"] = idx
        elif header in {"student_offer_submitted_at", "studentsubmittedat", "学生提交日期", "submited"}:
            header_index["student_offer_submitted_at"] = idx

    if "candidate_no" not in header_index:
        raise ValueError("导入文件必须包含 candidate_no 列")

    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        item: dict[str, Any] = {}
        for field in [
            "candidate_no",
            "plan_id",
            "is_agree",
            "reason",
            "is_sent_mail",
            "student_offer_submitted_at",
        ]:
            index = header_index.get(field)
            if index is None or index >= len(row):
                continue
            item[field] = _normalize_cell(row[index])
        result.append(item)
    return result


def build_recruitment_template(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "Worksheet"
    worksheet.append([label for _, label in RECRUITMENT_TEMPLATE_COLUMNS])
    for record in records:
        worksheet.append(_build_excel_row(record, RECRUITMENT_TEMPLATE_COLUMNS))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_registered_portal_students_template(records: list[dict[str, Any]]) -> bytes:
    columns = _resolve_registered_portal_student_export_columns(records)
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "注册学生"
    worksheet.append([label for _, label in columns])
    for record in records:
        worksheet.append(_build_excel_row(record, columns))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_advisor_screening_template(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = "导师初筛"
    worksheet.append([label for _, label in ADVISOR_SCREENING_EXPORT_COLUMNS])
    for index, record in enumerate(records, start=1):
        row_record = dict(record)
        row_record["row_no"] = index
        worksheet.append(_build_excel_row(row_record, ADVISOR_SCREENING_EXPORT_COLUMNS))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()