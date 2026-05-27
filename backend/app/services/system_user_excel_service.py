from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.services.recruitment_excel_service import _build_excel_row, _normalize_cell, _normalize_header


SYSTEM_USER_IMPORT_COLUMNS: list[tuple[str, str]] = [
    ("username", "账号"),
    ("full_name", "姓名"),
    ("role_name", "岗位角色"),
    ("role_code", "角色编码"),
    ("department_name", "部门"),
    ("email", "邮箱"),
    ("phone_number", "电话"),
    ("account_status", "账号状态"),
    ("introduction", "用户介绍"),
    ("password", "登录密码"),
]

LEGACY_SYSTEM_USER_IMPORT_COLUMNS: list[tuple[str, str]] = [
    ("id", "用户ID"),
    *SYSTEM_USER_IMPORT_COLUMNS[:8],
    ("last_login_at", "最近登录"),
    *SYSTEM_USER_IMPORT_COLUMNS[8:],
]


def parse_system_user_import_template(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    expected_headers = [_normalize_header(label) for _, label in SYSTEM_USER_IMPORT_COLUMNS]
    legacy_headers = [_normalize_header(label) for _, label in LEGACY_SYSTEM_USER_IMPORT_COLUMNS]
    normalized_header_row = [_normalize_header(value) for value in rows[0]]
    if normalized_header_row[: len(expected_headers)] == expected_headers:
        active_columns = SYSTEM_USER_IMPORT_COLUMNS
    elif normalized_header_row[: len(legacy_headers)] == legacy_headers:
        active_columns = LEGACY_SYSTEM_USER_IMPORT_COLUMNS
    else:
        raise ValueError("导入文件表头与系统用户导入模板不一致")

    result: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row[: len(active_columns)])
        if not any(value not in (None, "") for value in values):
            continue
        item = {
            field: _normalize_cell(value)
            for (field, _), value in zip(active_columns, values, strict=False)
        }
        item.pop("id", None)
        item.pop("last_login_at", None)
        item["row_number"] = row_number
        result.append(item)
    return result


def build_system_user_import_template(records: list[dict[str, Any]] | None = None) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "系统用户导入模板"
    worksheet.append([label for _, label in SYSTEM_USER_IMPORT_COLUMNS])
    for record in records or []:
        worksheet.append(_build_excel_row(record, SYSTEM_USER_IMPORT_COLUMNS))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()