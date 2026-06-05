from __future__ import annotations

from collections import Counter
from io import BytesIO
import sys
from pathlib import Path

from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))


from app.services.management_service import store


def _normalize_text(value: object) -> str:
    return str(value or "").strip()


def _extract_excel_counters(content: bytes) -> tuple[Counter[str], Counter[str]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return Counter(), Counter()

    headers = [_normalize_text(item) for item in rows[0]]
    first_choice_index = headers.index("志愿1导师") if "志愿1导师" in headers else None
    second_choice_index = headers.index("志愿2导师") if "志愿2导师" in headers else None

    first_counter: Counter[str] = Counter()
    second_counter: Counter[str] = Counter()

    for row in rows[1:]:
        values = list(row)
        if first_choice_index is not None and first_choice_index < len(values):
            advisor_name = _normalize_text(values[first_choice_index])
            if advisor_name:
                first_counter[advisor_name] += 1
        if second_choice_index is not None and second_choice_index < len(values):
            advisor_name = _normalize_text(values[second_choice_index])
            if advisor_name:
                second_counter[advisor_name] += 1

    return first_counter, second_counter


def _bucket_for_dashboard(counter: Counter[str]) -> list[tuple[str, int]]:
    ordered = sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    top10 = ordered[:10]
    remaining_total = sum(count for _, count in ordered[10:])
    bucket = list(top10)
    if remaining_total > 0:
        bucket.append(("其他导师", remaining_total))
    return bucket


def _dashboard_counter(choice_items: object) -> Counter[str]:
    counter: Counter[str] = Counter()
    for item in choice_items:
        advisor_name = _normalize_text(getattr(item, "advisor_name", None))
        count = int(getattr(item, "student_count", 0) or 0)
        if advisor_name:
            counter[advisor_name] = count
    return counter


def _print_counter(label: str, counter: Counter[str]) -> None:
    print(f"[{label}] total={sum(counter.values())}")
    for advisor_name, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])):
        print(f"  - {advisor_name}: {count}")


def main() -> int:
    print("[INFO] 正在导出注册学生全量数据（Excel口径）...")
    excel_content = store.export_registered_portal_students([])

    excel_first, excel_second = _extract_excel_counters(excel_content)
    excel_first_bucket = Counter(dict(_bucket_for_dashboard(excel_first)))
    excel_second_bucket = Counter(dict(_bucket_for_dashboard(excel_second)))

    print("[INFO] 正在读取 Dashboard 志愿统计...")
    dashboard = store.get_dashboard_recruitment_advisor_choice_distribution()
    dashboard_map = {choice.choice_round: _dashboard_counter(choice.items) for choice in dashboard.choices}
    dashboard_first = dashboard_map.get("first_choice", Counter())
    dashboard_second = dashboard_map.get("second_choice", Counter())

    first_all_keys = sorted(set(excel_first_bucket.keys()) | set(dashboard_first.keys()))
    second_all_keys = sorted(set(excel_second_bucket.keys()) | set(dashboard_second.keys()))

    first_diff = {
        key: (excel_first_bucket.get(key, 0), dashboard_first.get(key, 0))
        for key in first_all_keys
        if excel_first_bucket.get(key, 0) != dashboard_first.get(key, 0)
    }
    second_diff = {
        key: (excel_second_bucket.get(key, 0), dashboard_second.get(key, 0))
        for key in second_all_keys
        if excel_second_bucket.get(key, 0) != dashboard_second.get(key, 0)
    }

    print("\n[INFO] Excel口径（按Dashboard Top10+其他聚合后）")
    _print_counter("first_choice", excel_first_bucket)
    _print_counter("second_choice", excel_second_bucket)

    print("\n[INFO] Dashboard口径")
    _print_counter("first_choice", dashboard_first)
    _print_counter("second_choice", dashboard_second)

    if not first_diff and not second_diff:
        print("\n[PASS] 注册学生导出与 Dashboard 志愿统计一致。")
        return 0

    print("\n[FAIL] 发现不一致：")
    if first_diff:
        print("  [first_choice]")
        for key, (excel_count, dashboard_count) in first_diff.items():
            print(f"    - {key}: excel={excel_count}, dashboard={dashboard_count}")
    if second_diff:
        print("  [second_choice]")
        for key, (excel_count, dashboard_count) in second_diff.items():
            print(f"    - {key}: excel={excel_count}, dashboard={dashboard_count}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
