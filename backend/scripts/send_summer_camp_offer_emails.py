"""Send summer-camp offer emails to recruitment applicants.

Usage:
    python -m backend.scripts.send_summer_camp_offer_emails \
        --candidate-no SH20270001,SH20270002 [--sent-mail yes|no] [--choice first|second]

The script reads the SMTP configuration from ``.env`` (same as the FastAPI
app via ``app.core.config.settings``), looks up each candidate_no in
``dtlms_recruitment_applications.email``, converts ``frontend/public/images/offer.md``
into a styled HTML e-mail body, and dispatches it via the same SMTP service
used elsewhere in the project (``app.services.email_service``).

``--candidate-no`` accepts either a comma-separated list of candidate
numbers (e.g. ``SH20270001,SH20270002``) or the path to an ``.xlsx`` /
``.xls`` file whose ``candidate_no`` column contains the same numbers.
When an Excel file is given, the first row is treated as a header and the
values are read from the ``candidate_no`` column (case-insensitive,
whitespace trimmed, blanks skipped).

Two operating modes are supported via ``--sent-mail``:

* ``yes`` (default): real delivery. One e-mail is sent to each candidate
  address that is found in the database. Missing candidates are reported
  but do not abort the run.
* ``no``: dry run. By default **no e-mail is dispatched**: the script
  looks up every ``candidate_no`` in ``dtlms_recruitment_applications``
  and prints a ``candidate_no -> email`` table (plus a summary of
  missing/blank rows) so reviewers can verify the recipient list before
  a real send. If ``--mail-list`` is provided (semicolon-separated
  addresses), the script additionally dispatches a single test e-mail
  to **each** address in the list. The ``--test-recipient`` flag is
  ignored in this mode.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import smtplib
import ssl
import sys
from dataclasses import dataclass
from email.message import EmailMessage
from html import escape
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - dependency is in requirements.txt
    load_workbook = None  # type: ignore

import psycopg


# --------------------------------------------------------------------------- #
# Paths & project wiring
# --------------------------------------------------------------------------- #

BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = BACKEND_DIR.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.config import settings  # noqa: E402  (sys.path tweak above)


SCRIPT_DIR = Path(__file__).resolve().parent
OFFER_FIRST_CHOICE_MARKDOWN_PATH = SCRIPT_DIR / "offer.md"
OFFER_SECOND_CHOICE_MARKDOWN_PATH = SCRIPT_DIR / "offer2.md"


def _resolve_offer_markdown_path(choice: str = "first") -> Path:
    """Locate the offer Markdown template, defaulting to the script directory."""

    if str(choice or "").strip().lower() == "second":
        return OFFER_SECOND_CHOICE_MARKDOWN_PATH
    return OFFER_FIRST_CHOICE_MARKDOWN_PATH


OFFER_MARKDOWN_PATH = _resolve_offer_markdown_path()

TEST_RECIPIENT_EMAIL = "lk139@126.com"

EMAIL_SUBJECT_PREFIX = "【上海人工智能实验室】2026年夏令营入营通知"


# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #

logger = logging.getLogger("send_summer_camp_offer_emails")


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #


@dataclass(frozen=True)
class Candidate:
    """A single row resolved from ``dtlms_recruitment_applications``."""

    candidate_no: str
    email: str
    student_name: str | None = None


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Send the summer-camp offer e-mail to the listed candidate numbers. "
            "Reads the SMTP configuration from the project's .env file."
        ),
        epilog=(
            "Example: python -m backend.scripts.send_summer_camp_offer_emails "
            "--candidate-no SH20270001,SH20270002 --sent-mail yes"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--candidate-no",
        required=True,
        help=(
            "Comma-separated list of recruitment candidate numbers, or a path "
            "to an .xlsx/.xls file whose 'candidate_no' column contains the "
            "same numbers. Matches dtlms_recruitment_applications.candidate_no."
        ),
    )
    parser.add_argument(
        "--sent-mail",
        choices=("yes", "no"),
        default="yes",
        help=(
            "yes = send the offer to the candidate e-mail addresses; "
            "no = dry run: do NOT send, only print each candidate_no and "
            "its e-mail address resolved from dtlms_recruitment_applications. "
            "Defaults to yes."
        ),
    )
    parser.add_argument(
        "--test-recipient",
        default=TEST_RECIPIENT_EMAIL,
        help=(
            "保留参数以兼容旧调用;--sent-mail=no 时不再使用。 "
            f"Defaults to {TEST_RECIPIENT_EMAIL}."
        ),
    )
    parser.add_argument(
        "--mail-list",
        default="",
        help=(
            "Semicolon-separated list of e-mail addresses. Used together "
            "with --sent-mail=no: when provided, the script sends one "
            "test e-mail to each address in this list (in addition to "
            "the recipient preview). When omitted, --sent-mail=no is a "
            "pure dry run and dispatches no e-mails."
        ),
    )
    parser.add_argument(
        "--offer-md",
        type=Path,
        default=None,
        help=(
            "Override the path of the offer Markdown template. By default the "
            "script picks <script_dir>/offer.md for the first-choice "
            "template and <script_dir>/offer2.md for the second-choice "
            "template, depending on --choice."
        ),
    )
    parser.add_argument(
        "--choice",
        choices=("first", "second"),
        default="first",
        help=(
            "Which choice the notification is for. Controls which Markdown "
            "template is used when --offer-md is not given. Defaults to \"first\" (uses offer.md in the script directory)."
        ),
    )
    parser.add_argument(
        "--simulate-recipient",
        default="",
        help=(
            "If set, override the per-candidate recipient with this single "
            "address for every candidate in --candidate-no. Used to "
            "send notification mails to a test inbox without reaching the "
            "real applicants. The database is still updated to mark "
            "is_sent_mail on success so QA flows behave like a real run."
        ),
    )
    parser.add_argument(
        "--result-json",
        type=Path,
        default=None,
        help=(
            "If set, write a JSON file with per-candidate results to this "
            "path before the script exits. Schema: {success_count, "
            "failure_count, results: [{candidate_no, email, status, error}]}."
        ),
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug-level logging.",
    )
    return parser


# --------------------------------------------------------------------------- #
# Database helpers
# --------------------------------------------------------------------------- #


def _conninfo() -> str:
    return (
        f"host={settings.postgres_host} "
        f"port={settings.postgres_port} "
        f"dbname={settings.postgres_db} "
        f"user={settings.postgres_user} "
        f"password={settings.postgres_password}"
    )


def _normalize(value: object) -> str:
    return str(value or "").strip()


def fetch_candidates(candidate_nos: list[str]) -> list[Candidate]:
    """Look up the e-mail addresses for the given candidate numbers."""

    if not candidate_nos:
        return []

    sql = """
        SELECT
            ra.candidate_no,
            ra.email,
            ra.student_name
        FROM dtlms_recruitment_applications ra
        WHERE ra.candidate_no = ANY(%s)
    """

    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cursor:
            cursor.execute(sql, (candidate_nos,))
            rows = cursor.fetchall()

    return [
        Candidate(
            candidate_no=_normalize(row[0]),
            email=_normalize(row[1]),
            student_name=(_normalize(row[2]) or None),
        )
        for row in rows
        if _normalize(row[1])
    ]


def mark_offer_sent_mail(candidate_nos: list[str]) -> int:
    """Set `dtlms_plan_offer.is_sent_mail = TRUE` for the given candidate_nos.

    Returns the number of rows updated.
    """

    cleaned = [item for item in (_normalize(no) for no in candidate_nos) if item]
    if not cleaned:
        return 0

    sql = (
        "UPDATE dtlms_plan_offer SET is_sent_mail = TRUE, updated_at = CURRENT_TIMESTAMP "
        "WHERE candidate_no = ANY(%s)"
    )
    with psycopg.connect(_conninfo()) as conn:
        with conn.cursor() as cursor:
            cursor.execute(sql, (cleaned,))
            updated = cursor.rowcount or 0
        conn.commit()
    return int(updated)


# --------------------------------------------------------------------------- #
# Markdown -> HTML (dependency-free)
# --------------------------------------------------------------------------- #
#
# offer.md uses: ATX headings, paragraphs, **bold**, ordered/unordered lists,
# pipe tables, [text](url) links, `---` horizontal rules and one blockquote.
# The renderer below supports exactly those constructs - no images, no inline
# HTML, no nested lists. That is sufficient for the current template and
# avoids pulling in a third-party dependency.

_INLINE_BOLD_RE = re.compile(r"\*\*(.+?)\*\*")
_INLINE_CODE_RE = re.compile(r"`([^`]+)`")
_INLINE_LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
_ESCAPED_PUNCT_RE = re.compile(r"\\([\\.*_\[\]()#+\-!`>~])")


def _apply_inline(text: str) -> str:
    """Apply inline Markdown transforms and HTML-escape the rest of the line."""

    placeholders: list[str] = []

    def _stash(html: str) -> str:
        placeholders.append(html)
        return f"\x00{len(placeholders) - 1}\x00"

    def _unescape_punct(value: str) -> str:
        return _ESCAPED_PUNCT_RE.sub(r"\1", value)

    def _unescape(match: re.Match[str]) -> str:
        return _stash(f"<strong>{escape(_unescape_punct(match.group(1)))}</strong>")

    def _code(match: re.Match[str]) -> str:
        return _stash(f"<code>{escape(match.group(1))}</code>")

    def _link(match: re.Match[str]) -> str:
        label, url = match.group(1), match.group(2)
        return _stash(
            f'<a href="{escape(url, quote=True)}">{escape(_unescape_punct(label))}</a>'
        )

    text = _INLINE_BOLD_RE.sub(_unescape, text)
    text = _INLINE_CODE_RE.sub(_code, text)
    text = _INLINE_LINK_RE.sub(_link, text)
    text = _ESCAPED_PUNCT_RE.sub(r"\1", text)
    text = escape(text, quote=False)

    def _restore(match: re.Match[str]) -> str:
        return placeholders[int(match.group(1))]

    return re.sub(r"\x00(\d+)\x00", _restore, text)

def _split_table_row(line: str) -> list[str]:
    stripped = line.strip().strip("|")
    return [cell.strip() for cell in stripped.split("|")]


def _is_table_separator(line: str) -> bool:
    stripped = line.strip().strip("|")
    if not stripped:
        return False
    return all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in stripped.split("|"))


def render_markdown_to_html(markdown_text: str) -> str:
    """Render the offer.md template to a self-contained HTML document."""

    lines = markdown_text.replace("\r\n", "\n").split("\n")

    html_blocks: list[str] = []
    i = 0
    n = len(lines)

    def flush_paragraph(start: int) -> None:
        chunk = lines[start:i]
        text = " ".join(line.strip() for line in chunk if line.strip())
        if text:
            html_blocks.append(f"<p>{_apply_inline(text)}</p>")

    while i < n:
        line = lines[i]
        stripped = line.strip()

        # Blank line separates blocks
        if not stripped:
            i += 1
            continue

        # Horizontal rule
        if re.fullmatch(r"-{3,}|\*{3,}|_{3,}", stripped):
            html_blocks.append("<hr>")
            i += 1
            continue

        # Heading
        heading_match = re.match(r"^(#{1,6})\s+(.*)$", stripped)
        if heading_match:
            level = len(heading_match.group(1))
            html_blocks.append(
                f"<h{level}>{_apply_inline(heading_match.group(2))}</h{level}>"
            )
            i += 1
            continue

        # Blockquote
        if stripped.startswith(">"):
            quote_lines: list[str] = []
            while i < n and lines[i].lstrip().startswith(">"):
                quote_lines.append(lines[i].lstrip()[1:].lstrip())
                i += 1
            inner = " ".join(line.strip() for line in quote_lines if line.strip())
            if inner:
                html_blocks.append(f"<blockquote><p>{_apply_inline(inner)}</p></blockquote>")
            continue

        # Table
        if "|" in stripped and i + 1 < n and _is_table_separator(lines[i + 1]):
            header_cells = _split_table_row(stripped)
            i += 2  # skip header + separator
            body_rows: list[list[str]] = []
            while i < n and "|" in lines[i] and lines[i].strip():
                body_rows.append(_split_table_row(lines[i]))
                i += 1
            header_html = "".join(f"<th>{_apply_inline(cell)}</th>" for cell in header_cells)
            body_html = "".join(
                "<tr>"
                + "".join(f"<td>{_apply_inline(cell)}</td>" for cell in row)
                + "</tr>"
                for row in body_rows
            )
            html_blocks.append(
                "<table>"
                f"<thead><tr>{header_html}</tr></thead>"
                f"<tbody>{body_html}</tbody>"
                "</table>"
            )
            continue

        # Unordered list
        if re.match(r"^[-*+]\s+", stripped):
            items: list[str] = []
            while i < n:
                match = re.match(r"^\s*[-*+]\s+(.*)$", lines[i])
                if not match:
                    break
                items.append(match.group(1).strip())
                i += 1
            html_blocks.append(
                "<ul>"
                + "".join(f"<li>{_apply_inline(item)}</li>" for item in items)
                + "</ul>"
            )
            continue

        # Ordered list
        if re.match(r"^\d+\.\s+", stripped):
            items = []
            while i < n:
                match = re.match(r"^\s*\d+\.\s+(.*)$", lines[i])
                if not match:
                    break
                items.append(match.group(1).strip())
                i += 1
            html_blocks.append(
                "<ol>"
                + "".join(f"<li>{_apply_inline(item)}</li>" for item in items)
                + "</ol>"
            )
            continue

        # Paragraph: collect contiguous non-blank, non-block lines
        para_start = i
        while i < n:
            peek = lines[i].strip()
            if not peek:
                break
            if re.match(r"^(#{1,6})\s+", peek):
                break
            if re.fullmatch(r"-{3,}|\*{3,}|_{3,}", peek):
                break
            if peek.startswith(">"):
                break
            if "|" in peek and i + 1 < n and _is_table_separator(lines[i + 1]):
                break
            if re.match(r"^\s*[-*+]\s+", lines[i]) or re.match(
                r"^\s*\d+\.\s+", lines[i]
            ):
                break
            i += 1
        flush_paragraph(para_start)

    body = "\n".join(html_blocks)
    return _wrap_html_document(body)


def _wrap_html_document(body_html: str) -> str:
    style = (
        "body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',"
        "'PingFang SC','Hiragino Sans GB','Microsoft YaHei',sans-serif;"
        "color:#1e293b;line-height:1.7;max-width:760px;margin:0 auto;"
        "padding:24px;background:#ffffff;}"
        "h1,h2,h3,h4{color:#0f172a;line-height:1.4;margin-top:1.6em;margin-bottom:0.6em;}"
        "h1{font-size:24px;}"
        "h2{font-size:20px;border-bottom:1px solid #e2e8f0;padding-bottom:6px;}"
        "h3{font-size:17px;}"
        "p{margin:0 0 1em;}"
        "ul,ol{margin:0 0 1em 1.4em;padding:0;}"
        "li{margin-bottom:0.4em;}"
        "strong{color:#0f172a;}"
        "a{color:#1d4ed8;text-decoration:none;}"
        "a:hover{text-decoration:underline;}"
        "hr{border:none;border-top:1px solid #e2e8f0;margin:24px 0;}"
        "blockquote{border-left:4px solid #fcd34d;background:#fefce8;"
        "margin:0 0 1em;padding:10px 16px;color:#475569;border-radius:4px;}"
        "table{border-collapse:collapse;width:100%;margin:0 0 1em;font-size:14px;}"
        "th,td{border:1px solid #e2e8f0;padding:8px 12px;text-align:left;vertical-align:top;}"
        "th{background:#f1f5f9;color:#0f172a;}"
        "code{background:#f1f5f9;padding:1px 6px;border-radius:4px;font-size:13px;}"
    )
    return (
        "<!DOCTYPE html>"
        '<html lang="zh-CN">'
        "<head>"
        '<meta charset="utf-8">'
        "<title>" + escape(EMAIL_SUBJECT_PREFIX) + "</title>"
        f"<style>{style}</style>"
        "</head>"
        f"<body>{body_html}</body>"
        "</html>"
    )


# --------------------------------------------------------------------------- #
# SMTP helpers
# --------------------------------------------------------------------------- #


def _format_from_address() -> str:
    from_name = settings.smtp_from_name.strip()
    from_email = settings.smtp_from_email.strip()
    if from_name:
        return f"{from_name} <{from_email}>"
    return from_email


def _build_message(*, to_email: str, subject: str, html_body: str) -> EmailMessage:
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = _format_from_address()
    message["To"] = to_email
    # Provide a plain-text fallback by stripping tags for clients that refuse HTML.
    text_fallback = re.sub(r"<[^>]+>", "", html_body)
    text_fallback = re.sub(r"\s+\n", "\n", text_fallback).strip()
    message.set_content(text_fallback)
    message.add_alternative(html_body, subtype="html")
    return message


def _send_via_smtp(message: EmailMessage) -> None:
    timeout = settings.smtp_timeout_seconds
    if settings.smtp_use_ssl:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(
            settings.smtp_host, settings.smtp_port, timeout=timeout, context=context
        ) as server:
            _login_if_needed(server)
            server.send_message(message)
        return

    with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=timeout) as server:
        if settings.smtp_use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        _login_if_needed(server)
        server.send_message(message)


def _login_if_needed(server: smtplib.SMTP) -> None:
    username = settings.smtp_username.strip()
    if username:
        server.login(username, settings.smtp_password)


def _smtp_configured() -> bool:
    return bool(
        True
        #settings.smtp_enabled
        and settings.smtp_host
        and settings.smtp_port
        and settings.smtp_from_email
    )


# --------------------------------------------------------------------------- #
# Main flow
# --------------------------------------------------------------------------- #


def _print_recipient_preview(candidate_nos: list[str]) -> int:
    """Dry-run mode: print ``candidate_no -> email`` without sending.

    Returns 0 when at least one row was resolved, otherwise 1. The
    caller is expected to exit with this return code.
    """

    candidates = fetch_candidates(candidate_nos)
    resolved = {c.candidate_no: c.email for c in candidates}
    missing = [no for no in candidate_nos if no not in resolved]
    blank_email = [c.candidate_no for c in candidates if not c.email]

    candidate_width = max(
        [len("candidate_no")]
        + [len(no) for no in candidate_nos]
        + [len(no) for no in missing + blank_email]
    )
    candidate_width = max(candidate_width, 12)

    header = f"{'candidate_no'.ljust(candidate_width)}  email"
    sep = f"{'-' * candidate_width}  {'-' * 40}"

    logger.info("测试模式(dry-run): 不会发送邮件,以下为 candidate_no -> email 映射:")
    print(header)
    print(sep)

    for no in candidate_nos:
        email = resolved.get(no, "")
        if not email:
            continue
        print(f"{no.ljust(candidate_width)}  {email}")

    if missing:
        print(sep)
        for no in missing:
            print(f"{no.ljust(candidate_width)}  <NOT FOUND IN dtlms_recruitment_applications>")
    if blank_email:
        for no in blank_email:
            print(f"{no.ljust(candidate_width)}  <EMAIL IS BLANK>")

    print(sep)
    print(
        f"汇总: 共 {len(candidate_nos)} 个报名号, "
        f"已解析 {len(candidates) - len(blank_email)} 个有效邮箱, "
        f"未找到 {len(missing)} 个, "
        f"邮箱为空 {len(blank_email)} 个"
    )
    return 0 if (candidates and not missing) else 1


def _split_candidate_nos(raw: str) -> list[str]:
    return [item.strip() for item in raw.split(",") if item.strip()]


def _split_mail_list(raw: str) -> list[str]:
    """Parse a semicolon-separated mail list, preserving order, deduping."""

    seen: set[str] = set()
    result: list[str] = []
    for chunk in (raw or "").split(";"):
        addr = chunk.strip()
        if not addr or addr in seen:
            continue
        seen.add(addr)
        result.append(addr)
    return result


_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}


def _is_excel_path(value: str) -> bool:
    if not value:
        return False
    # Disambiguate: only treat the value as a file when it explicitly looks
    # like a path (contains a separator) and ends with an Excel suffix. This
    # way a raw list like 'SH20270001,SH20270002' is never misread as a file.
    if "," in value:
        return False
    try:
        suffix = Path(value).suffix.lower()
    except ValueError:
        return False
    return suffix in _EXCEL_SUFFIXES


def _load_candidate_nos_from_excel(path: Path) -> list[str]:
    """Read the ``candidate_no`` column from an Excel workbook."""

    load_workbook_fn = load_workbook
    if load_workbook_fn is None:
        raise RuntimeError(
            "openpyxl is required to read Excel candidate lists; "
            "install it via `pip install openpyxl`."
        )
    if not path.exists():
        raise FileNotFoundError(f"找不到候选号 Excel 文件：{path}")

    workbook = load_workbook_fn(filename=str(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            return []
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []

        target_column: int | None = None
        for index, cell in enumerate(header):
            if cell is None:
                continue
            if str(cell).strip().lower() == "candidate_no":
                target_column = index
                break
        if target_column is None:
            raise ValueError(
                f"Excel {path.name} 中找不到 candidate_no 列（首行表头）"
            )

        result: list[str] = []
        for row in rows:
            if not row or target_column >= len(row):
                continue
            value = row[target_column]
            if value is None:
                continue
            text = str(value).strip()
            if text and text not in result:
                result.append(text)
        return result
    finally:
        workbook.close()


def _resolve_candidate_nos(raw: str) -> list[str]:
    """Resolve ``--candidate-no`` into a deduplicated, ordered list."""

    value = (raw or "").strip()
    if not value:
        return []
    if _is_excel_path(value):
        return _load_candidate_nos_from_excel(Path(value))
    return _split_candidate_nos(value)


def _build_subject(candidate_nos: Iterable[str]) -> str:
    numbers = list(candidate_nos)
    if not numbers:
        return EMAIL_SUBJECT_PREFIX
    if len(numbers) == 1:
        return f"{EMAIL_SUBJECT_PREFIX}（{numbers[0]}）"
    return f"{EMAIL_SUBJECT_PREFIX}（{len(numbers)} 人）"


def send_to_recipient(
    *,
    recipient: str,
    subject: str,
    html_body: str,
    recipient_label: str,
) -> bool:
    if not _smtp_configured():
        logger.error(
            "SMTP 未启用或配置不完整（请检查 .env 中的 SMTP_ENABLED / SMTP_HOST / "
            "SMTP_PORT / SMTP_FROM_EMAIL）。收件人 %s 未发送。",
            recipient_label,
        )
        return False

    message = _build_message(to_email=recipient, subject=subject, html_body=html_body)
    try:
        _send_via_smtp(message)
    except Exception as exc:  # noqa: BLE001
        logger.error("发送失败：收件人=%s 错误=%s", recipient_label, exc)
        return False
    logger.info("发送成功：收件人=%s 主题=%s", recipient_label, subject)
    return True


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    try:
        candidate_nos = _resolve_candidate_nos(args.candidate_no)
    except (FileNotFoundError, ValueError) as exc:
        logger.error("%s", exc)
        return 2
    if not candidate_nos:
        logger.error("--candidate-no 至少需要一个报名号")
        return 2
    logger.info("共解析到 %d 个报名号", len(candidate_nos))

    offer_md_path = args.offer_md or _resolve_offer_markdown_path(args.choice)
    if not offer_md_path.exists():
        logger.error("找不到 offer.md 模板：%s", offer_md_path)
        return 2
    logger.info("使用模板: %s (choice=%s)", offer_md_path, args.choice)

    markdown_text = offer_md_path.read_text(encoding="utf-8")
    html_body = render_markdown_to_html(markdown_text)

    simulate_recipient = str(getattr(args, "simulate_recipient", "") or "").strip()
    result_records: list[dict[str, str]] = []
    def _flush_results(success_nos: list[str], failures: list[dict[str, str]]) -> None:
        out_path = getattr(args, "result_json", None)
        if not out_path:
            return
        payload = {
            "success_count": len(success_nos),
            "failure_count": len(failures),
            "results": result_records,
        }
        try:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as exc:  # noqa: BLE001
            logger.error("写入结果 JSON 失败：%s 错误=%s", out_path, exc)

    if args.sent_mail == "no":
        # Always print the candidate_no -> email map first.
        preview_ok = _print_recipient_preview(candidate_nos) == 0
        mail_list = _split_mail_list(args.mail_list)
        if not mail_list:
            # Pure dry run.
            return 0 if preview_ok else 1
        # Send one test e-mail per address in --mail-list.
        # Subject is the bare formal subject - no test markers,
        # no candidate_no suffix, no count summary. Test e-mails must
        # be byte-identical to a real-delivery message.
        subject = EMAIL_SUBJECT_PREFIX
        logger.info(
            "测试模式(--mail-list): 向 %d 个邮箱发送测试邮件, 主题=%s",
            len(mail_list),
            subject,
        )
        sent = 0
        for recipient in mail_list:
            ok = send_to_recipient(
                recipient=recipient,
                subject=subject,
                html_body=html_body,
                recipient_label=recipient,
            )
            if ok:
                sent += 1
        logger.info("测试发送完成: 成功 %d/%d", sent, len(mail_list))
        # Both the preview and the dispatch need to succeed for exit 0.
        return 0 if (preview_ok and sent == len(mail_list)) else 1

    # Real delivery: one e-mail per candidate
    if simulate_recipient:
        logger.info("模拟发送: 收件人统一覆盖为 %s (真实邮箱不会收到邮件)", simulate_recipient)
        dispatch_targets = [
            Candidate(candidate_no=no, email=simulate_recipient, student_name=None)
            for no in candidate_nos
        ]
        candidates = dispatch_targets
        missing: list[str] = []
    else:
        candidates = fetch_candidates(candidate_nos)
        found_nos = {c.candidate_no for c in candidates}
        missing = [no for no in candidate_nos if no not in found_nos]

        if missing:
            logger.warning(
                "以下报名号在 dtlms_recruitment_applications 中未找到或邮箱为空：%s",
                ", ".join(missing),
            )

        if not candidates:
            logger.error("没有可发送的候选人，程序退出")
            failures: list[dict[str, str]] = [
                {"candidate_no": no, "email": "", "status": "missing", "error": "not_found"}
                for no in missing
            ]
            result_records.extend(failures)
            _flush_results([], failures)
            return 1

    success = 0
    success_nos: list[str] = []
    failures = []
    for candidate in candidates:
        subject = _build_subject([candidate.candidate_no])
        ok = send_to_recipient(
            recipient=candidate.email,
            subject=subject,
            html_body=html_body,
            recipient_label=f"{candidate.candidate_no}<{candidate.email}>",
        )
        if ok:
            success += 1
            success_nos.append(candidate.candidate_no)
            result_records.append({
                "candidate_no": candidate.candidate_no,
                "email": candidate.email,
                "status": "sent",
                "error": "",
            })
        else:
            failures.append({
                "candidate_no": candidate.candidate_no,
                "email": candidate.email,
                "status": "failed",
                "error": "send_failed",
            })
    for no in missing:
        failures.append({
            "candidate_no": no,
            "email": "",
            "status": "missing",
            "error": "not_found",
        })
    result_records.extend(failures)

    if success_nos:
        try:
            updated = mark_offer_sent_mail(success_nos)
            logger.info("已将 %d 条 dtlms_plan_offer.is_sent_mail 置为 TRUE", updated)
        except Exception as exc:  # noqa: BLE001
            logger.error("更新 is_sent_mail 失败：%s", exc)

    _flush_results(success_nos, failures)

    logger.info(
        "发送完成：成功 %d/%d，未找到 %d",
        success,
        len(candidates),
        len(missing),
    )
    return 0 if (success == len(candidates) and not failures) else 1


if __name__ == "__main__":
    raise SystemExit(main())

